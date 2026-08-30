# -*- coding: utf-8 -*-
"""
Historial de cambios -- reemplaza el 8_Historial_Cambios.xlsx que se editaba
a mano para forzar un valor distinto al Maestro/Patrón (hora de entrada/
salida, canal, supervisor, zona, refrigerio) para un DNI en un rango de
fechas. La lectura ya vive en Postgres desde Fase 6 (ver
pipeline/historial_cambios.py::cargar_historial()); lo que faltaba era la
pantalla para poder cargar/editar/borrar esas filas sin tocar la base a
mano -- Davor, 2026-08-25 ("teniamos un excel historial_cambios... eso no
hemos migrado a la plataforma").

Carga masiva por Excel (2026-08-25, Davor: "si van haber muchos cambios,
podria tambien subir un excel") -- mismo patrón que cargas.py (Cargar
Headcount): plantilla descargable + parseo con reporte de errores fila por
fila, sin frenar toda la carga por una fila mal completada.
"""
import datetime as dt
import io
from functools import wraps

import openpyxl
import pandas as pd
from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl.styles import Font

from dimension_models import HistorialCambio, Persona, get_session
from permisos import requiere_pagina
from scoping import condicion_scope

bp = Blueprint("historial", __name__, url_prefix="/historial")

# Mismo vocabulario que documentaba MAC/historial_cambios.py -- son los
# únicos "Campo" que algún módulo del pipeline realmente lee vía
# valor_efectivo() (ver motor_clasificacion.py, reporte_diario_9am.py,
# horas_semanales.py, export_sharepoint.py, cobertura.py). "Rol" estaba
# documentado en el Excel original pero ningún módulo lo consulta -- no se
# incluye acá para no ofrecer una opción que no hace nada en silencio.
CAMPOS_VALIDOS = [
    "Hora entrada programada", "Hora salida programada",
    "Canal", "Canal del día", "Supervisor", "Zona", "Refrigerio",
]
DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

# Un solo lugar para el formato esperado de "Valor nuevo" por Campo -- se
# usa tanto para la ayuda en pantalla (historial.html) como para la hoja
# "Instrucciones" de la plantilla de carga masiva, para que no queden
# desincronizadas entre sí.
FORMATO_POR_CAMPO = {
    "Hora entrada programada": "Formato HH:MM (ej. 09:00).",
    "Hora salida programada": "Formato HH:MM (ej. 17:30).",
    "Canal": "El texto exacto del canal (ej. Tradicional, Farmacia, Autoservicio).",
    "Canal del día": "El texto exacto del canal (ej. Tradicional, Farmacia, Autoservicio).",
    "Supervisor": "El nombre completo tal como figura en el Maestro.",
    "Zona": "El texto de zona/ruta tal como figura en el Maestro.",
    "Refrigerio": '"Con refrigerio", "Medio refrigerio" o "Sin refrigerio".',
}

COLUMNAS_HISTORIAL_ESPERADAS = [
    "DNI", "Campo", "Valor nuevo", "Fecha desde",
    "Fecha hasta (opcional)", "Día de la semana (opcional)", "Comentario (opcional)",
]


def _analista_requerido(f):
    """Mismo criterio de acceso que "Cargar Headcount" (cargas.py) -- un
    override de Historial es tan sensible como subir Headcount nuevo, así
    que se abre a analista+admin, no solo admin."""
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if current_user.rol not in ("analista", "admin"):
            flash("Esta sección es solo para analistas.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapper


def _persona_en_scope(session, dni):
    """Busca la Persona y confirma que está dentro del scope de
    current_user (mismo criterio que reportes.py::ficha()) -- evita que un
    analista de un cliente cree/edite/borre un cambio sobre gente de otro
    cliente escribiendo el DNI a mano en el formulario o en el Excel."""
    query = session.query(Persona).filter(Persona.dni == dni)
    cond_scope = condicion_scope(Persona, current_user)
    if cond_scope is not None:
        query = query.filter(cond_scope)
    return query.first()


def _leer_datos_comunes(form):
    """Valida los campos que son iguales sea cual sea la cantidad de días
    elegidos (crear() los repite una vez por día; editar() los usa tal
    cual). El Día de la semana se valida aparte porque crear() acepta
    varios y editar() uno solo."""
    dni = form.get("dni", "").strip()
    campo = form.get("campo", "").strip()
    valor_nuevo = form.get("valor_nuevo", "").strip()
    fecha_desde = form.get("fecha_desde", "").strip()
    fecha_hasta = form.get("fecha_hasta", "").strip() or None
    comentario = form.get("comentario", "").strip() or None

    if not dni or campo not in CAMPOS_VALIDOS or not valor_nuevo or not fecha_desde:
        return None, "Completá DNI, Campo, Valor nuevo y Fecha desde."
    if campo in CAMPOS_HORA:
        valor_normalizado = _normalizar_hora(valor_nuevo)
        if valor_normalizado is None:
            return None, f'"{valor_nuevo}" no es una hora válida -- usá el formato HH:MM (ej. 09:00).'
        valor_nuevo = valor_normalizado
    try:
        fecha_desde = dt.date.fromisoformat(fecha_desde)
        fecha_hasta = dt.date.fromisoformat(fecha_hasta) if fecha_hasta else None
    except ValueError:
        return None, "Fecha inválida."
    if fecha_hasta and fecha_hasta < fecha_desde:
        return None, "La fecha hasta no puede ser anterior a la fecha desde."
    return {
        "dni": dni, "campo": campo, "valor_nuevo": valor_nuevo,
        "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta, "comentario": comentario,
    }, None


def _fecha(valor):
    if pd.isna(valor):
        return None
    if isinstance(valor, dt.date) and not isinstance(valor, dt.datetime):
        return valor
    return pd.Timestamp(valor).date()


def _texto(valor):
    if pd.isna(valor):
        return None
    return str(valor).strip() or None


CAMPOS_HORA = ("Hora entrada programada", "Hora salida programada")


def _normalizar_hora(valor):
    """"Hora entrada/salida programada" termina en pd.to_timedelta() dentro
    del motor de clasificación -- pandas 3.x exige "hh:mm:ss", pero la ayuda
    en pantalla le pide al usuario "HH:MM" (ej. "09:00"). Sin esto, un valor
    así tipeado se guardaba tal cual y recién tumbaba el motor de
    clasificación HORAS después, en producción, sin ningún aviso al momento
    de cargarlo (apagón real 2026-08-25/26, dos veces con este mismo
    origen). Se completa con ":00" si falta y se valida que sea una hora
    real -- devuelve None si no se puede salvar, para frenar la carga con
    un error claro en vez de guardar basura."""
    texto = valor.strip()
    if texto.count(":") == 1:
        texto = texto + ":00"
    try:
        dt.time.fromisoformat(texto)
    except ValueError:
        return None
    return texto


def _plantilla_historial_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de cambios"
    ws.cell(row=1, column=1, value="Historial de cambios -- carga masiva").font = Font(bold=True, size=13)
    ws.cell(
        row=2, column=1,
        value='Plantilla en blanco -- completá una fila por cambio y subila en "Historial de cambios". '
              'Ver la hoja "Instrucciones" para el formato de cada Campo.',
    )
    for i, col in enumerate(COLUMNAS_HISTORIAL_ESPERADAS, start=1):
        celda = ws.cell(row=4, column=i, value=col)
        celda.font = Font(bold=True)
        ws.column_dimensions[celda.column_letter].width = max(16, len(col) + 2)
    fila_ejemplo = [
        "EJEMPLO", "Hora entrada programada", "09:00", dt.date.today(),
        None, "Lunes", "Cambio de horario temporal por capacitación",
    ]
    for i, valor in enumerate(fila_ejemplo, start=1):
        celda = ws.cell(row=5, column=i, value=valor)
        celda.font = Font(italic=True, color="9C6B13")
    ws.cell(row=6, column=1, value="↑ borrá esta fila de ejemplo antes de subir el archivo").font = Font(italic=True, size=9, color="9C6B13")

    wi = wb.create_sheet("Instrucciones")
    wi.cell(row=1, column=1, value="Qué va en cada Campo").font = Font(bold=True, size=13)
    wi.cell(row=3, column=1, value="Campo").font = Font(bold=True)
    wi.cell(row=3, column=2, value="Formato del Valor nuevo").font = Font(bold=True)
    fila = 4
    for campo, formato in FORMATO_POR_CAMPO.items():
        wi.cell(row=fila, column=1, value=campo)
        wi.cell(row=fila, column=2, value=formato)
        fila += 1
    fila += 1
    wi.cell(row=fila, column=1, value="Otras columnas").font = Font(bold=True)
    fila += 1
    for texto in [
        "DNI: sin ceros a la izquierda, tal como figura en el Maestro.",
        "Fecha desde: obligatoria.",
        "Fecha hasta (opcional): vacío = el cambio sigue vigente sin fecha de fin.",
        "Día de la semana (opcional): vacío = aplica todos los días del rango. Si querés que "
        "aplique solo ciertos días (ej. todos los martes y jueves), repetí la fila una vez por día.",
        "Comentario (opcional): libre, para dejar contexto de por qué se hizo el cambio.",
    ]:
        wi.cell(row=fila, column=1, value=texto)
        fila += 1
    wi.column_dimensions["A"].width = 30
    wi.column_dimensions["B"].width = 75

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parsear_historial_excel(fuente):
    df = pd.read_excel(fuente, sheet_name="Historial de cambios", header=3).dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    faltantes = [c for c in COLUMNAS_HISTORIAL_ESPERADAS if c not in df.columns]
    if faltantes:
        raise ValueError(
            f"el archivo no tiene el formato esperado -- faltan las columnas: {', '.join(faltantes)}. "
            f'Usá la plantilla del botón "Descargar plantilla".'
        )
    # Mismo gotcha de siempre: si ALGUNA fila de DNI viene vacía, pandas sube
    # la columna entera a float64 y ".astype(str)" sobre la columna original
    # deja "18074336.0" en vez de "18074336" para TODOS los DNIs válidos.
    dni_num = pd.to_numeric(df["DNI"], errors="coerce")
    df = df[dni_num.notna()].copy()
    df["DNI"] = dni_num[dni_num.notna()].astype("int64").astype(str)
    return df


@bp.route("/", methods=["GET"])
@requiere_pagina("historial")
def listar():
    buscar = request.args.get("buscar", "").strip()
    campo_filtro = request.args.get("campo", "").strip()

    session = get_session()
    try:
        query = session.query(HistorialCambio, Persona.nombre_completo).join(
            Persona, Persona.dni == HistorialCambio.dni
        )
        cond_scope = condicion_scope(Persona, current_user)
        if cond_scope is not None:
            query = query.filter(cond_scope)
        if buscar:
            like = f"%{buscar}%"
            query = query.filter((HistorialCambio.dni.ilike(like)) | (Persona.nombre_completo.ilike(like)))
        if campo_filtro:
            query = query.filter(HistorialCambio.campo == campo_filtro)
        filas = query.order_by(HistorialCambio.fecha_desde.desc()).all()
    finally:
        session.close()

    cambios = [{"c": c, "nombre": nombre} for c, nombre in filas]
    return render_template(
        "historial.html", usuario=current_user, cambios=cambios,
        campos=CAMPOS_VALIDOS, dias_semana=DIAS_SEMANA, formato_por_campo=FORMATO_POR_CAMPO,
        buscar=buscar, campo_filtro=campo_filtro,
    )


@bp.route("/crear", methods=["POST"])
@_analista_requerido
def crear():
    volver = request.form.get("volver") or url_for("historial.listar")
    datos, error = _leer_datos_comunes(request.form)
    if error:
        flash(error, "error")
        return redirect(volver)

    # Selección múltiple de días (Davor, 2026-08-25: "los dias, que sea
    # opción multiple") -- se guarda una fila por día elegido, mismo criterio
    # que ya usa cargar_historial() para leerlas (una fila = un día, o
    # dia_semana=None = todos los días del rango). Sin ningún día tildado,
    # queda como "todos los días" (una sola fila con None).
    dias = [d for d in request.form.getlist("dia_semana") if d in DIAS_SEMANA]
    dias_a_guardar = dias or [None]

    session = get_session()
    try:
        persona = _persona_en_scope(session, datos["dni"])
        if not persona:
            flash(f"No se encontró a nadie con DNI {datos['dni']} en tu alcance.", "error")
            return redirect(volver)
        for dia in dias_a_guardar:
            session.add(HistorialCambio(**datos, dia_semana=dia))
        session.commit()
        sufijo = f" ({len(dias_a_guardar)} días)" if len(dias_a_guardar) > 1 else ""
        flash(f"Cambio registrado para {persona.nombre_completo.title()}{sufijo}.", "ok")
    finally:
        session.close()
    return redirect(volver)


@bp.route("/<int:cambio_id>/editar", methods=["GET", "POST"])
@_analista_requerido
def editar(cambio_id):
    session = get_session()
    try:
        cambio = session.query(HistorialCambio).filter(HistorialCambio.id == cambio_id).first()
        if not cambio or not _persona_en_scope(session, cambio.dni):
            flash("No se encontró ese cambio o no tenés acceso a él.", "error")
            return redirect(url_for("historial.listar"))

        if request.method == "POST":
            datos, error = _leer_datos_comunes(request.form)
            if error:
                flash(error, "error")
                return redirect(url_for("historial.editar", cambio_id=cambio_id))
            dia_semana = request.form.get("dia_semana", "").strip() or None
            if dia_semana and dia_semana not in DIAS_SEMANA:
                flash("Día de la semana inválido.", "error")
                return redirect(url_for("historial.editar", cambio_id=cambio_id))
            if datos["dni"] != cambio.dni and not _persona_en_scope(session, datos["dni"]):
                flash(f"No se encontró a nadie con DNI {datos['dni']} en tu alcance.", "error")
                return redirect(url_for("historial.editar", cambio_id=cambio_id))
            datos["dia_semana"] = dia_semana
            for campo, valor in datos.items():
                setattr(cambio, campo, valor)
            session.commit()
            flash("Cambio actualizado.", "ok")
            return redirect(url_for("historial.listar"))

        persona = session.query(Persona).filter(Persona.dni == cambio.dni).first()
        return render_template(
            "historial_editar.html", usuario=current_user, cambio=cambio,
            nombre=persona.nombre_completo if persona else cambio.dni,
            campos=CAMPOS_VALIDOS, dias_semana=DIAS_SEMANA,
        )
    finally:
        session.close()


@bp.route("/<int:cambio_id>/eliminar", methods=["POST"])
@_analista_requerido
def eliminar(cambio_id):
    volver = request.form.get("volver") or url_for("historial.listar")
    session = get_session()
    try:
        cambio = session.query(HistorialCambio).filter(HistorialCambio.id == cambio_id).first()
        if not cambio or not _persona_en_scope(session, cambio.dni):
            flash("No se encontró ese cambio o no tenés acceso a él.", "error")
        else:
            session.delete(cambio)
            session.commit()
            flash("Cambio eliminado.", "ok")
    finally:
        session.close()
    return redirect(volver)


@bp.get("/plantilla.xlsx")
@_analista_requerido
def plantilla():
    buf = _plantilla_historial_excel()
    return send_file(
        buf, as_attachment=True, download_name="Plantilla_Historial_Cambios.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.post("/cargar")
@_analista_requerido
def cargar():
    archivo = request.files.get("archivo")
    if not archivo or archivo.filename == "":
        flash("Tenés que subir un archivo.", "error")
        return redirect(url_for("historial.listar"))

    try:
        df = _parsear_historial_excel(io.BytesIO(archivo.read()))
    except Exception as e:
        flash(f"Error leyendo el archivo: {e}", "error")
        return redirect(url_for("historial.listar"))

    creados, errores = 0, []
    session = get_session()
    try:
        for i, row in df.iterrows():
            fila_excel = i + 5  # fila 4 = encabezados, la primera fila de datos es la 5
            dni = row["DNI"]
            campo = _texto(row["Campo"]) or ""
            valor_nuevo = _texto(row["Valor nuevo"]) or ""
            dia_semana = _texto(row["Día de la semana (opcional)"])
            fecha_desde = _fecha(row["Fecha desde"])
            fecha_hasta = _fecha(row["Fecha hasta (opcional)"])

            if campo not in CAMPOS_VALIDOS:
                errores.append(f"Fila {fila_excel}: Campo \"{campo}\" inválido.")
                continue
            if not valor_nuevo:
                errores.append(f"Fila {fila_excel}: falta Valor nuevo.")
                continue
            if campo in CAMPOS_HORA:
                valor_normalizado = _normalizar_hora(valor_nuevo)
                if valor_normalizado is None:
                    errores.append(f"Fila {fila_excel}: \"{valor_nuevo}\" no es una hora válida -- usá el formato HH:MM (ej. 09:00).")
                    continue
                valor_nuevo = valor_normalizado
            if not fecha_desde:
                errores.append(f"Fila {fila_excel}: falta Fecha desde.")
                continue
            if fecha_hasta and fecha_hasta < fecha_desde:
                errores.append(f"Fila {fila_excel}: Fecha hasta anterior a Fecha desde.")
                continue
            if dia_semana and dia_semana not in DIAS_SEMANA:
                errores.append(f"Fila {fila_excel}: Día de la semana \"{dia_semana}\" inválido.")
                continue

            persona = _persona_en_scope(session, dni)
            if not persona:
                errores.append(f"Fila {fila_excel}: DNI {dni} no encontrado o fuera de tu alcance.")
                continue

            session.add(HistorialCambio(
                dni=dni, campo=campo, valor_nuevo=valor_nuevo, fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta, dia_semana=dia_semana, comentario=_texto(row["Comentario (opcional)"]),
            ))
            creados += 1
        session.commit()
    finally:
        session.close()

    mensaje = f"Carga completa: {creados} cambios registrados."
    if errores:
        mensaje += f" {len(errores)} filas con error -- " + " | ".join(errores[:10])
        if len(errores) > 10:
            mensaje += f" (+{len(errores) - 10} más)"
    flash(mensaje, "ok" if not errores else "error")
    return redirect(url_for("historial.listar"))
