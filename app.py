# -*- coding: utf-8 -*-
"""
Esqueleto en la nube del Proyecto MAC -- Fase 0 + Fase 1 del plan de
migración (ver DEPLOY.md). Este es un proyecto NUEVO e independiente del
app.py local que sigue corriendo en la laptop; no toca ni reemplaza nada de
lo existente todavía. Por ahora solo resuelve login/roles/multiusuario --
el reporte diario real llega en fases futuras (2 y 3 del roadmap).
"""
import datetime as dt
import io
import os
import re
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl.styles import Font
from sqlalchemy import text
from sqlalchemy.orm import aliased

PERU_TZ = ZoneInfo("America/Lima")  # sin horario de verano -- offset fijo UTC-5

from extensions import csrf, db, limiter, login_manager
from models import Usuario
from dimension_models import Feriado, Persona, PersonaSupervisorCanal, PersonaZonaCanal, SUBCANALES
from permisos import requiere_pagina, tiene_acceso
from dimension_models import get_session as get_dim_session
from fact_models import ClasificacionDiaria
from github_actions import disparar_workflow
from scoping import (
    CANALES_FILTRABLES, aplicar_filtros_extra, condicion_canal, condicion_scope,
    todos_overrides_supervisor_canal, todos_overrides_zona_canal,
)
from vacaciones import calcular_viajes_vacaciones

load_dotenv()  # lee .env en local; en PythonAnywhere las variables se cargan desde su panel, no de este archivo

_RE_PREFIJO_FALTA = re.compile(r"^falta\s*[-–]?\s*", re.IGNORECASE)


def _motivo_limpio(comentario):
    """Version chica de asistencia.py::_homologar_motivo() -- separada a
    proposito (evita acoplar app.py al modulo del blueprint) para agrupar
    "Faltas por motivo" del dashboard.

    pd.isna() en vez de "not comentario": una fila FALTA sin comentario del
    supervisor todavia llega como NaN de pandas (no None), y "not NaN" da
    False -- sin este chequeo, str(NaN) = "nan" terminaba mostrandose
    literal como motivo "Nan" en vez de agruparse en "Sin motivo"."""
    if pd.isna(comentario) or not str(comentario).strip():
        return "Sin motivo"
    texto = str(comentario).strip()
    while True:
        m = _RE_PREFIJO_FALTA.match(texto)
        if not m or m.end() == 0:
            break
        resto = texto[m.end():].strip()
        if resto == texto:
            break
        texto = resto
    texto = re.sub(r"\s*\([^)]*\)\s*$", "", texto).strip()
    texto = (texto[0].upper() + texto[1:]) if texto else "Sin motivo"
    # Homologaciones -- mismo motivo real, tipeado distinto por cada
    # supervisor. Se agregan acá a medida que se detectan (ver "Tráfico /
    # clima" == "Paro / clima": ambos son transito cortado por el clima).
    HOMOLOGACIONES = {
        "tráfico / clima": "Paro / clima", "trafico / clima": "Paro / clima",
        "feriado regional": "Feriado",
    }
    return HOMOLOGACIONES.get(texto.lower(), texto)


def create_app():
    app = Flask(__name__)
    _secret_key = os.environ.get("SECRET_KEY")
    if not _secret_key:
        # El repo es público -- este valor de respaldo queda visible en el
        # código fuente, así que si SECRET_KEY llegara a faltar en Render
        # (env var borrada, redeploy mal configurado) cualquiera podría
        # firmar una cookie de sesión válida. No se aborta el arranque
        # (rompería `python app.py` local sin nada configurado) pero se
        # avisa fuerte en los logs para que se note enseguida.
        print("ADVERTENCIA: SECRET_KEY no está seteada -- usando un valor de "
              "respaldo público (visible en el repo). Configurar SECRET_KEY "
              "en las variables de entorno de Render.")
        _secret_key = "dev-only-no-usar-en-produccion"
    app.config["SECRET_KEY"] = _secret_key
    app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///dev.db")
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    # Cookie de sesion solo por HTTPS -- Render sirve todo por HTTPS, asi
    # que esto no rompe produccion. SESSION_COOKIE_SECURE=false permite
    # override para correr local por http (python app.py en localhost).
    app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "true").lower() == "true"
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    # 10 MB -- de sobra para un Excel de Headcount, evita que alguien suba
    # un archivo gigante y cuelgue el worker (Render free tier, un solo
    # proceso web para todos los usuarios).
    app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

    db.init_app(app)
    login_manager.init_app(app)
    csrf.init_app(app)
    limiter.init_app(app)

    @app.after_request
    def _headers_seguridad(response):
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        # 'unsafe-inline' en script/style porque las plantillas usan <style>
        # y <script> embebidos (sin build step de assets) -- igual bloquea
        # cargar recursos desde dominios externos no autorizados, que es el
        # riesgo principal que mitiga un CSP.
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline'; img-src 'self' data:; "
            "frame-ancestors 'none'"
        )
        return response

    @login_manager.user_loader
    def load_user(user_id):
        # Flask-Login espera None (no una excepción) para "no autenticado" --
        # una cookie de sesión vieja/corrupta con un user_id no numérico
        # tiraba 500 en TODA request, incluso al intentar entrar a /login.
        try:
            return db.session.get(Usuario, int(user_id))
        except (TypeError, ValueError):
            return None

    from auth import bp as auth_bp
    from admin import bp as admin_bp
    from cargas import bp as cargas_bp
    from asistencia import bp as asistencia_bp
    from reportes import bp as reportes_bp
    from historial import bp as historial_bp
    app.register_blueprint(auth_bp)
    app.register_blueprint(admin_bp)
    app.register_blueprint(reportes_bp)
    app.register_blueprint(cargas_bp)
    app.register_blueprint(asistencia_bp)
    app.register_blueprint(historial_bp)

    # Disponible en todas las plantillas como {% if tiene_acceso(usuario, 'clave') %}
    # -- ver permisos.py.
    app.jinja_env.globals["tiene_acceso"] = tiene_acceso

    @app.errorhandler(405)
    def _metodo_no_permitido(e):
        # Pasa cuando el navegador reintenta con GET una URL que solo acepta
        # POST -- típico de una pestaña en segundo plano que Chrome/Safari
        # recarga solo, reproduciendo la última URL de la barra de
        # direcciones sin el método original (ej. /historial/crear después
        # de enviar un formulario). Sin esto se veía la página cruda de
        # error de Werkzeug en vez de volver a un lugar útil. No se usa
        # flash() acá -- el Dashboard (destino más común) no renderiza
        # mensajes flash, y quedaría colgado para aparecer sin contexto en
        # la próxima página que sí los muestre.
        destino = "dashboard" if current_user.is_authenticated else "auth.login"
        return redirect(url_for(destino))

    @app.get("/health")
    def health():
        try:
            db.session.execute(text("SELECT 1"))
            return jsonify({"status": "ok", "db": True}), 200
        except Exception as e:
            return jsonify({"status": "error", "db": False, "detalle": str(e)}), 500

    @app.get("/cron/pipeline")
    def cron_disparar_pipeline():
        """Disparador externo confiable para pipeline_completo.yml (Davor,
        2026-08-29) -- el `schedule` nativo de GitHub Actions resultó no ser
        confiable para esto (37 de las últimas 40 corridas fueron
        workflow_dispatch, casi nada por el cron de cada 5 min; se detectó
        un hueco real de 92 min sin ninguna corrida). Pensado para que un
        servicio externo (cron-job.org o similar) le pegue cada 5 min.

        Protegido por token en la URL (?token=...) -- sin login, porque un
        servicio de cron externo no tiene cómo autenticarse con nuestra
        sesión. Sin CRON_TRIGGER_TOKEN configurado en el entorno, este
        endpoint queda deshabilitado (404) en vez de aceptar cualquier
        pedido sin token. Es seguro pegarle más seguido de lo necesario --
        pipeline_completo.py ya tiene su propio candado (db_lock) y no
        corre dos veces en paralelo aunque este disparo coincida con uno
        manual o del cron nativo (que sigue activo, esto es un respaldo, no
        un reemplazo)."""
        token_esperado = os.environ.get("CRON_TRIGGER_TOKEN")
        if not token_esperado:
            return jsonify({"error": "CRON_TRIGGER_TOKEN no configurado"}), 404
        if request.args.get("token") != token_esperado:
            return jsonify({"error": "token inválido"}), 403
        ok, mensaje = disparar_workflow("pipeline_completo.yml")
        return jsonify({"ok": ok, "mensaje": mensaje}), (200 if ok else 502)

    @app.get("/")
    @login_required
    def dashboard():
        # Fase 6 (2026-08-21): dashboard real -- consulta Postgres en vivo en
        # cada visita, no depende de dashboard_data.json (ese archivo lo
        # genera export_dashboard_data.py en cada corrida del pipeline, pero
        # vive en SharePoint/local, no aca).
        #
        # 2026-08-22: el periodo (desde/hasta) ahora es seleccionable -- antes
        # el indicador de efectividad mezclaba TODO el historico desde el
        # inicio del sistema en un solo numero, sin poder acotarlo.
        hoy_peru = dt.datetime.now(PERU_TZ).date()
        try:
            hasta = dt.date.fromisoformat(request.args.get("hasta", "")) if request.args.get("hasta") else hoy_peru
        except ValueError:
            hasta = hoy_peru
        try:
            desde = dt.date.fromisoformat(request.args.get("desde", "")) if request.args.get("desde") else hoy_peru.replace(day=1)
        except ValueError:
            desde = hoy_peru.replace(day=1)
        if desde > hasta:
            desde, hasta = hasta, desde

        # Jinja no tiene timedelta a mano -- se precalculan los 3 atajos de
        # periodo acá en vez de hacer aritmetica de fechas en la plantilla.
        presets = {
            "7d": {"desde": (hoy_peru - dt.timedelta(days=6)).isoformat(), "hasta": hoy_peru.isoformat()},
            "30d": {"desde": (hoy_peru - dt.timedelta(days=29)).isoformat(), "hasta": hoy_peru.isoformat()},
            "mes": {"desde": hoy_peru.replace(day=1).isoformat(), "hasta": hoy_peru.isoformat()},
        }

        cond_scope = condicion_scope(Persona, current_user)  # None = sin restriccion

        # Filtros adicionales (2026-08-22, pedido explicito) -- Rol/Región/
        # Supervisor, encima del scope de acceso (cond_scope). "" en el query
        # string se trata como "sin filtro" (opción "Todos" del <select>).
        rol_filtro = request.args.get("rol") or None
        region_filtro = request.args.get("region") or None
        supervisor_filtro = request.args.get("supervisor") or None
        # Canal (Davor, 2026-08-29) -- SOLO admin: "debo tener un filtro
        # para ver Tradicional, Farmacia y AU" -- un analista de canal ya
        # está acotado por condicion_scope(), no lo necesita.
        es_admin = current_user.rol == "admin"
        canal_filtro = (request.args.get("canal") or None) if es_admin else None

        dim_session = get_dim_session()
        try:
            def _con_filtros(q):
                if cond_scope is not None:
                    q = q.filter(cond_scope)
                if rol_filtro:
                    q = q.filter(Persona.rol == rol_filtro)
                if region_filtro:
                    q = q.filter(Persona.region == region_filtro)
                if supervisor_filtro:
                    q = q.filter(Persona.supervisor_dni == supervisor_filtro)
                if canal_filtro:
                    q = q.filter(condicion_canal(Persona, canal_filtro))
                return q

            query = (
                dim_session.query(ClasificacionDiaria.dni, Persona.nombre_completo, Persona.region,
                                   ClasificacionDiaria.fecha, ClasificacionDiaria.estado,
                                   ClasificacionDiaria.entrada_real,
                                   ClasificacionDiaria.salida_real, ClasificacionDiaria.comentario_supervisor)
                .join(Persona, Persona.dni == ClasificacionDiaria.dni)
                .filter(ClasificacionDiaria.fecha >= desde, ClasificacionDiaria.fecha <= hasta)
            )
            filas = _con_filtros(query).all()

            # "Hoy" es independiente del periodo elegido (si elegis un rango
            # pasado, igual queres ver como viene el dia de hoy) -- consulta
            # aparte, chica.
            query_hoy = (
                dim_session.query(ClasificacionDiaria.estado)
                .join(Persona, Persona.dni == ClasificacionDiaria.dni)
                .filter(ClasificacionDiaria.fecha == hoy_peru)
            )
            filas_hoy = _con_filtros(query_hoy).all()

            headcount_query = dim_session.query(Persona.dni).filter(Persona.estado == "Activo")
            headcount_actual = _con_filtros(headcount_query).count()

            # Opciones de los 3 <select> -- acotadas al mismo scope de acceso
            # (cond_scope) para que un analista/supervisor no vea roles/
            # regiones/supervisores de gente que de todos modos no puede ver.
            q_roles = dim_session.query(Persona.rol).filter(Persona.rol.isnot(None))
            if cond_scope is not None:
                q_roles = q_roles.filter(cond_scope)
            roles_disponibles = sorted({r for (r,) in q_roles.distinct().all()})

            q_regiones = dim_session.query(Persona.region).filter(Persona.region.isnot(None))
            if cond_scope is not None:
                q_regiones = q_regiones.filter(cond_scope)
            regiones_disponibles = sorted({r for (r,) in q_regiones.distinct().all()})

            SupervisorPersona = aliased(Persona)
            q_sup = (
                dim_session.query(Persona.supervisor_dni, SupervisorPersona.nombre_completo)
                .join(SupervisorPersona, SupervisorPersona.dni == Persona.supervisor_dni)
                .filter(Persona.supervisor_dni.isnot(None))
            )
            if cond_scope is not None:
                q_sup = q_sup.filter(cond_scope)
            supervisores_disponibles = sorted(set(q_sup.distinct().all()), key=lambda t: t[1].title())
        finally:
            dim_session.close()

        periodo_args = {"desde": desde.isoformat(), "hasta": hasta.isoformat()}
        filtro_args = {
            "rol": rol_filtro or "", "region": region_filtro or "", "supervisor": supervisor_filtro or "",
            "canal": canal_filtro or "",
        }
        canales_disponibles = CANALES_FILTRABLES if es_admin else []

        if not filas:
            return render_template(
                "dashboard.html", usuario=current_user, hay_datos=False,
                periodo_desde=desde, periodo_hasta=hasta, periodo_args=periodo_args, presets=presets,
                filtro_args=filtro_args, roles_disponibles=roles_disponibles,
                regiones_disponibles=regiones_disponibles, supervisores_disponibles=supervisores_disponibles,
                canales_disponibles=canales_disponibles,
            )

        r = pd.DataFrame(filas, columns=[
            "dni", "nombre", "region", "fecha", "estado", "entrada_real", "salida_real", "comentario",
        ])
        r["fecha"] = pd.to_datetime(r["fecha"])
        r["estado_base"] = r["estado"].apply(lambda s: s.split(" (")[0])
        r["region"] = r["region"].fillna("Sin región")

        hoy_es = [f.split(" (")[0] for (f,) in filas_hoy]
        resumen_hoy = {
            "asistio": hoy_es.count("ASISTIÓ A TIEMPO"),
            "tardanza": hoy_es.count("TARDANZA"),
            "falta": hoy_es.count("FALTA"),
            "vacante": hoy_es.count("VACANTE"),
            "vacaciones": hoy_es.count("VACACIONES"),
            "total": len(hoy_es),
        }

        total = len(r)
        n_asistio = int((r["estado_base"] == "ASISTIÓ A TIEMPO").sum())
        n_tardanza = int((r["estado_base"] == "TARDANZA").sum())
        n_falta = int((r["estado_base"] == "FALTA").sum())
        n_vacante = int((r["estado_base"] == "VACANTE").sum())
        n_vacaciones = int((r["estado_base"] == "VACACIONES").sum())
        pct_efectividad = round((n_asistio + n_tardanza) / total * 100, 1) if total else 0.0
        n_personas_evaluadas = r["dni"].nunique()
        n_incidencias = n_falta + n_vacante + n_vacaciones

        # Jornada promedio: solo dias con entrada Y salida real (asistio o
        # tardanza) -- ambas guardadas como texto "HH:MM:SS".
        con_jornada = r[r["entrada_real"].notna() & r["salida_real"].notna()].copy()
        if len(con_jornada):
            entrada_td = pd.to_timedelta(con_jornada["entrada_real"])
            salida_td = pd.to_timedelta(con_jornada["salida_real"])
            duracion = (salida_td - entrada_td).dt.total_seconds() / 3600
            duracion = duracion[(duracion > 0) & (duracion < 16)]  # descarta datos corruptos/turnos nocturnos raros
            jornada_promedio_h = duracion.mean() if len(duracion) else None
        else:
            jornada_promedio_h = None
        if jornada_promedio_h is not None:
            horas = int(jornada_promedio_h)
            minutos = int(round((jornada_promedio_h - horas) * 60))
            jornada_promedio = f"{horas}h {minutos:02d}m"
        else:
            jornada_promedio = "—"

        # Tendencia diaria = % de efectividad (asistió a tiempo + tardanza,
        # sobre el total evaluado ese día), no el conteo crudo de falta/
        # tardanza -- pedido explícito de Davor (2026-08-24), la misma
        # métrica que ya muestra el aro de arriba pero día a día.
        tendencia = r.groupby(r["fecha"].dt.date)["estado_base"].value_counts().unstack(fill_value=0)
        for col in ["ASISTIÓ A TIEMPO", "TARDANZA"]:
            if col not in tendencia.columns:
                tendencia[col] = 0
        tendencia = tendencia.sort_index()
        tendencia["total_dia"] = tendencia.sum(axis=1)
        tendencia["pct_dia"] = (
            (tendencia["ASISTIÓ A TIEMPO"] + tendencia["TARDANZA"]) / tendencia["total_dia"] * 100
        ).round(1)
        serie_tendencia = [
            {"dia": d.strftime("%d/%m"), "pct": float(fila["pct_dia"])}
            for d, fila in tendencia.iterrows()
        ]

        # Puntos del gráfico de líneas, en % (0-100 en x e y) -- los puntos
        # se dibujan como <div> posicionados por porcentaje (no <circle> de
        # SVG) para que no salgan ovalados: el SVG usa preserveAspectRatio
        # "none" para llenar todo el ancho del panel, así que un <circle>
        # ahí adentro se estira de forma no uniforme y deja de ser redondo.
        # El rango vertical usa el min/max real de los datos (con margen) en
        # vez de 0-100 fijo, que aplastaba la línea contra el techo (el % de
        # efectividad real casi siempre anda entre 75-100%).
        if serie_tendencia:
            pcts_dia = [d["pct"] for d in serie_tendencia]
            min_pct, max_pct = min(pcts_dia), max(pcts_dia)
            if max_pct == min_pct:
                min_pct, max_pct = max(0.0, min_pct - 5), min(100.0, max_pct + 5)
            else:
                margen = (max_pct - min_pct) * 0.15
                min_pct, max_pct = max(0.0, min_pct - margen), min(100.0, max_pct + margen)
            n = len(serie_tendencia)
            for i, d in enumerate(serie_tendencia):
                d["x"] = round(i / (n - 1) * 100, 2) if n > 1 else 50.0
                d["y"] = round(100 - (d["pct"] - min_pct) / (max_pct - min_pct) * 100, 1)
            tendencia_puntos = " ".join(f"{d['x']},{d['y']}" for d in serie_tendencia)
        else:
            tendencia_puntos = ""

        # "Tiene reemplazo - {motivo cese}" no es un motivo de falta -- es el
        # reporte de baja/reemplazo que un supervisor manda desde la app
        # móvil (ver reporte_diario_9am.py::comentarios_supervisor_dia()),
        # que queda como comentario en los días que la persona sale Falta
        # hasta que se procesa el reemplazo. No cuenta como "motivo".
        faltas_sin_reemplazo = r[
            (r["estado_base"] == "FALTA")
            & ~r["comentario"].astype(str).str.strip().str.lower().str.startswith("tiene reemplazo")
        ]
        motivo_falta = faltas_sin_reemplazo["comentario"].apply(_motivo_limpio)
        faltas_por_motivo = motivo_falta.value_counts().head(8)

        # Tabla "Faltas" del detalle del periodo: descanso médico y licencia
        # son ausencias justificadas y programadas, no fallas de asistencia
        # -- no deben sumar al ranking de faltas por persona (pedido
        # explícito de Davor, 2026-08-24).
        MOTIVOS_NO_CONTABLES_FALTAS = {"Descanso médico", "Licencia"}
        r["_falta_contable"] = r["estado_base"] == "FALTA"
        r.loc[faltas_sin_reemplazo.index, "_falta_contable"] = ~motivo_falta.isin(MOTIVOS_NO_CONTABLES_FALTAS)

        resumen_persona = r.groupby(["dni", "nombre"]).agg(
            dias=("estado_base", "size"),
            falta=("_falta_contable", "sum"),
            tardanza=("estado_base", lambda s: int((s == "TARDANZA").sum())),
        ).reset_index()
        resumen_persona["falta"] = resumen_persona["falta"].astype(int)
        # Se muestran todas las personas con al menos 1 falta/tardanza en el
        # periodo, no solo un "Top N" -- pedido explícito: el detalle del
        # periodo debe mostrar todos los datos aunque sea un solo registro.
        top_falta = (
            resumen_persona[resumen_persona["falta"] > 0]
            .sort_values(["falta", "tardanza"], ascending=False).to_dict("records")
        )
        top_tardanza = (
            resumen_persona[resumen_persona["tardanza"] > 0]
            .sort_values(["tardanza", "falta"], ascending=False).to_dict("records")
        )

        # Agrupa los días VACACIONES sueltos en "viajes" contiguos por
        # persona y calcula la duración en días CALENDARIO (salida -> regreso
        # a marcar), no solo el conteo de filas VACACIONES -- el motor no
        # genera fila para domingos/feriados dentro del periodo, así que
        # contar filas subestimaba la duración real (pedido explícito:
        # "si salio el 10.08 y regreso el 17.08 a marcar, salio de
        # vacaciones 7 dias", contando domingos de por medio).
        vacaciones_detalle = calcular_viajes_vacaciones(dim_session, r, hasta)

        # Dias marcados VACANTE dentro del periodo elegido -- antes solo
        # miraba Personas que HOY siguen con estado="Vacante" en el Maestro,
        # asi que una vacante ya cubierta durante el periodo no aparecia
        # aunque si hubiera tenido dias vacante reales.
        vacantes_detalle = (
            r[r["estado_base"] == "VACANTE"].groupby(["dni", "nombre"]).size()
            .reset_index(name="dias").sort_values("dias", ascending=False)
            .to_dict("records")
        )

        data = {
            "periodo": {"desde": desde.strftime("%d/%m/%Y"), "hasta": hasta.strftime("%d/%m/%Y")},
            "resumen_hoy": resumen_hoy,
            "resumen": {
                "total": total, "asistio": n_asistio, "tardanza": n_tardanza, "falta": n_falta,
                "vacante": n_vacante, "vacaciones": n_vacaciones, "pct_efectividad": pct_efectividad,
            },
            "kpis": {
                "headcount_actual": headcount_actual, "personas_evaluadas": int(n_personas_evaluadas),
                "incidencias": n_incidencias, "jornada_promedio": jornada_promedio,
            },
            "tendencia": serie_tendencia,
            "tendencia_puntos": tendencia_puntos,
            "faltas_por_motivo": list(faltas_por_motivo.items()),
            "top_falta": top_falta,
            "top_tardanza": top_tardanza,
            "vacaciones_detalle": vacaciones_detalle,
            "vacantes_detalle": vacantes_detalle,
        }
        return render_template(
            "dashboard.html", usuario=current_user, hay_datos=True, data=data,
            periodo_desde=desde, periodo_hasta=hasta, periodo_args=periodo_args, presets=presets,
            filtro_args=filtro_args, roles_disponibles=roles_disponibles,
            regiones_disponibles=regiones_disponibles, supervisores_disponibles=supervisores_disponibles,
            canales_disponibles=canales_disponibles,
        )

    def _filtros_personal(dim_session):
        # Región/Ciudad/Subcanal (Davor, 2026-08-29: "agregar filtro,
        # región, ciudad, canal y subcanal") -- mismo criterio que
        # reportes.py::_filtros_admin(): admin y analista sí filtran,
        # supervisor no (ya está acotado a su equipo, filtrar no le sirve).
        # Canal sigue siendo SOLO admin (un analista de canal ya está
        # acotado por condicion_scope()).
        puede_filtrar = current_user.rol in ("admin", "analista")
        es_admin = current_user.rol == "admin"
        filtro_args = {
            "region": (request.args.get("region") or "") if puede_filtrar else "",
            "ciudad": (request.args.get("ciudad") or "") if puede_filtrar else "",
            "canal": (request.args.get("canal") or "") if es_admin else "",
            "subcanal": (request.args.get("subcanal") or "") if puede_filtrar else "",
            # Estado (Davor, 2026-09-04: "agregar un filtro de Estado, para
            # filtrar solo activos") -- sin filtro, "Personal" siempre
            # mostro tambien Inactivo/Vacante mezclados con Activo.
            "estado": (request.args.get("estado") or "") if puede_filtrar else "",
            # Supervisor (Davor, 2026-09-04: "Agregale filtro supervisor
            # también") -- mismo patrón que asistencia.py::_filtros_marcar().
            "supervisor": (request.args.get("supervisor") or "") if puede_filtrar else "",
        }
        if not puede_filtrar:
            return filtro_args, [], [], [], [], [], []
        cond_scope = condicion_scope(Persona, current_user)

        # Filtros cruzados (Davor, 2026-09-04: "Que sean filtros cruzados")
        # -- las opciones de Región/Ciudad/Supervisor se calculan aplicando
        # los DEMÁS filtros ya elegidos (todos menos el propio), no solo el
        # scope base; sin esto, elegir Canal=Farmacia seguía mostrando en
        # "Supervisor" a gente que no tiene NADIE en Farmacia. Canal/
        # Subcanal/Estado quedan fuera a propósito -- son listas fijas de
        # valores posibles (CANALES_FILTRABLES/SUBCANALES/3 estados), no
        # "lo que hay ahora mismo", así que cruzarlas no aporta.
        def _query_cruzada(excepto):
            query = dim_session.query(Persona)
            if cond_scope is not None:
                query = query.filter(cond_scope)
            return aplicar_filtros_extra(
                query, Persona,
                region_filtro=None if excepto == "region" else filtro_args["region"],
                ciudad_filtro=None if excepto == "ciudad" else filtro_args["ciudad"],
                canal_filtro=None if excepto == "canal" else filtro_args["canal"],
                subcanal_filtro=None if excepto == "subcanal" else filtro_args["subcanal"],
                estado_filtro=None if excepto == "estado" else filtro_args["estado"],
                supervisor_filtro=None if excepto == "supervisor" else filtro_args["supervisor"],
            )

        def _valores_cruzados(columna, excepto):
            q = _query_cruzada(excepto).filter(columna.isnot(None)).with_entities(columna)
            return sorted({v for (v,) in q.distinct().all() if v})

        regiones_disponibles = _valores_cruzados(Persona.region, "region")
        ciudades_disponibles = _valores_cruzados(Persona.ciudad, "ciudad")
        canales_disponibles = CANALES_FILTRABLES if es_admin else []
        # Lista fija (no _valores_cruzados(Persona.estado, ...)) -- son solo
        # estos 3 valores posibles en todo el sistema, ver
        # dar_de_baja_submit()/cargas.py.
        estados_disponibles = ["Activo", "Inactivo", "Vacante"]
        SupervisorPersona = aliased(Persona)
        q_sup = (
            _query_cruzada("supervisor")
            .with_entities(Persona.supervisor_dni, SupervisorPersona.nombre_completo)
            .join(SupervisorPersona, SupervisorPersona.dni == Persona.supervisor_dni)
            .filter(Persona.supervisor_dni.isnot(None))
        )
        supervisores_disponibles = sorted(set(q_sup.distinct().all()), key=lambda t: (t[1] or "").title())
        return (
            filtro_args, regiones_disponibles, ciudades_disponibles, canales_disponibles, SUBCANALES,
            estados_disponibles, supervisores_disponibles,
        )

    def _consultar_personal(dim_session, filtro_args):
        query = dim_session.query(Persona)
        cond_scope = condicion_scope(Persona, current_user)
        if cond_scope is not None:
            query = query.filter(cond_scope)
        query = aplicar_filtros_extra(
            query, Persona, region_filtro=filtro_args["region"], ciudad_filtro=filtro_args["ciudad"],
            canal_filtro=filtro_args["canal"], subcanal_filtro=filtro_args["subcanal"],
            estado_filtro=filtro_args["estado"], supervisor_filtro=filtro_args["supervisor"],
        )
        return query.order_by(Persona.estado, Persona.nombre_completo).all()

    _TODOS_LOS_CANALES = ["Tradicional", "Farmacia", "Autoservicio"]

    def _resolver_supervisores(dim_session, personas):
        """{dni: {"principal": nombre_o_None, "principal_canal": etiqueta_o_None,
        "overrides": [(canal, nombre, supervisor_dni), ...]}} (Davor,
        2026-08-29: "que aparezca el nombre y también si hay 2 supers para
        ese mercaderista"). El nombre se busca SIN aplicar condicion_scope()
        -- el supervisor "de otro canal" de un compartido puede pertenecer a
        otro analista, y aun así hay que poder mostrar su nombre acá (mismo
        criterio que ya usa asistencia.py para este mismo problema).

        `principal_canal` (Davor, 2026-09-04: para un Multicanal con
        overrides de Farmacia/AU, el principal ES el de Tradicional -- "Pon
        en el item de supervisor... Tradicional: Daniela cabanillas") --
        cuando hay overrides que cubren SOLO algunos canales, el principal
        se etiqueta con el/los canal(es) que quedan sin override, en vez de
        mostrarlo suelto sin aclarar a cuál corresponde.

        El `supervisor_dni` en cada override (Davor, 2026-09-04: "ponle
        editar a supervisores también") es lo que necesita el <select> de
        edición en Personal para preseleccionar la opción correcta -- el
        nombre solo no alcanza para eso."""
        overrides = todos_overrides_supervisor_canal(dim_session, [p.dni for p in personas])
        dnis_necesarios = {p.supervisor_dni for p in personas if p.supervisor_dni}
        for filas in overrides.values():
            dnis_necesarios.update(sup_dni for _canal, sup_dni in filas)
        nombre_de = dict(
            dim_session.query(Persona.dni, Persona.nombre_completo).filter(Persona.dni.in_(dnis_necesarios)).all()
        ) if dnis_necesarios else {}
        resultado = {}
        for p in personas:
            principal_nombre = (nombre_de.get(p.supervisor_dni) or p.supervisor_dni) if p.supervisor_dni else None
            overrides_p = [
                (canal.title(), nombre_de.get(sup_dni) or sup_dni, sup_dni)
                for canal, sup_dni in overrides.get(p.dni, [])
            ]
            principal_canal = None
            if overrides_p and principal_nombre:
                cubiertos = {canal for canal, _, _ in overrides_p}
                faltantes = [c for c in _TODOS_LOS_CANALES if c not in cubiertos]
                if faltantes:
                    principal_canal = ", ".join(faltantes)
            resultado[p.dni] = {
                "principal": principal_nombre,
                "principal_canal": principal_canal,
                "overrides": overrides_p,
            }
        return resultado

    def _resolver_zonas(dim_session, personas):
        """{dni: {"principal": zona_o_None, "principal_canal": etiqueta_o_None,
        "overrides": [(canal, zona), ...]}} -- mismo patrón que
        _resolver_supervisores() pero para Zona (Davor, 2026-09-04: "Misma
        lógica para zonas, ya que tienen zonas por canal también")."""
        overrides = todos_overrides_zona_canal(dim_session, [p.dni for p in personas])
        resultado = {}
        for p in personas:
            overrides_p = [(canal.title(), zona) for canal, zona in overrides.get(p.dni, [])]
            principal_canal = None
            if overrides_p and p.zona:
                cubiertos = {canal for canal, _ in overrides_p}
                faltantes = [c for c in _TODOS_LOS_CANALES if c not in cubiertos]
                if faltantes:
                    principal_canal = ", ".join(faltantes)
            resultado[p.dni] = {
                "principal": p.zona,
                "principal_canal": principal_canal,
                "overrides": overrides_p,
            }
        return resultado

    @app.get("/personal")
    @requiere_pagina("personal")
    def personal():
        # Fase 2: primera pantalla que lee directo de Postgres (dimension_models,
        # las mismas tablas que migrar_dimensiones_a_postgres.py pobló) --
        # solo lectura por ahora, el alta/baja/reemplazo sigue siendo
        # agregar_reemplazo.py (CLI local) hasta que haya un formulario acá.
        dim_session = get_dim_session()
        try:
            filtro_args, regiones_disp, ciudades_disp, canales_disp, subcanales_disp, estados_disp, supervisores_disp = _filtros_personal(dim_session)
            personas = _consultar_personal(dim_session, filtro_args)
            supervisores_por_dni = _resolver_supervisores(dim_session, personas)
            zonas_por_dni = _resolver_zonas(dim_session, personas)
            puede_editar = current_user.rol in ("admin", "analista")
            # Lista GLOBAL (no acotada por condicion_scope), a propósito
            # (Davor, 2026-09-04: "ponle editar a supervisores también") --
            # un supervisor puede tener gente a cargo en el headcount de MÁS
            # DE UN analista (ver cargas.py::supervisores_propios, mismo
            # criterio), así que el <select> de edición debe poder asignar
            # a cualquiera, no solo a los visibles en el scope de quien edita.
            supervisores_editables = sorted(
                (
                    (dni, nombre) for dni, nombre, rol in
                    dim_session.query(Persona.dni, Persona.nombre_completo, Persona.rol)
                    if (rol or "").strip().upper().startswith("SUPERVISOR")
                ),
                key=lambda t: (t[1] or "").title(),
            ) if puede_editar else []
        finally:
            dim_session.close()
        filtro_qs = {k: v for k, v in filtro_args.items() if v}
        return render_template(
            "personal.html", usuario=current_user, personas=personas,
            filtro_args=filtro_args, filtro_qs=filtro_qs, regiones_disponibles=regiones_disp, ciudades_disponibles=ciudades_disp,
            canales_disponibles=canales_disp, subcanales_disponibles=subcanales_disp, estados_disponibles=estados_disp,
            supervisores_disponibles=supervisores_disp, supervisores_editables=supervisores_editables,
            subcanales=SUBCANALES, puede_editar=puede_editar, supervisores_por_dni=supervisores_por_dni,
            zonas_por_dni=zonas_por_dni,
        )

    @app.post("/personal/editar")
    @requiere_pagina("personal")
    def personal_editar():
        # Edición inline (Davor, 2026-08-29): Subcanal (lista fija) + Zona
        # (texto libre) con un botón "Guardar" por fila -- antes Subcanal
        # se guardaba solo con el cambio de select, pero al sumar Zona (un
        # <input type=text> no dispara onchange de forma útil) hace falta
        # un submit explícito. Se reusa condicion_scope() para que un
        # analista solo pueda tocar DNIs de su propio equipo.
        if current_user.rol not in ("admin", "analista"):
            flash("No tenés permiso para editar Personal.", "error")
            return redirect(url_for("personal"))
        dni = request.form.get("dni", "")
        subcanal = (request.form.get("subcanal") or "").strip()
        # modo_canal=1 (Davor, 2026-09-04: "Misma lógica para zonas... con
        # la opción de editarlo en el maestro ahora" / "ponle editar a
        # supervisores también") -- Multicanal manda 2 campos por canal
        # (Tradicional / Farmacia+AU) tanto para Zona como para Supervisor,
        # en vez del campo único de siempre; el resto de las personas
        # (un solo canal) sigue con el campo único, sin cambio de
        # comportamiento.
        modo_canal = request.form.get("modo_canal") == "1"
        zona = (request.form.get("zona") or "").strip()
        zona_tradicional = (request.form.get("zona_tradicional") or "").strip()
        zona_farmacia_au = (request.form.get("zona_farmacia_au") or "").strip()
        supervisor = (request.form.get("supervisor") or "").strip()
        supervisor_tradicional = (request.form.get("supervisor_tradicional") or "").strip()
        supervisor_farmacia_au = (request.form.get("supervisor_farmacia_au") or "").strip()
        if subcanal and subcanal not in SUBCANALES:
            flash("Subcanal inválido.", "error")
            return redirect(url_for("personal"))
        dim_session = get_dim_session()
        try:
            # Los DNI de supervisor vienen de un <select>, pero se validan
            # igual antes de guardar (Persona.supervisor_dni es FK -- un
            # valor inválido revienta el commit con un error feo en vez de
            # un mensaje claro).
            dnis_sup_pedidos = {v for v in (supervisor, supervisor_tradicional, supervisor_farmacia_au) if v}
            if dnis_sup_pedidos:
                dnis_validos = {
                    d for (d,) in dim_session.query(Persona.dni).filter(Persona.dni.in_(dnis_sup_pedidos)).all()
                }
                if dnis_sup_pedidos - dnis_validos:
                    flash("Supervisor inválido.", "error")
                    return redirect(url_for("personal"))

            query = dim_session.query(Persona).filter(Persona.dni == dni)
            cond_scope = condicion_scope(Persona, current_user)
            if cond_scope is not None:
                query = query.filter(cond_scope)
            persona = query.first()
            if persona is None:
                flash("No se encontró a esa persona en tu equipo.", "error")
                return redirect(url_for("personal"))
            persona.subcanal = subcanal or None
            if modo_canal:
                persona.zona = zona_tradicional or None
                persona.supervisor_dni = supervisor_tradicional or None
                if zona_farmacia_au:
                    for canal in ("FARMACIA", "AUTOSERVICIO"):
                        fila_zc = dim_session.query(PersonaZonaCanal).filter_by(dni=dni, canal=canal).first()
                        if fila_zc:
                            fila_zc.zona = zona_farmacia_au
                        else:
                            dim_session.add(PersonaZonaCanal(dni=dni, canal=canal, zona=zona_farmacia_au))
                else:
                    dim_session.query(PersonaZonaCanal).filter(
                        PersonaZonaCanal.dni == dni, PersonaZonaCanal.canal.in_(("FARMACIA", "AUTOSERVICIO")),
                    ).delete(synchronize_session=False)
                if supervisor_farmacia_au:
                    for canal in ("FARMACIA", "AUTOSERVICIO"):
                        fila_sc = dim_session.query(PersonaSupervisorCanal).filter_by(dni=dni, canal=canal).first()
                        if fila_sc:
                            fila_sc.supervisor_dni = supervisor_farmacia_au
                        else:
                            dim_session.add(PersonaSupervisorCanal(dni=dni, canal=canal, supervisor_dni=supervisor_farmacia_au))
                else:
                    dim_session.query(PersonaSupervisorCanal).filter(
                        PersonaSupervisorCanal.dni == dni, PersonaSupervisorCanal.canal.in_(("FARMACIA", "AUTOSERVICIO")),
                    ).delete(synchronize_session=False)
            else:
                persona.zona = zona or None
                persona.supervisor_dni = supervisor or None
            dim_session.commit()
            flash(f"{persona.nombre_completo.title()} actualizado.", "ok")
        finally:
            dim_session.close()
        return redirect(url_for("personal", **request.args))

    @app.post("/personal/reactivar")
    @requiere_pagina("personal")
    def personal_reactivar():
        # Davor, 2026-08-30: caso Jose Quiñones (descanso médico largo,
        # cubierto por Yuridia, vuelve en setiembre) -- el sistema no tiene
        # un estado "en licencia", así que una baja temporal se modela
        # igual que una baja definitiva (Vacante -> reemplazo -> Inactivo).
        # "Agregar reemplazo" no sirve para reactivar a la MISMA persona en
        # su propio puesto (reemplazos.py lo bloquea a propósito, ver
        # comentario ahí), así que hace falta este botón aparte: vuelve a
        # Activo y limpia fecha/motivo de baja, sin tocar canal/zona/
        # supervisor/patrón -- son su propia fila, nunca se borraron.
        if current_user.rol not in ("admin", "analista"):
            flash("No tenés permiso para reactivar en Personal.", "error")
            return redirect(url_for("personal"))
        dni = request.form.get("dni", "")
        dim_session = get_dim_session()
        try:
            query = dim_session.query(Persona).filter(Persona.dni == dni)
            cond_scope = condicion_scope(Persona, current_user)
            if cond_scope is not None:
                query = query.filter(cond_scope)
            persona = query.first()
            if persona is None:
                flash("No se encontró a esa persona en tu equipo.", "error")
            elif persona.estado == "Activo":
                flash(f"{persona.nombre_completo.title()} ya está Activo.", "error")
            else:
                persona.estado = "Activo"
                persona.fecha_baja = None
                persona.motivo_baja = None
                persona.dado_de_baja_por = None
                dim_session.commit()
                flash(f"{persona.nombre_completo.title()} reactivado.", "ok")
        finally:
            dim_session.close()
        return redirect(url_for("personal", **request.args))

    @app.get("/personal/exportar")
    @requiere_pagina("personal")
    def personal_exportar():
        dim_session = get_dim_session()
        try:
            filtro_args, _regiones_disp, _ciudades_disp, _canales_disp, _subcanales_disp, _estados_disp, _supervisores_disp = _filtros_personal(dim_session)
            personas = _consultar_personal(dim_session, filtro_args)
            supervisores_por_dni = _resolver_supervisores(dim_session, personas)
        finally:
            dim_session.close()

        columnas = [
            "DNI", "Nombre", "Estado", "Fecha de baja", "Rol", "Canal", "Subcanal",
            "Región", "Ciudad", "Zona", "Supervisor", "Supervisores adicionales",
            "Correo", "Analista",
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Personal"
        ws.append(columnas)
        for celda in ws[1]:
            celda.font = Font(bold=True)
        for p in personas:
            sup = supervisores_por_dni.get(p.dni, {"principal": None, "overrides": []})
            adicionales = ", ".join(f"{c}: {n}" for c, n in sup["overrides"])
            ws.append([
                p.dni, p.nombre_completo, p.estado, p.fecha_baja, p.rol, p.canal, p.subcanal,
                p.region, p.ciudad, p.zona, sup["principal"], adicionales,
                p.correo, p.analista_propietario,
            ])
        for i, titulo in enumerate(columnas, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(titulo) + 2)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        hoy = dt.datetime.now(PERU_TZ).strftime("%Y-%m-%d")
        return send_file(
            buf, as_attachment=True, download_name=f"Personal_{hoy}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return app


app = create_app()

if __name__ == "__main__":
    app.run(debug=True)
