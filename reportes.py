# -*- coding: utf-8 -*-
"""Blueprint de reportes nuevos para supervisores (2026-08-22, lista de
ideas de Davor) -- Horas semanales (#2) primero; Alertas (#1) y Ficha del
trabajador (#6) se agregan en los siguientes pasos del mismo plan."""
import calendar
import datetime as dt
import io
import re

import openpyxl
import pandas as pd
from flask import Blueprint, flash, redirect, request, render_template, send_file, url_for
from flask_login import current_user, login_required
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import aliased

from alertas import alertas_periodo, SALIDA_ANTICIPADA_MIN
from asistencia import _cargar_reporte, _estado_base, _fecha_mas_reciente_con_datos, _homologar_motivo, historico_persona
from cobertura import _cargar_visitas, marcaciones_del_dia, matriz_cobertura
from dimension_models import HistorialCambio, Persona, get_session
from excel_safety import fila_segura
from fact_models import ClasificacionDiaria
from historial import CAMPOS_VALIDOS, DIAS_SEMANA as DIAS_SEMANA_HISTORIAL
from horas_semanales import semana_iso, calcular_detalle_semana, resumen_por_persona
from permisos import requiere_pagina
from proyecciones import estacionalidad_faltas, necesidad_contratacion, ranking_proxima_falta, score_riesgo_rotacion, tasa_rotacion_por_supervisor
from recomendaciones import insights_equipo, resumen_perfil_equipo
from scoping import CANALES_FILTRABLES, condicion_scope, overrides_supervisor_canal
from vacaciones import calcular_viajes_vacaciones

bp = Blueprint("reportes", __name__, url_prefix="/reportes")

DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

# Mismo texto corto que ya usa Asistencia diaria (asistencia.html) para el
# badge de estado -- Davor pidió que la Ficha del trabajador "jale el mismo
# formato" en vez de mostrar el estado crudo ("ASISTIÓ A TIEMPO", 2026-08-24).
ESTADO_CORTO = {"ASISTIÓ A TIEMPO": "Asistió", "TARDANZA": "Tardanza", "FALTA": "Falta", "VACANTE": "Vacante", "VACACIONES": "Vacaciones"}

# Mismo vocabulario que ya usa la vista "Tareo" de One Page (F/PC/T/V/DM/A/—)
# -- Davor, 2026-08-26: "cuando entro a ver el personal... me deberia salir
# como un tareo... algo asi como la vista de One Page".
def _codigo_tareo(estado_base, comentario):
    comentario_norm = str(comentario).lower() if pd.notna(comentario) else ""
    if estado_base == "FALTA":
        if "descanso" in comentario_norm and "medic" in comentario_norm:
            return "DM", "info"
        return "F", "bad"
    if estado_base == "TARDANZA":
        return "T", "warn"
    if estado_base == "VACANTE":
        return "PC", "muted"
    if estado_base == "VACACIONES":
        return "V", "info"
    if estado_base == "ASISTIÓ A TIEMPO":
        return "A", "ok"
    return "?", "muted"

COLUMNAS_HORAS = [
    ("nombre", "Nombre"), ("supervisor", "Supervisor"), ("ciudad", "Ciudad"), ("region", "Región"),
    ("dias_falta_vacante", "Días Falta/Vacante"), ("horas_trabajadas", "Horas trabajadas"),
    ("horas_a_trabajar", "Horas a trabajar"), ("horas_a_trabajar_sin_faltas", "Horas a trabajar sin faltas"),
    ("diferencia_h", "Diferencia (h)"), ("pct_cumplimiento", "% Cumplimiento"),
    ("pct_cumplimiento_sin_faltas", "% Cumplimiento sin faltas"),
]


def _filtros_admin():
    """Filtros de Supervisor/Región/Rol/Ciudad para las 4 pestañas de
    Reportes -- admin y analista (Davor, 2026-08-24: "solo debe haber
    filtros para el perfil admin, para los supervisores no debe haber
    filtros"; 2026-08-25, unificado con Marcar asistencia: "solo para el
    analista y admin, para supervisor no" + "agrega tambien filtros de
    ciudad"). Para supervisor devuelve todo vacío -- las plantillas ocultan
    la barra de filtros cuando `roles_disponibles` (etc.) viene vacío. Los
    desplegables se acotan con condicion_scope() igual que la data real --
    sin esto, un analista vería nombres de supervisores/regiones de OTROS
    clientes en el desplegable aunque elegirlos no devolviera nada.
    Devuelve (filtro_args, roles_disponibles, regiones_disponibles,
    supervisores_disponibles, ciudades_disponibles, canales_disponibles).

    `canal` (Davor, 2026-08-29) -- SOLO para admin: "en mi caso que soy
    admin, debo tener un filtro para ver Tradicional, Farmacia y AU". Un
    analista de canal ya está acotado a su canal_asignado por
    condicion_scope(), no tiene sentido que además elija -- por eso
    `canales_disponibles` viene vacío para cualquiera que no sea admin
    (mismo criterio que ya oculta la barra entera para supervisor)."""
    if current_user.rol == "supervisor":
        return {"rol": "", "region": "", "supervisor": "", "ciudad": "", "canal": ""}, [], [], [], [], []

    es_admin = current_user.rol == "admin"
    filtro_args = {
        "rol": request.args.get("rol") or "",
        "region": request.args.get("region") or "",
        "supervisor": request.args.get("supervisor") or "",
        "ciudad": request.args.get("ciudad") or "",
        "canal": (request.args.get("canal") or "") if es_admin else "",
    }
    session = get_session()
    try:
        cond_scope = condicion_scope(Persona, current_user)

        def _valores(columna):
            q = session.query(columna).filter(columna.isnot(None))
            if cond_scope is not None:
                q = q.filter(cond_scope)
            return sorted({v for (v,) in q.distinct().all() if v})

        roles_disponibles = _valores(Persona.rol)
        regiones_disponibles = _valores(Persona.region)
        ciudades_disponibles = _valores(Persona.ciudad)

        SupervisorPersona = aliased(Persona)
        q_sup = (
            session.query(Persona.supervisor_dni, SupervisorPersona.nombre_completo)
            .join(SupervisorPersona, SupervisorPersona.dni == Persona.supervisor_dni)
            .filter(Persona.supervisor_dni.isnot(None))
        )
        if cond_scope is not None:
            q_sup = q_sup.filter(cond_scope)
        supervisores_disponibles = sorted(set(q_sup.distinct().all()), key=lambda t: (t[1] or "").title())
    finally:
        session.close()
    canales_disponibles = CANALES_FILTRABLES if es_admin else []
    return filtro_args, roles_disponibles, regiones_disponibles, supervisores_disponibles, ciudades_disponibles, canales_disponibles


def _semana_desde_query():
    """Lee ?semana=2026-W34 (formato de <input type=week>) -- si falta o es
    inválido, usa la semana ISO actual."""
    hoy = dt.date.today()
    semana_str = request.args.get("semana", "")
    if len(semana_str) == 8 and semana_str[4] == "-" and semana_str[5] == "W":
        try:
            anio, num = int(semana_str[:4]), int(semana_str[6:])
            desde, hasta = semana_iso(anio, num)
            return desde, hasta, semana_str
        except ValueError:
            pass
    anio, num, _ = hoy.isocalendar()
    desde, hasta = semana_iso(anio, num)
    return desde, hasta, f"{anio}-W{num:02d}"


@bp.get("/horas")
@requiere_pagina("reportes_horas")
def horas():
    desde, hasta, semana_str = _semana_desde_query()
    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    detalle = calcular_detalle_semana(
        desde, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    resumen = resumen_por_persona(detalle)
    filas = resumen.to_dict("records") if len(resumen) else []
    return render_template(
        "reportes_horas.html", usuario=current_user, activo="horas",
        semana_str=semana_str, desde=desde, hasta=hasta, filas=filas,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


@bp.get("/horas/exportar")
@requiere_pagina("reportes_horas")
def horas_exportar():
    desde, hasta, semana_str = _semana_desde_query()
    filtro_args, _roles_disp, _regiones_disp, _supervisores_disp, _ciudades_disp, _canales_disp = _filtros_admin()
    detalle = calcular_detalle_semana(
        desde, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    resumen = resumen_por_persona(detalle)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas semanales"
    ws.append([titulo for _clave, titulo in COLUMNAS_HORAS])
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for fila in resumen.to_dict("records"):
        ws.append(fila_segura([fila.get(clave) for clave, _titulo in COLUMNAS_HORAS]))
    for i, (_clave, titulo) in enumerate(COLUMNAS_HORAS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(titulo) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=f"Horas_Semanales_{semana_str}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _mes_desde_query():
    """Lee ?mes=2026-08 -- si falta o es inválido, usa el mes calendario actual."""
    hoy = dt.date.today()
    mes_str = request.args.get("mes", "")
    if re.match(r"^\d{4}-\d{2}$", mes_str):
        anio, mes = int(mes_str[:4]), int(mes_str[5:7])
    else:
        anio, mes = hoy.year, hoy.month
    desde = dt.date(anio, mes, 1)
    hasta = dt.date(anio, mes, calendar.monthrange(anio, mes)[1])
    return desde, hasta, f"{anio}-{mes:02d}"


def _dia_foco_desde_query():
    """Día de foco de Alertas (?dia=YYYY-MM-DD) -- sin el parámetro (primera
    carga) default a hoy; presente pero vacío ("Ver todo el mes" lo dejó
    así) = sin foco (None); inválido = hoy."""
    if "dia" not in request.args:
        return dt.date.today()
    dia_str = request.args.get("dia", "")
    if not dia_str:
        return None
    try:
        return dt.date.fromisoformat(dia_str)
    except ValueError:
        return dt.date.today()


@bp.get("/alertas")
@requiere_pagina("reportes_alertas")
def alertas():
    desde, hasta, mes_str = _mes_desde_query()
    dia_foco = _dia_foco_desde_query()
    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    lista = alertas_periodo(
        desde, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    return render_template(
        "reportes_alertas.html", usuario=current_user, activo="alertas",
        mes_str=mes_str, desde=desde, hasta=hasta, alertas=lista, dia_foco=dia_foco,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


def _rango_mes_actual_por_defecto():
    """Como _rango_desde_query() (Cobertura), pero el default de `desde` es
    el día 1 del mes de `hasta` en vez de una cantidad fija de días atrás
    -- Reportes > Desempeño mostraba el mes calendario completo sin poder
    acotarlo (Davor, 2026-08-26: "que sea calendario, no mes completo, que
    pueda filtrar hasta ciertos dias"). Con ?desde=&hasta= se puede pedir
    cualquier rango de días; sin params, sigue viéndose el mes en curso
    hasta hoy (mismo comportamiento de antes)."""
    hoy = dt.date.today()
    try:
        hasta = dt.date.fromisoformat(request.args["hasta"]) if request.args.get("hasta") else hoy
    except ValueError:
        hasta = hoy
    try:
        desde = dt.date.fromisoformat(request.args["desde"]) if request.args.get("desde") else hasta.replace(day=1)
    except ValueError:
        desde = hasta.replace(day=1)
    if desde > hasta:
        desde, hasta = hasta, desde
    return desde, hasta


@bp.get("/recomendaciones")
@requiere_pagina("reportes_recomendaciones")
def recomendaciones():
    desde, hasta = _rango_mes_actual_por_defecto()
    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    # calcular_detalle_semana() es lo mas caro de este reporte (~5s incluso
    # con equipos chicos) -- insights_equipo() y resumen_perfil_equipo() lo
    # llamaban cada uno por su cuenta con rangos que se superponen casi
    # del todo, duplicando esa consulta y sumando por encima del timeout
    # del worker en Render (Davor, 2026-09-01: 503 al filtrar por
    # supervisor). Se trae UNA vez, con el rango mas amplio de los dos
    # (el de insights_equipo, que mira 4 semanas antes de "hasta"), y se
    # reusa en ambas funciones.
    anio_actual, num_actual, _ = hasta.isocalendar()
    inicio_actual, _ = semana_iso(anio_actual, num_actual)
    desde_semanas = inicio_actual - dt.timedelta(weeks=4)
    detalle = calcular_detalle_semana(
        desde_semanas, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    lista = insights_equipo(
        current_user, desde=desde, hasta=hasta, detalle=detalle,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    perfiles = resumen_perfil_equipo(
        current_user, desde=desde, hasta=hasta, detalle=detalle,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    return render_template(
        "reportes_recomendaciones.html", usuario=current_user, activo="recomendaciones",
        insights=lista, perfiles=perfiles, desde=desde, hasta=hasta,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


@bp.get("/proyecciones")
@requiere_pagina("reportes_proyecciones")
def proyecciones():
    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    kwargs_filtro = dict(
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    # Cada pieza se calcula UNA sola vez y se pasa a las funciones que la
    # necesitan (tasas_sup/señales/riesgo/estacionalidad_equipo) -- sin
    # esto insights_equipo() (lo más pesado, llama a calcular_detalle_semana())
    # se recalculaba 2 veces y tasa_rotacion_por_supervisor() 3 veces por
    # request, la misma duplicación que ya tumbó Desempeño con 503 en
    # Render (Davor, 2026-09-01: 18.5s la primera versión de esta pantalla).
    estacionalidad = estacionalidad_faltas(current_user, **kwargs_filtro)
    tasas_sup = tasa_rotacion_por_supervisor(current_user, **kwargs_filtro)
    señales = insights_equipo(current_user, **kwargs_filtro)
    riesgo = score_riesgo_rotacion(current_user, tasas_sup=tasas_sup, señales=señales, **kwargs_filtro)
    contratacion = necesidad_contratacion(current_user, tasas_sup=tasas_sup, **kwargs_filtro)
    proxima_falta = ranking_proxima_falta(current_user, riesgo=riesgo, estacionalidad_equipo=estacionalidad, **kwargs_filtro)
    return render_template(
        "reportes_proyecciones.html", usuario=current_user, activo="proyecciones",
        estacionalidad=estacionalidad, riesgo=riesgo, contratacion=contratacion, proxima_falta=proxima_falta,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


def _rango_desde_query(dias_por_defecto=21):
    """Lee ?desde=&hasta= -- si faltan o son inválidos, usa los últimos
    `dias_por_defecto` días hasta hoy (mismo patrón que el Dashboard)."""
    hoy = dt.date.today()
    try:
        hasta = dt.date.fromisoformat(request.args["hasta"]) if request.args.get("hasta") else hoy
    except ValueError:
        hasta = hoy
    try:
        desde = dt.date.fromisoformat(request.args["desde"]) if request.args.get("desde") else hasta - dt.timedelta(days=dias_por_defecto)
    except ValueError:
        desde = hasta - dt.timedelta(days=dias_por_defecto)
    if desde > hasta:
        desde, hasta = hasta, desde
    return desde, hasta


@bp.get("/cobertura")
@requiere_pagina("reportes_cobertura")
def cobertura():
    desde, hasta = _rango_desde_query()
    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    personas, fechas, celdas, categorias = matriz_cobertura(
        desde, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )

    # Resaltado simple: celda por debajo de la mitad del promedio de ESA
    # persona en el periodo elegido -- mismo criterio "umbral simple" del
    # plan, no intenta reproducir el cruce con Falta/Vacaciones del mockup.
    promedios = {}
    for p in personas:
        valores = [celdas[(p["dni"], f)] for f in fechas if (p["dni"], f) in celdas]
        promedios[p["dni"]] = (sum(valores) / len(valores)) if valores else 0

    return render_template(
        "reportes_cobertura.html", usuario=current_user, activo="cobertura",
        desde=desde, hasta=hasta, personas=personas, fechas=fechas, celdas=celdas,
        categorias=categorias, promedios=promedios,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


@bp.get("/cobertura/exportar")
@requiere_pagina("reportes_cobertura")
def cobertura_exportar():
    desde, hasta = _rango_desde_query()
    filtro_args, _roles_disp, _regiones_disp, _supervisores_disp, _ciudades_disp, _canales_disp = _filtros_admin()
    personas, fechas, celdas, _categorias = matriz_cobertura(
        desde, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobertura Diaria"
    encabezado = ["DNI", "Nombre", "Ciudad", "Supervisor"] + [f.strftime("%d/%m") for f in fechas]
    ws.append(encabezado)
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for p in personas:
        fila = [p["dni"], p["nombre"], p["ciudad"], p["supervisor"]]
        fila += [celdas.get((p["dni"], f), "") for f in fechas]
        ws.append(fila_segura(fila))
    for i in range(1, 5):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=f"Cobertura_Diaria_{desde}_a_{hasta}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _fecha_desde_query():
    """Lee ?fecha= (formato ISO) -- si falta o es inválida, usa hoy. Mismo
    patrón que asistencia.marcar_vista() (vista de un solo día, no un rango)."""
    fecha_str = request.args.get("fecha") or dt.date.today().isoformat()
    try:
        return dt.date.fromisoformat(fecha_str), fecha_str
    except ValueError:
        hoy = dt.date.today()
        return hoy, hoy.isoformat()


@bp.get("/marcaciones")
@requiere_pagina("reportes_marcaciones")
def marcaciones():
    fecha, fecha_str = _fecha_desde_query()

    # "Ver ruta del día" desde Cobertura (clic en el número) -- acota a una
    # sola persona. Mismo criterio de acceso que reportes.ficha(): validar
    # el scope acá, no solo confiar en dni_filtro de marcaciones_del_dia()
    # (que en _cargar_visitas se salta condicion_scope cuando hay dni_filtro
    # -- sin este chequeo cualquier usuario logueado podría ver la ruta de
    # cualquier DNI armando la URL a mano).
    dni_foco = request.args.get("dni") or None
    nombre_foco = None
    if dni_foco:
        session = get_session()
        try:
            query = session.query(Persona).filter(Persona.dni == dni_foco)
            cond_scope = condicion_scope(Persona, current_user)
            if cond_scope is not None:
                query = query.filter(cond_scope)
            persona_foco = query.first()
        finally:
            session.close()
        if persona_foco:
            nombre_foco = persona_foco.nombre_completo
        else:
            flash("No se encontró a esa persona o no tenés acceso a su ruta.", "error")
            dni_foco = None

    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    personas = marcaciones_del_dia(
        fecha, current_user, dni_filtro=dni_foco,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    return render_template(
        "reportes_marcaciones.html", usuario=current_user, activo="marcaciones",
        fecha=fecha, fecha_str=fecha_str, personas=personas, dni_foco=dni_foco, nombre_foco=nombre_foco,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


LEYENDA_TAREO = [
    ("A", "ok", "Asistió a tiempo"), ("T", "warn", "Tardanza"), ("F", "bad", "Falta"),
    ("DM", "info", "Descanso médico"), ("V", "info", "Vacaciones"), ("PC", "muted", "Posición por cubrir"),
    ("—", "muted", "Sin marcación (no le tocaba trabajar)"),
]


def _matriz_tareo(desde, hasta, filtro_args):
    """Detalle día-persona del rango (reusa calcular_detalle_semana, mismo
    scope/filtros que el resto de Reportes) reducido a un código de tareo
    por (dni, fecha) -- mismo vocabulario que reportes.ficha() ya usa para
    "Tareo del mes" de UNA persona (Davor, 2026-08-28: "cuando entro a ver
    el personal... que sea como un tareo", ahora para TODO el equipo con
    los mismos filtros que ya tiene el resto de Reportes."""
    detalle = calcular_detalle_semana(
        desde, hasta, current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    fechas, fecha_cursor = [], desde
    while fecha_cursor <= hasta:
        fechas.append(fecha_cursor)
        fecha_cursor += dt.timedelta(days=1)

    personas, celdas = [], {}
    if len(detalle):
        for dni, grupo in detalle.groupby("dni"):
            personas.append({
                "dni": dni, "nombre": grupo["nombre"].iloc[0], "supervisor": grupo["supervisor"].iloc[0],
                "ciudad": grupo["ciudad"].iloc[0], "region": grupo["region"].iloc[0],
            })
            for _, row in grupo.iterrows():
                codigo, clase = _codigo_tareo(row["estado_base"], row["comentario"])
                celdas[(dni, row["fecha"].date())] = {"codigo": codigo, "clase": clase, "titulo": row["estado"]}
        personas.sort(key=lambda p: (p["nombre"] or "").title())
    return personas, fechas, celdas


@bp.get("/tareo")
@requiere_pagina("reportes_tareo")
def tareo():
    desde, hasta, mes_str = _mes_desde_query()
    filtro_args, roles_disp, regiones_disp, supervisores_disp, ciudades_disp, canales_disp = _filtros_admin()
    personas, fechas, celdas = _matriz_tareo(desde, hasta, filtro_args)
    return render_template(
        "reportes_tareo.html", usuario=current_user, activo="tareo",
        mes_str=mes_str, desde=desde, hasta=hasta, personas=personas, fechas=fechas, celdas=celdas,
        leyenda_tareo=LEYENDA_TAREO,
        filtro_args=filtro_args, roles_disponibles=roles_disp,
        regiones_disponibles=regiones_disp, supervisores_disponibles=supervisores_disp,
        ciudades_disponibles=ciudades_disp, canales_disponibles=canales_disp,
    )


# Mismos colores que .cod.<clase> en reportes_tareo.html -- Davor, 2026-08-28:
# "que el exportar salgo con... los colores de celdas, letras centradas y
# visuales" (el Excel salía en texto plano, sin nada del color/vocabulario
# visual que ya tiene la vista web).
_COLOR_TAREO = {
    "ok": ("DCFCE7", "16A34A"), "warn": ("FEF3C7", "D97706"), "bad": ("FEE2E2", "DC2626"),
    "info": ("DBEAFE", "2563EB"), "muted": ("D7E3EF", "5B7186"),
}


@bp.get("/tareo/exportar")
@requiere_pagina("reportes_tareo")
def tareo_exportar():
    desde, hasta, mes_str = _mes_desde_query()
    filtro_args, _roles_disp, _regiones_disp, _supervisores_disp, _ciudades_disp, _canales_disp = _filtros_admin()
    personas, fechas, celdas = _matriz_tareo(desde, hasta, filtro_args)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareo"
    ws.sheet_view.zoomScale = 80
    ws.freeze_panes = "E2"
    centrado = Alignment(horizontal="center", vertical="center")

    encabezado = ["DNI", "Nombre", "Ciudad", "Supervisor"] + [f.strftime("%d/%m") for f in fechas]
    ws.append(encabezado)
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.alignment = centrado
    N_COLS_FIJAS = 4
    for p in personas:
        fila = [p["dni"], p["nombre"], p["ciudad"], p["supervisor"]]
        fila += [celdas.get((p["dni"], f), {}).get("codigo", "—") for f in fechas]
        ws.append(fila_segura(fila))
        fila_excel = ws.max_row
        for i, f in enumerate(fechas):
            info = celdas.get((p["dni"], f))
            clase = info["clase"] if info else "muted"
            fill_hex, font_hex = _COLOR_TAREO[clase]
            celda = ws.cell(row=fila_excel, column=N_COLS_FIJAS + 1 + i)
            celda.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            celda.font = Font(bold=True, color=font_hex)
            celda.alignment = centrado
    for i in range(1, 5):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=f"Tareo_{mes_str}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/ficha/<dni>")
@login_required
def ficha(dni):
    # Mismo criterio de acceso que el resto del sitio: si la Persona no
    # entra en el scope de current_user (su equipo/cliente), 404 lógico en
    # vez de dejar entrar por URL directa a alguien fuera de su alcance.
    session = get_session()
    try:
        query = session.query(Persona).filter(Persona.dni == dni)
        cond_scope = condicion_scope(Persona, current_user)
        if cond_scope is not None:
            query = query.filter(cond_scope)
        persona = query.first()

        nombre_supervisor = None
        if persona:
            # Override de supervisor por canal (Davor, 2026-08-29) -- ver
            # scoping.overrides_supervisor_canal().
            overrides_sup = overrides_supervisor_canal(session, [dni], current_user)
            supervisor_dni_efectivo = overrides_sup.get(dni, persona.supervisor_dni)
            if supervisor_dni_efectivo:
                sup = session.query(Persona.nombre_completo).filter(Persona.dni == supervisor_dni_efectivo).first()
                nombre_supervisor = sup[0] if sup else None
    finally:
        session.close()

    if not persona:
        flash("No se encontró a esa persona o no tenés acceso a su ficha.", "error")
        return redirect(url_for("reportes.horas"))

    hoy = dt.date.today()
    mes_desde = hoy.replace(day=1)
    mes_hasta = dt.date(hoy.year, hoy.month, calendar.monthrange(hoy.year, hoy.month)[1])
    ventana_desde = hoy - dt.timedelta(days=180)

    # Indicador de asistencia del mes -- misma regla de horas ya validada
    # (Horas semanales), aplicada al rango de un mes completo en vez de una
    # semana, para UNA sola persona.
    detalle_mes = calcular_detalle_semana(mes_desde, mes_hasta, current_user, dni_filtro=dni)
    resumen_mes = resumen_por_persona(detalle_mes)
    indicador_mes = resumen_mes.iloc[0].to_dict() if len(resumen_mes) else None

    # Tareo del mes -- calendario compacto día por día (mismo estilo que la
    # vista "Tareo" de One Page), reusando detalle_mes (ya calculado arriba,
    # no hace falta una consulta nueva). Un día sin fila (domingo, feriado,
    # antes del ingreso, o todavía no procesado) se muestra como "—".
    por_fecha = {row["fecha"].date(): row for _, row in detalle_mes.iterrows()} if len(detalle_mes) else {}
    tareo_mes = []
    fecha_cursor = mes_desde
    while fecha_cursor <= mes_hasta:
        row = por_fecha.get(fecha_cursor)
        if row is None:
            codigo, clase, titulo = "—", "muted", "Sin marcación (no le tocaba trabajar)"
        else:
            estado_base = row["estado"].split(" (")[0]
            codigo, clase = _codigo_tareo(estado_base, row["comentario"])
            titulo = row["estado"]
        tareo_mes.append({
            "fecha": fecha_cursor, "dia": fecha_cursor.day, "dia_semana": DIAS_ES[fecha_cursor.weekday()][:2],
            "codigo": codigo, "clase": clase, "titulo": titulo,
        })
        fecha_cursor += dt.timedelta(days=1)

    # Vacaciones (últimos 180 días) -- reusa el mismo agrupador en "viajes"
    # que ya usa el Dashboard.
    session = get_session()
    try:
        filas_periodo = (
            session.query(ClasificacionDiaria.dni, Persona.nombre_completo, ClasificacionDiaria.fecha, ClasificacionDiaria.estado)
            .join(Persona, Persona.dni == ClasificacionDiaria.dni)
            .filter(ClasificacionDiaria.dni == dni, ClasificacionDiaria.fecha >= ventana_desde, ClasificacionDiaria.fecha <= hoy)
            .all()
        )
        r_periodo = pd.DataFrame(filas_periodo, columns=["dni", "nombre", "fecha", "estado"])
        if len(r_periodo):
            r_periodo["fecha"] = pd.to_datetime(r_periodo["fecha"])
            r_periodo["estado_base"] = r_periodo["estado"].apply(lambda s: s.split(" (")[0])
        viajes_vacaciones = calcular_viajes_vacaciones(session, r_periodo, hoy) if len(r_periodo) else []

        # Descansos médicos (últimos 180 días) -- subconjunto de Falta cuyo
        # comentario del supervisor menciona "descanso médico".
        descansos = (
            session.query(ClasificacionDiaria.fecha, ClasificacionDiaria.comentario_supervisor)
            .filter(
                ClasificacionDiaria.dni == dni,
                ClasificacionDiaria.fecha >= ventana_desde, ClasificacionDiaria.fecha <= hoy,
                ClasificacionDiaria.estado.ilike("FALTA%"),
                ClasificacionDiaria.comentario_supervisor.ilike("%descanso%medic%"),
            )
            .order_by(ClasificacionDiaria.fecha.desc())
            .all()
        )

        # Historial de cambios (overrides de horario/canal/supervisor/zona/
        # refrigerio) vigentes para esta persona -- ya se validó el scope de
        # `persona` arriba, no hace falta repetirlo acá.
        historial_persona = (
            session.query(HistorialCambio)
            .filter(HistorialCambio.dni == dni)
            .order_by(HistorialCambio.fecha_desde.desc())
            .all()
        )
    finally:
        session.close()

    # Cumplimiento semanal, últimas 6 semanas ISO (incluye la actual) --
    # mismo cálculo de Horas semanales, una corrida por semana. Cada semana
    # queda linkeable (?semana=) para ver su detalle día a día abajo.
    anio_actual, num_actual, _ = hoy.isocalendar()
    cumplimiento_semanal = []
    for i in range(5, -1, -1):
        fecha_ref = hoy - dt.timedelta(weeks=i)
        anio_s, num_s, _ = fecha_ref.isocalendar()
        desde_s, hasta_s = semana_iso(anio_s, num_s)
        detalle_s = calcular_detalle_semana(desde_s, hasta_s, current_user, dni_filtro=dni)
        resumen_s = resumen_por_persona(detalle_s)
        pct = resumen_s.iloc[0]["pct_cumplimiento_sin_faltas"] if len(resumen_s) else None
        cumplimiento_semanal.append({
            "semana": f"S{num_s}", "semana_str": f"{anio_s}-W{num_s:02d}",
            "desde": desde_s, "pct": None if pct is None or pd.isna(pct) else pct,
        })

    # Tendencia de cobertura, últimas 6 semanas, día por día -- mismo
    # cálculo que Reportes > Cobertura (PDVs distintos por día, Punto Censo
    # aparte, ver cobertura.matriz_cobertura) y mismos símbolos T/A/F/C de
    # ahí, pero en tira diaria como "Tareo del mes" en vez de barras
    # semanales (Davor, 2026-08-29: "que sea asi como el tareo, de forma
    # diaria... la cantidad de PDVs visitados, con su simbolo").
    cobertura_desde = hoy - dt.timedelta(days=41)
    v_cob = _cargar_visitas(cobertura_desde, hoy, None, dni_filtro=dni)
    if len(v_cob):
        v_normal = v_cob[v_cob["geofence_ok"] & ~v_cob["es_censo"]]
        total_normal = v_normal.groupby("fecha_inicio")["punto_venta_id"].nunique()
        v_censo = v_cob[v_cob["es_censo"]]
        total_censo = v_censo.groupby("fecha_inicio").size()
        total_por_dia = total_normal.add(total_censo, fill_value=0)
        dias_au = set(v_normal[v_normal["tipo_negocio"] == "AUTOSERVICIOS"]["fecha_inicio"].dt.date)
        dias_farma = set(v_normal[v_normal["tipo_negocio"] == "FARMACIA"]["fecha_inicio"].dt.date)
        dias_tradicional = set(v_normal[v_normal["tipo_negocio"] == "TRADICIONAL"]["fecha_inicio"].dt.date)
        dias_censo = set(total_censo.index.date) if len(total_censo) else set()
    else:
        total_por_dia = pd.Series(dtype=float)
        dias_au = dias_farma = dias_tradicional = dias_censo = set()

    valores_no_cero = [int(v) for v in total_por_dia if v > 0]
    prom_cobertura = (sum(valores_no_cero) / len(valores_no_cero)) if valores_no_cero else 0

    tendencia_cobertura = []
    fecha_cursor = cobertura_desde
    while fecha_cursor <= hoy:
        total = int(total_por_dia.get(pd.Timestamp(fecha_cursor), 0))
        tags = []
        if fecha_cursor in dias_tradicional:
            tags.append("tradicional")
        if fecha_cursor in dias_au:
            tags.append("au")
        if fecha_cursor in dias_farma:
            tags.append("farma")
        if fecha_cursor in dias_censo:
            tags.append("censo")
        clase = "muted" if total == 0 else (
            "bad" if prom_cobertura > 0 and total < prom_cobertura * 0.5 else "ok"
        )
        tendencia_cobertura.append({
            "fecha": fecha_cursor, "dia": fecha_cursor.day, "dia_semana": DIAS_ES[fecha_cursor.weekday()][:2],
            "total": total, "clase": clase, "tags": tags,
        })
        fecha_cursor += dt.timedelta(days=1)

    # Detalle día a día -- para ver DÓNDE está el cuello de botella detrás
    # de un % bajo (Davor, 2026-08-23), no solo el número agregado. Semana
    # seleccionable igual que Horas semanales (?semana=2026-W34); si no se
    # pasa, la última semana con datos reales (no necesariamente la
    # calendario actual, que puede no tener nada cargado todavía).
    semana_str_qs = request.args.get("semana", "")
    if len(semana_str_qs) == 8 and semana_str_qs[4] == "-" and semana_str_qs[5] == "W":
        try:
            anio_v, num_v = int(semana_str_qs[:4]), int(semana_str_qs[6:])
        except ValueError:
            anio_v, num_v = anio_actual, num_actual
    else:
        anio_v, num_v = anio_actual, num_actual
    desde_v, hasta_v = semana_iso(anio_v, num_v)
    detalle_semana_vista = calcular_detalle_semana(desde_v, hasta_v, current_user, dni_filtro=dni)
    detalle_dias = []
    if len(detalle_semana_vista):
        for _, row in detalle_semana_vista.sort_values("fecha").iterrows():
            estado_base = row["estado"].split(" (")[0]
            detalle_dias.append({
                "fecha": row["fecha"].strftime("%d/%m"),
                "dia_semana": DIAS_ES[row["fecha"].weekday()],
                "estado": row["estado"],
                "estado_base": estado_base,
                "estado_corto": ESTADO_CORTO.get(estado_base, estado_base.title()),
                "motivo": _homologar_motivo(row["comentario"]) if estado_base == "FALTA" else None,
                "entrada_real": row["entrada_real"] or "—",
                "entrada_esperada": row["entrada_esperada"] or "—",
                "salida_real": row["salida_real"] or "—",
                "salida_esperada": row["salida_esperada"] or "—",
                "horas_trabajadas": row["horas_trabajadas"] if pd.notna(row["horas_trabajadas"]) else None,
                "horas_a_trabajar": row["horas_a_trabajar"],
            })
    semana_vista_str = f"{anio_v}-W{num_v:02d}"

    # Tardanzas Y salidas anticipadas del mes, combinadas en una sola tabla
    # con las 4 horas (entrada/salida real y programada) -- Davor,
    # 2026-08-26: "deberia salir detalle de las salidas anticipadas quiza
    # junto con tardanzas... Detalle entrada, entrada programada, salida,
    # salida programada". salida_anticipada_min no viene en detalle_mes
    # (calcular_detalle_semana no la trae), se consulta aparte -- mismo
    # patrón que resumen_perfil_equipo() en recomendaciones.py.
    session = get_session()
    try:
        salida_anticipada_por_fecha = dict(
            session.query(ClasificacionDiaria.fecha, ClasificacionDiaria.salida_anticipada_min)
            .filter(ClasificacionDiaria.dni == dni, ClasificacionDiaria.fecha >= mes_desde, ClasificacionDiaria.fecha <= mes_hasta)
            .all()
        )
    finally:
        session.close()

    tardanzas_mes = []
    if len(detalle_mes):
        for _, row in detalle_mes.sort_values("fecha", ascending=False).iterrows():
            estado_base = row["estado"].split(" (")[0]
            es_tardanza = estado_base == "TARDANZA"
            sal_ant = salida_anticipada_por_fecha.get(row["fecha"].date())
            es_salida_temprana = sal_ant is not None and sal_ant > SALIDA_ANTICIPADA_MIN
            if not es_tardanza and not es_salida_temprana:
                continue
            tardanzas_mes.append({
                "fecha_iso": row["fecha"].date().isoformat(),
                "fecha": row["fecha"].strftime("%d/%m"),
                "estado": row["estado"],
                "estado_corto": ESTADO_CORTO.get(estado_base, estado_base.title()),
                "tardanza": es_tardanza, "salida_temprana": es_salida_temprana,
                "entrada_real": row["entrada_real"] or "—", "entrada_esperada": row["entrada_esperada"] or "—",
                "salida_real": row["salida_real"] or "—", "salida_esperada": row["salida_esperada"] or "—",
            })

    # Faltas del mes -- mismo criterio que Tardanzas del mes (Davor,
    # 2026-08-24: "tiene una falta pero no me sale en el detalle del
    # perfil" -- el indicador de arriba ya contaba la falta, pero "Detalle
    # día a día" solo muestra la semana seleccionada, y esta era la única
    # sección que le faltaba su propia lista explícita). Cuenta TODAS las
    # faltas del mes (con o sin sustento), igual que indicador_mes.dias_falta_vacante.
    faltas_mes = []
    if len(detalle_mes):
        for _, row in detalle_mes[detalle_mes["estado"].str.startswith("FALTA")].sort_values("fecha", ascending=False).iterrows():
            faltas_mes.append({
                "fecha": row["fecha"].strftime("%d/%m"),
                "motivo": _homologar_motivo(row["comentario"]) or "Sin motivo",
            })

    alertas_mes = alertas_periodo(mes_desde, mes_hasta, None, dni_filtro=dni)
    insights = insights_equipo(None, dni_filtro=dni, hasta=hoy)

    return render_template(
        "reportes_ficha.html", usuario=current_user, activo="ficha",
        persona=persona, nombre_supervisor=nombre_supervisor,
        indicador_mes=indicador_mes, tareo_mes=tareo_mes, viajes_vacaciones=viajes_vacaciones,
        descansos=descansos, cumplimiento_semanal=cumplimiento_semanal, tendencia_cobertura=tendencia_cobertura,
        detalle_dias=detalle_dias, semana_vista_str=semana_vista_str,
        desde_v=desde_v, hasta_v=hasta_v, tardanzas_mes=tardanzas_mes, faltas_mes=faltas_mes,
        alertas_mes=alertas_mes, insights=insights,
        historial_persona=historial_persona, campos_historial=CAMPOS_VALIDOS, dias_semana_historial=DIAS_SEMANA_HISTORIAL,
    )


HISTORICO_MAX_DIAS_RANGO = 31  # ~1 mes -- cada dia es una consulta aparte (ver mas abajo), un rango gigante seria muy lento


@bp.get("/historico")
@requiere_pagina("reportes_historico")
def historico():
    """"Histórico de asistencia diaria" (Davor, 2026-08-29) -- copia de
    "Reporte diario" (asistencia.py) pero viviendo en Reportes, con dos
    diferencias a propósito: la Salida es del MISMO día que la entrada
    (ver _cargar_reporte(..., salida_mismo_dia=True) -- "Reporte diario"
    usa la de ayer porque se manda al cliente a media tarde, antes de que
    el turno de hoy cierre; acá se mira un día puntual del pasado, así que
    hace falta ver SU PROPIA salida), y se puede editar la Hora entrada/
    salida PROGRAMADA para ese día puntual con un boton chico ("solo un
    boton que salga editar y abajo se escriba la nueva hora, algo mas
    pequeño") -- postea directo a historial.crear(), mismo motor que ya
    usa "Historial de cambios" para forzar un horario distinto al Maestro/
    Patrón, sin duplicar esa lógica acá.

    Modo mercaderista (Davor, 2026-09-01: "debo poner un rango de fechas y
    me debe salir un filtro de mercaderista... me debe salir de forma
    diaria en modo horizontal cada día su detalle") -- si se elige una
    persona puntual, la vista cambia de "todos en un solo día" a "un día
    por fila para esa sola persona", en el rango elegido (tope
    HISTORICO_MAX_DIAS_RANGO). Reusa _cargar_reporte(dni_filtro=...) UNA
    VEZ POR DÍA -- cada llamada ya es barata (acotada a un solo DNI), pero
    son varias consultas por día, por eso el tope al rango."""
    fecha_str = request.args.get("fecha") or dt.date.today().isoformat()
    try:
        fecha = dt.date.fromisoformat(fecha_str)
    except ValueError:
        fecha = dt.date.today()
        fecha_str = fecha.isoformat()

    filtro_args, roles_disponibles, regiones_disponibles, supervisores_disponibles, ciudades_disponibles, canales_disponibles = _filtros_admin()
    solo_incidencias = request.args.get("solo_incidencias") == "1"

    # El filtro de mercaderista y el rango de fechas son solo para quien YA
    # tiene los demas filtros (admin/analista) -- mismo criterio que
    # _filtros_admin(). Davor, 2026-09-01: "falta las fechas de rango" --
    # desde/hasta se muestran SIEMPRE para admin/analista (no solo despues
    # de elegir mercaderista), asi se pueden elegir ambos en un solo paso.
    mercaderista_filtro = (request.args.get("mercaderista") or "") if roles_disponibles else ""
    mercaderistas_disponibles = []
    if roles_disponibles:
        session = get_session()
        try:
            cond_scope = condicion_scope(Persona, current_user)
            query_m = session.query(Persona.dni, Persona.nombre_completo)
            if cond_scope is not None:
                query_m = query_m.filter(cond_scope)
            mercaderistas_disponibles = sorted(
                ((dni, nombre) for dni, nombre in query_m.all()), key=lambda p: (p[1] or "").title(),
            )
        finally:
            session.close()

    if roles_disponibles:
        desde_str = request.args.get("desde") or fecha_str
        hasta_str = request.args.get("hasta") or fecha_str
        try:
            desde, hasta = dt.date.fromisoformat(desde_str), dt.date.fromisoformat(hasta_str)
        except ValueError:
            desde = hasta = fecha
        if hasta < desde:
            desde, hasta = hasta, desde
        if mercaderista_filtro and (hasta - desde).days > HISTORICO_MAX_DIAS_RANGO:
            desde = hasta - dt.timedelta(days=HISTORICO_MAX_DIAS_RANGO)
        desde_str, hasta_str = desde.isoformat(), hasta.isoformat()
    else:
        desde = hasta = fecha
        desde_str = hasta_str = fecha_str

    if mercaderista_filtro:
        filas, ultima_sync = historico_persona(mercaderista_filtro, desde, hasta, usuario_actual=current_user)
        hay_datos_del_dia = bool(filas)
        if solo_incidencias and filas:
            filas = [f for f in filas if _estado_base(f["estado"]) == "TARDANZA" or f["salida_temprana"]]
        mercaderista_nombre = dict(mercaderistas_disponibles).get(mercaderista_filtro, mercaderista_filtro)
        return render_template(
            "reportes_historico.html", usuario=current_user, activo="historico",
            modo_mercaderista=True, mercaderista_filtro=mercaderista_filtro, mercaderistas_disponibles=mercaderistas_disponibles,
            mercaderista_nombre=mercaderista_nombre,
            desde_str=desde_str, hasta_str=hasta_str,
            fecha_str=fecha_str, resumen=None, filas=filas, frescura={"ultima_sync": ultima_sync} if ultima_sync else None,
            fecha_reciente=None, hay_datos_del_dia=hay_datos_del_dia,
            filtro_args=filtro_args, roles_disponibles=roles_disponibles, regiones_disponibles=regiones_disponibles,
            supervisores_disponibles=supervisores_disponibles, ciudades_disponibles=ciudades_disponibles,
            canales_disponibles=canales_disponibles, solo_incidencias=solo_incidencias,
        )

    # Sin mercaderista elegido: "todos en un solo dia" -- para admin/analista
    # ese dia es `hasta` (el rango de fechas sigue visible en pantalla, pero
    # sin una persona puntual mostrar varios dias x todas las personas seria
    # una tabla gigante, no lo que se pidio); para supervisor sigue siendo
    # el selector simple de un dia (`fecha`, sin rango).
    dia_efectivo = hasta if roles_disponibles else fecha
    resumen, filas, frescura = _cargar_reporte(
        dia_efectivo, usuario_actual=current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
        salida_mismo_dia=True,
    )
    hay_datos_del_dia = bool(filas)
    # "Solo tardanzas y salidas antes de hora" (Davor, 2026-08-29) --
    # checkbox sobre las filas ya cargadas, no un filtro de _cargar_reporte()
    # (no acota qué se trae de la base, solo qué se muestra) -- el resumen
    # de arriba (asistio/tardanza/falta) sigue mostrando el total del día
    # completo, sin importar el check.
    if solo_incidencias and filas:
        filas = [f for f in filas if _estado_base(f["estado"]) == "TARDANZA" or f["salida_temprana"]]
    fecha_reciente = None if hay_datos_del_dia else _fecha_mas_reciente_con_datos()
    fecha_str_efectiva = dia_efectivo.isoformat()
    return render_template(
        "reportes_historico.html", usuario=current_user, activo="historico",
        modo_mercaderista=False, mercaderista_filtro=mercaderista_filtro, mercaderistas_disponibles=mercaderistas_disponibles,
        desde_str=desde_str, hasta_str=hasta_str,
        fecha_str=fecha_str_efectiva, resumen=resumen, filas=filas, frescura=frescura, fecha_reciente=fecha_reciente,
        hay_datos_del_dia=hay_datos_del_dia,
        filtro_args=filtro_args, roles_disponibles=roles_disponibles, regiones_disponibles=regiones_disponibles,
        supervisores_disponibles=supervisores_disponibles, ciudades_disponibles=ciudades_disponibles,
        canales_disponibles=canales_disponibles, solo_incidencias=solo_incidencias,
    )


@bp.get("/perfil")
@requiere_pagina("reportes_perfil")
def perfil():
    """Punto de entrada para abrir la Ficha de un mercaderista puntual sin
    tener que encontrarlo antes en otra lista y hacerle clic al nombre
    (Davor, 2026-08-29: "esta vista del mercaderista no hay una parte
    donde visualizarla, solo cuando damos click... deberia haber en
    reportes una pestaña... donde pongo el mercaderista y se activa toda
    esa ventana"). Mismo scope que el resto del sitio -- un supervisor solo
    ve su propio equipo en el desplegable, igual que ya le pasa en
    cualquier otra lista.

    `ciudad` (Davor, 2026-08-29: "agregale un filtro ciudad") -- acota el
    desplegable/buscador antes de tipear, útil cuando dos personas
    comparten nombre o hay muchas para una sola ciudad."""
    ciudad_filtro = request.args.get("ciudad") or ""
    session = get_session()
    try:
        cond_scope = condicion_scope(Persona, current_user)

        query_ciudades = session.query(Persona.ciudad).filter(Persona.ciudad.isnot(None))
        if cond_scope is not None:
            query_ciudades = query_ciudades.filter(cond_scope)
        ciudades_disponibles = sorted({c for (c,) in query_ciudades.distinct().all() if c})

        query = session.query(Persona.dni, Persona.nombre_completo)
        if cond_scope is not None:
            query = query.filter(cond_scope)
        if ciudad_filtro:
            query = query.filter(Persona.ciudad == ciudad_filtro)
        # Tupla plana (no el Row de SQLAlchemy que devuelve .all()) -- hace
        # falta para que |tojson la pueda serializar tal cual en el buscador
        # con sugerencias del template.
        personas_disponibles = sorted(
            ((dni, nombre) for dni, nombre in query.all()), key=lambda p: (p[1] or "").title(),
        )
    finally:
        session.close()
    return render_template(
        "reportes_perfil.html", usuario=current_user, activo="perfil",
        personas_disponibles=personas_disponibles,
        ciudad_filtro=ciudad_filtro, ciudades_disponibles=ciudades_disponibles,
    )
