# -*- coding: utf-8 -*-
"""
Fase 6: version portable de MAC/reporte_diario_9am.py -- las 5 lecturas de
clasificacion (Fase 3, ya migradas hoy) quedan igual. Cambios de
portabilidad: Athena/Maestro/Patron/Feriados/Tabla3 por los clientes
portables, imports relativos.

Cambio real de diseno: `escribir_excel()` ahora TAMBIEN sube el reporte por
Graph (ademas de guardarlo local, que sigue sirviendo para el paso siguiente
DENTRO de la misma corrida del workflow). Esto es necesario porque, a
diferencia de la laptop (el mismo disco vive entre corridas), un runner de
GitHub Actions es efimero -- sin subirlo, la SIGUIENTE corrida (que en el
ciclo real es aplicar_correcciones.py, buscando lo que el analista corrigio
a mano) jamas veria el archivo. Mientras dure la validacion en paralelo (ver
plan Fase 6), sube a una ruta de auditoria aparte, NO al
Reportes_Diarios/Reporte_Asistencia_*.xlsx real que ya lee/edita el
analista.
"""
import os
import re
import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dimension_models import get_session  # noqa: E402
from fact_models import ClasificacionDiaria  # noqa: E402
from graph_client import leer_excel, subir_creando_si_no_existe  # noqa: E402

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from athena_client import traer_visitas  # noqa: E402
from historial_cambios import cargar_historial, valor_efectivo  # noqa: E402
from motor_clasificacion import TIPO_NEGOCIO_A_CANAL  # noqa: E402
import graph_lists  # noqa: E402

RUTA_GRAPH_MAC = "ASISTENCIA/MAC/"
CARPETA_AUDITORIA_GRAPH = "_nube_Reportes_Diarios/"


def fmt_hora(td):
    if td is None or pd.isna(td):
        return None
    total_seg = int(pd.Timedelta(td).total_seconds())
    h, resto = divmod(total_seg, 3600)
    m, s = divmod(resto, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def limpiar_motivo(texto, estado):
    if texto is None or estado is None:
        return texto
    texto = str(texto)
    prefijo = f"{estado} - "
    if texto.startswith(prefijo):
        texto = texto[len(prefijo):]
    if texto.lower().startswith(estado.lower() + " "):
        texto = texto[len(estado) + 1:]
    texto = re.sub(r"\s*\([^)]*\)\s*$", "", texto).strip()
    if not texto:
        return None
    return texto[0].upper() + texto[1:]


GEOFENCE_MAX_M = 200
TOLERANCIA_MIN = 15
CIERRE_AUTOMATICO_STR = "23:30:00"

CARPETA_SALIDA = "Reportes_Diarios"

COLOR_HEADER = "263544"
COLOR_ASISTIO = "C6EFCE"
COLOR_TARDANZA = "FFEB9C"
COLOR_FALTA = "FFC7CE"
COLOR_VACANTE = "D9D9D9"
COLOR_VACACIONES = "B3EBF2"
COLOR_CORREGIDO = "E1D5E7"
COLOR_EDITABLE = "FFF9DB"


def dia_habil_anterior(fecha, feriados_set=None):
    feriados_set = feriados_set or set()
    d = fecha - pd.Timedelta(days=1)
    while d.weekday() == 6 or d.date() in feriados_set:
        d -= pd.Timedelta(days=1)
    return d


def cargar_maestro(fecha=None):
    m = leer_excel("ASISTENCIA/MAC/1_Maestro_Headcount.xlsx", sheet_name="Maestro Headcount", header=3).dropna(how="all")
    m.columns = [str(c).strip() for c in m.columns]
    m = m.rename(columns={m.columns[0]: "DNI"})
    m = m[pd.to_numeric(m["DNI"], errors="coerce").notna()].copy()
    m["DNI"] = m["DNI"].astype(str).str.strip()
    estado = m["Estado"].astype(str).str.strip()
    incluir = estado.isin(["Activo", "Vacante"])
    if fecha is not None:
        ingreso = pd.to_datetime(m["Fecha de ingreso"], errors="coerce")
        baja = pd.to_datetime(m["Fecha de baja"], errors="coerce")
        estuvo_esa_fecha = (
            (estado == "Inactivo")
            & (ingreso.isna() | (ingreso <= fecha))
            & (baja.isna() | (baja >= fecha))
        )
        incluir = incluir | estuvo_esa_fecha
    return m[incluir].copy()


def cargar_patron():
    p = leer_excel("ASISTENCIA/MAC/2A_Patron_Recurrente.xlsx", sheet_name="Patrón recurrente", header=3).dropna(how="all")
    p.columns = [str(c).strip() for c in p.columns]
    p = p.rename(columns={p.columns[0]: "DNI", p.columns[1]: "Dia"})
    p["DNI"] = p["DNI"].astype(str).str.strip()
    p["Dia"] = p["Dia"].astype(str).str.strip().str.lower()
    dia_map = {"lunes": 0, "martes": 1, "miércoles": 2, "miercoles": 2, "jueves": 3, "viernes": 4, "sábado": 5, "sabado": 5}
    p["weekday"] = p["Dia"].map(dia_map)
    col_ent = [c for c in p.columns if "entrada" in c.lower()][0]
    col_sal = [c for c in p.columns if "salida" in c.lower()][0]
    return p.set_index(["DNI", "weekday"]), col_ent, col_sal


def cargar_feriados():
    feriados = leer_excel("ASISTENCIA/MAC/5_Feriados_2026.xlsx", sheet_name="Feriados 2026", header=3)
    return set(pd.to_datetime(feriados["Fecha"], format="%d/%m/%Y", dayfirst=True).dt.date)


def salida_anterior_por_persona(fecha_anterior):
    session = get_session()
    try:
        filas = (
            session.query(ClasificacionDiaria.dni, ClasificacionDiaria.salida_real)
            .filter(ClasificacionDiaria.fecha == fecha_anterior.date())
            .all()
        )
    finally:
        session.close()
    resultado = {}
    for dni, salida_real in filas:
        td = pd.to_timedelta(str(salida_real), errors="coerce")
        if pd.notna(td):
            resultado[dni] = td
    return resultado


def canal_anterior_por_persona(fecha_anterior):
    session = get_session()
    try:
        filas = (
            session.query(ClasificacionDiaria.dni, ClasificacionDiaria.canales_marcados)
            .filter(ClasificacionDiaria.fecha == fecha_anterior.date())
            .all()
        )
    finally:
        session.close()
    return {dni: canal for dni, canal in filas if canal}


def dnis_asistio_anterior(fecha_anterior):
    session = get_session()
    try:
        filas = (
            session.query(ClasificacionDiaria.dni)
            .filter(ClasificacionDiaria.fecha == fecha_anterior.date(), ClasificacionDiaria.entrada_real.isnot(None))
            .all()
        )
    finally:
        session.close()
    return {dni for (dni,) in filas}


def correcciones_tabla3(hoy, anterior):
    t3 = leer_excel("ASISTENCIA/MAC/3_Registro_Diario_Supervisor.xlsx", sheet_name="Registro diario supervisor", header=3).dropna(how="all")
    t3.columns = [str(c).strip() for c in t3.columns]
    t3 = t3.rename(columns={t3.columns[0]: "DNI"})
    t3 = t3[pd.to_numeric(t3["DNI"], errors="coerce").notna()].copy()
    t3["DNI"] = t3["DNI"].astype(str).str.strip()
    t3["Fecha_dt"] = pd.to_datetime(t3["Fecha"], errors="coerce").dt.date

    ent_hoy = t3[(t3["Fecha_dt"] == hoy.date()) & t3["Hora entrada corregida"].notna()]
    sal_ant = t3[(t3["Fecha_dt"] == anterior.date()) & t3["Hora salida corregida"].notna()]
    return (
        dict(zip(ent_hoy["DNI"], ent_hoy["Hora entrada corregida"])),
        dict(zip(sal_ant["DNI"], sal_ant["Hora salida corregida"])),
    )


def dnis_alguna_vez_en_maestro():
    m = leer_excel("ASISTENCIA/MAC/1_Maestro_Headcount.xlsx", sheet_name="Maestro Headcount", header=3).dropna(how="all")
    m.columns = [str(c).strip() for c in m.columns]
    m = m.rename(columns={m.columns[0]: "DNI"})
    m = m[pd.to_numeric(m["DNI"], errors="coerce").notna()].copy()
    return set(m["DNI"].astype(str).str.strip())


def _leer_registro_supervisor():
    try:
        items = graph_lists.listar_items("RegistroSupervisor")
    except Exception as e:
        print(f"Aviso: no se pudo leer la Lista 'RegistroSupervisor' por Graph ({e}) -- se sigue sin reportes de supervisor por ahora.")
        return pd.DataFrame()
    r = pd.DataFrame(items)
    if r.empty or "DNI" not in r.columns:
        return pd.DataFrame()
    for col in ("EstadoReportado", "Motivo", "MotivoCese", "DNIReemplazo",
                "NombreReemplazo", "Comentario", "RegistradoPor", "FechaRegistro"):
        if col not in r.columns:
            r[col] = pd.NA
    r = r.dropna(subset=["DNI"]).copy()
    r["Fecha_dt"] = pd.to_datetime(r["Fecha"], utc=True, errors="coerce").dt.date
    return r


def reemplazos_pendientes(dnis_ya_en_maestro):
    r = _leer_registro_supervisor()
    if r.empty:
        return []
    r = r[r["EstadoReportado"] == "Tiene reemplazo"]
    r = r[r["DNIReemplazo"].notna()]
    if r.empty:
        return []
    r["DNIReemplazo"] = r["DNIReemplazo"].astype(float).astype(int).astype(str)
    r["DNI"] = r["DNI"].astype(float).astype(int).astype(str)

    pendientes = []
    vistos = set()
    for _, row in r.iterrows():
        dni_reemplazo = row["DNIReemplazo"]
        if dni_reemplazo in dnis_ya_en_maestro or dni_reemplazo in vistos:
            continue
        vistos.add(dni_reemplazo)
        pendientes.append({
            "dni_vacante": row["DNI"],
            "dni_reemplazo": dni_reemplazo,
            "nombre_reemplazo": row.get("NombreReemplazo"),
            "motivo_cese": row.get("MotivoCese"),
            "fecha": row.get("Fecha"),
        })
    return pendientes


def comentarios_supervisor_dia(fecha):
    r = _leer_registro_supervisor()
    if r.empty:
        return {}
    r_dia = r[r["Fecha_dt"] == fecha.date()]
    if r_dia.empty:
        return {}
    r_dia = r_dia.copy()
    r_dia["DNI"] = r_dia["DNI"].astype(float).astype(int).astype(str)

    comentarios = {}
    for _, row in r_dia.iterrows():
        motivo = row.get("Motivo")
        estado = row.get("EstadoReportado")
        motivo_cese = row.get("MotivoCese")
        comentario_libre = row.get("Comentario")
        if pd.notna(motivo):
            texto = f"{estado} - {motivo}" if pd.notna(estado) else str(motivo)
        elif pd.notna(motivo_cese):
            texto = f"{estado} - {motivo_cese}" if pd.notna(estado) else str(motivo_cese)
        else:
            texto = str(estado) if pd.notna(estado) else None
        if pd.notna(comentario_libre) and str(comentario_libre).strip():
            texto = f"{texto} ({comentario_libre})" if texto else str(comentario_libre)
        if texto:
            comentarios[row["DNI"]] = texto
    return comentarios


def comentarios_supervisor_hoy(hoy):
    return comentarios_supervisor_dia(hoy)


def comentario_historico_dia(fecha):
    session = get_session()
    try:
        filas = (
            session.query(ClasificacionDiaria.dni, ClasificacionDiaria.comentario_supervisor)
            .filter(ClasificacionDiaria.fecha == fecha.date())
            .all()
        )
    finally:
        session.close()
    return {dni: c for dni, c in filas if c}


def main():
    hoy = pd.Timestamp(sys.argv[1]) if len(sys.argv) > 1 else pd.Timestamp.today().normalize()
    feriados_set = cargar_feriados()
    anterior = dia_habil_anterior(hoy, feriados_set)

    activos = cargar_maestro(hoy)
    p_idx, col_ent, col_sal = cargar_patron()
    idx_historial = cargar_historial()

    print("Trayendo visitas de hoy desde Athena (tabla rápida)...")
    filtro_hoy = f"AND date_parse(v.fecha_inicio, '%d-%m-%Y') = date '{hoy.date()}'"
    v = traer_visitas("dim_lf_general_visitas_diarias", filtro_hoy)
    v["nro_documento"] = v["nro_documento"].astype(str).str.strip()

    v["_hora_fin_dt"] = pd.to_datetime(v["fecha_fin"] + " " + v["hora_fin"], format="%d-%m-%Y %H:%M:%S", errors="coerce")
    ultima_carga_athena = v["_hora_fin_dt"].max()

    print("Completando con la tabla histórico (respaldo, solo hoy)...")
    v_hist = traer_visitas("dim_lf_general_visitas", filtro_hoy)
    v_hist["nro_documento"] = v_hist["nro_documento"].astype(str).str.strip()
    dnis_en_diaria = set(v["nro_documento"].unique())
    v_respaldo = v_hist[~v_hist["nro_documento"].isin(dnis_en_diaria)]
    if len(v_respaldo):
        print(f"{v_respaldo['nro_documento'].nunique()} persona(s) sin marcación en 'diarias' completada(s) desde 'histórico'")
        v = pd.concat([v, v_respaldo], ignore_index=True)

    v["hora_inicio_td"] = pd.to_timedelta(v["hora_inicio"].astype(str), errors="coerce")
    v["geofence_ok"] = v["distancia_metros_inicio"] <= GEOFENCE_MAX_M
    entrada_real_por_dni = v.groupby("nro_documento")["hora_inicio_td"].min().to_dict()

    v["canal_visita"] = v["tipo_negocio"].map(TIPO_NEGOCIO_A_CANAL).fillna("Otro")
    canal_hoy_por_dni = (v.groupby("nro_documento")["canal_visita"]
                         .apply(lambda s: ", ".join(sorted(s.unique()))).to_dict())

    session = get_session()
    try:
        _filas_clasif_hoy = (
            session.query(ClasificacionDiaria.dni, ClasificacionDiaria.entrada_real)
            .filter(ClasificacionDiaria.fecha == hoy.date())
            .all()
        )
    finally:
        session.close()
    entrada_real_clasif = {dni: er for dni, er in _filas_clasif_hoy if er}

    salida_ant_por_dni = salida_anterior_por_persona(anterior)
    canal_ant_por_dni = canal_anterior_por_persona(anterior)
    dnis_asistio_ant = dnis_asistio_anterior(anterior)
    comentarios_hoy = comentarios_supervisor_hoy(hoy)
    comentarios_ayer = comentarios_supervisor_dia(anterior)
    comentario_hist_hoy = comentario_historico_dia(hoy)
    comentario_hist_ayer = comentario_historico_dia(anterior)
    correcciones_entrada, correcciones_salida = correcciones_tabla3(hoy, anterior)

    filas = []
    dnis_entrada_corregida = set()
    dnis_salida_corregida = set()
    for _, persona in activos.iterrows():
        dni = persona["DNI"]
        weekday_hoy = hoy.weekday()
        weekday_ant = anterior.weekday()

        if weekday_hoy == 6 or hoy.date() in feriados_set:
            continue

        fecha_ingreso = persona.get("Fecha de ingreso")
        fecha_baja = persona.get("Fecha de baja")
        if pd.notna(fecha_ingreso) and hoy < pd.Timestamp(fecha_ingreso):
            continue
        if pd.notna(fecha_baja) and hoy > pd.Timestamp(fecha_baja):
            continue

        pat_hoy = p_idx.loc[(dni, weekday_hoy)] if (dni, weekday_hoy) in p_idx.index else None
        if pat_hoy is None:
            continue
        if isinstance(pat_hoy, pd.DataFrame):
            pat_hoy = pat_hoy.iloc[0]

        es_vacante = str(persona.get("Estado")).strip() == "Vacante"
        comentario_entrada_hoy = comentarios_hoy.get(dni) or comentario_hist_hoy.get(dni)
        comentario_entrada_norm = str(comentario_entrada_hoy).upper() if comentario_entrada_hoy is not None else ""

        entrada_esp_td = pd.to_timedelta(str(
            valor_efectivo(idx_historial, dni, "Hora entrada programada", hoy, pat_hoy[col_ent])
        ))
        entrada_real = entrada_real_por_dni.get(dni)
        if entrada_real is None and dni in entrada_real_clasif:
            entrada_real = pd.to_timedelta(str(entrada_real_clasif[dni]))
        entrada_desde_correccion = dni in correcciones_entrada
        if entrada_desde_correccion:
            entrada_real = pd.to_timedelta(str(correcciones_entrada[dni]))
            dnis_entrada_corregida.add(dni)

        if not entrada_desde_correccion and (es_vacante or "VACANTE" in comentario_entrada_norm):
            estado_entrada = "Vacante"
            comentario_entrada_hoy = None
        elif not entrada_desde_correccion and "VACACIONES" in comentario_entrada_norm:
            estado_entrada = "Vacaciones"
            comentario_entrada_hoy = None
        elif not entrada_desde_correccion and comentario_entrada_norm.startswith("FALTA"):
            estado_entrada = "Falta"
            comentario_entrada_hoy = limpiar_motivo(comentario_entrada_hoy, "Falta")
        elif not entrada_desde_correccion and entrada_real is None and "ASISTIÓ" in comentario_entrada_norm:
            estado_entrada = "Asistió (por confirmar)"
            comentario_entrada_hoy = "Asistió (reportado por supervisor -- confirmar en \"Hora entrada corregida\")"
        elif entrada_real is None:
            estado_entrada = "Falta"
            comentario_entrada_hoy = limpiar_motivo(comentario_entrada_hoy, "Falta")
        else:
            diff_min = (entrada_real - entrada_esp_td).total_seconds() / 60
            estado_entrada = "Tardanza" if diff_min > TOLERANCIA_MIN else "Asistió"
            texto_auto_superado = comentario_entrada_norm in ("ASISTIÓ", "VACANTE", "VACACIONES") or comentario_entrada_norm.startswith((
                "ASISTIÓ (REPORTADO POR SUPERVISOR",
                "CORRECCIÓN DESDE REPORTE DIARIO 9AM",
            ))
            if (estado_entrada == "Asistió" or entrada_desde_correccion) and comentario_entrada_hoy and texto_auto_superado:
                comentario_entrada_hoy = None

        ingreso_despues_de_ayer = pd.notna(fecha_ingreso) and anterior < pd.Timestamp(fecha_ingreso)

        pat_ant = p_idx.loc[(dni, weekday_ant)] if (dni, weekday_ant) in p_idx.index and not ingreso_despues_de_ayer else None
        if isinstance(pat_ant, pd.DataFrame):
            pat_ant = pat_ant.iloc[0]
        salida_esp_ant = str(
            valor_efectivo(idx_historial, dni, "Hora salida programada", anterior, pat_ant[col_sal])
        ) if pat_ant is not None else None
        salida_real_ant = None if ingreso_despues_de_ayer else salida_ant_por_dni.get(dni)
        salida_desde_correccion = dni in correcciones_salida and not ingreso_despues_de_ayer
        if salida_desde_correccion:
            salida_real_ant = pd.to_timedelta(str(correcciones_salida[dni]))
            dnis_salida_corregida.add(dni)

        comentario_salida = None
        if es_vacante:
            comentario_salida = "Vacante"
        elif pat_ant is not None:
            comentario_salida = comentario_hist_ayer.get(dni) or comentarios_ayer.get(dni)
            if salida_real_ant is None and comentario_salida and "ASISTIÓ" in comentario_salida.upper():
                comentario_salida = "Asistió (reportado por supervisor -- confirmar en \"Hora salida corregida\")"
            elif comentario_salida is None and salida_real_ant is None:
                comentario_salida = ("Salida no registrada (sí asistió)" if dni in dnis_asistio_ant
                                      else "Faltó (sin motivo registrado)")
            elif salida_real_ant is not None and comentario_salida and (
                comentario_salida.strip().upper() in ("ASISTIÓ", "FALTÓ (SIN MOTIVO REGISTRADO)")
                or comentario_salida.strip().upper().startswith((
                    "CORRECCIÓN DESDE REPORTE DIARIO 9AM", "SALIDA NO REGISTRADA",
                ))
            ):
                comentario_salida = None

        rol = str(persona.get("Rol")).strip() if pd.notna(persona.get("Rol")) else None
        canal_o_rol = (
            valor_efectivo(idx_historial, dni, "Canal", hoy, persona.get("Canal"))
            if rol == "MERCADERISTAS" else rol
        )

        filas.append({
            "Región": persona.get("Región"),
            "Supervisor": valor_efectivo(idx_historial, dni, "Supervisor", hoy, persona.get("Supervisor asignado")),
            "Ciudad": persona.get("Ciudad / Mercado"),
            "CANAL": canal_o_rol,
            "Canal(es) marcado(s) hoy": canal_hoy_por_dni.get(dni) or None,
            "Mercado": valor_efectivo(idx_historial, dni, "Zona", hoy, persona.get("Zona / Ruta asignada")),
            "Dni": dni,
            "Mercaderista": persona["Nombre completo"],
            "Fecha": hoy.date(),
            "Hora entrada prog": fmt_hora(entrada_esp_td),
            "Hora entrada real": fmt_hora(entrada_real),
            "Estado_Entrada": estado_entrada,
            "Comentario Entrada": comentario_entrada_hoy,
            "Hora salida prog anterior": fmt_hora(pd.to_timedelta(str(salida_esp_ant))) if salida_esp_ant else None,
            "Hora salida real anterior": fmt_hora(salida_real_ant),
            "Canal(es) marcado(s) ayer": canal_ant_por_dni.get(dni) or None,
            "Comentario salida": comentario_salida,
            "Hora entrada corregida": None,
            "Hora salida corregida": None,
        })

    df = pd.DataFrame(filas)
    df = df.sort_values(["Región", "Supervisor", "Ciudad", "CANAL", "Mercado", "Mercaderista"]).reset_index(drop=True)

    for col in ["Región", "Supervisor", "Ciudad", "CANAL", "Mercado", "Mercaderista"]:
        df[col] = df[col].astype(str).str.title().replace("Nan", None)

    n_asistio = (df["Estado_Entrada"] == "Asistió").sum()
    n_tardanza = (df["Estado_Entrada"] == "Tardanza").sum()
    n_falta = (df["Estado_Entrada"] == "Falta").sum()
    n_vacante = (df["Estado_Entrada"] == "Vacante").sum()
    n_vacaciones = (df["Estado_Entrada"] == "Vacaciones").sum()
    n_pendiente = (df["Estado_Entrada"] == "Asistió (por confirmar)").sum()

    pendientes = reemplazos_pendientes(dnis_alguna_vez_en_maestro())
    if pendientes:
        print(f"⚠ {len(pendientes)} reemplazo(s) pendiente(s) de procesar con ventana_reemplazos.py")

    os.makedirs(CARPETA_SALIDA, exist_ok=True)
    destino = os.path.join(CARPETA_SALIDA, f"Reporte_Asistencia_{hoy.date()}.xlsx")
    escribir_excel(df, destino, hoy, n_asistio, n_tardanza, n_falta, n_vacante, n_vacaciones, n_pendiente, pendientes,
                    dnis_entrada_corregida, dnis_salida_corregida, ultima_carga_athena)
    print(f"Guardado: {destino} ({len(df)} filas -- Asistió={n_asistio}, Tardanza={n_tardanza}, Falta={n_falta}, "
          f"Vacante={n_vacante}, Vacaciones={n_vacaciones}, Por_confirmar={n_pendiente})")


def escribir_excel(df, path, hoy, n_asistio, n_tardanza, n_falta, n_vacante=0, n_vacaciones=0, n_pendiente=0,
                    pendientes=None, dnis_entrada_corregida=None, dnis_salida_corregida=None, ultima_carga_athena=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    FUENTE = "Aptos"

    ws["A1"] = f"Reporte de Asistencia -- {hoy.date()}"
    ws["A1"].font = Font(name=FUENTE, bold=True, size=14)
    total = n_asistio + n_tardanza + n_falta + n_vacante + n_vacaciones + n_pendiente
    texto_vacante = f" · {n_vacante} Vacante" if n_vacante else ""
    texto_vacaciones = f" · {n_vacaciones} Vacaciones" if n_vacaciones else ""
    texto_pendiente = f" · {n_pendiente} Por confirmar" if n_pendiente else ""
    ws["A2"] = (f"{n_asistio} Asistieron · {n_tardanza} Tardanza · {n_falta} Falta{texto_vacante}"
                f"{texto_vacaciones}{texto_pendiente}  (Total: {total})")
    ws["A2"].font = Font(name=FUENTE, bold=True, size=11, color="52514E")
    ws["A3"] = "Si el aplicativo marcó mal, llena \"Hora entrada corregida\" / \"Hora salida corregida\" (columnas amarillas) y corre aplicar_correcciones.py"
    ws["A3"].font = Font(name=FUENTE, italic=True, size=9, color="8A6D00")
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:H3")

    if ultima_carga_athena is not None and pd.notna(ultima_carga_athena):
        atraso_min = (pd.Timestamp.now() - ultima_carga_athena).total_seconds() / 60
        texto_atraso = f"Athena actualizado hasta las {ultima_carga_athena:%H:%M:%S} (hace {int(atraso_min)} min)"
        color_atraso = "9C0006" if atraso_min > 45 else "808080"
    else:
        texto_atraso = "Athena: sin marcaciones de hoy para calcular frescura"
        color_atraso = "9C0006"
    ws["F2"] = texto_atraso
    ws["F2"].font = Font(name=FUENTE, italic=True, size=9, color=color_atraso)
    ws.merge_cells("F2:H2")

    if pendientes:
        nombres = ", ".join(f"{p['nombre_reemplazo']} (DNI {p['dni_reemplazo']}, reemplaza a {p['dni_vacante']})" for p in pendientes)
        ws["A4"] = f"⚠ {len(pendientes)} REEMPLAZO(S) PENDIENTE(S) DE PROCESAR con ventana_reemplazos.py: {nombres}"
        ws["A4"].font = Font(name=FUENTE, bold=True, size=10, color="9C0006")
        ws["A4"].fill = PatternFill("solid", fgColor="FFC7CE")
        ws.merge_cells("A4:H4")

    fila_header = 5
    for c, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=fila_header, column=c, value=col_name)
        cell.font = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_asistio = PatternFill("solid", fgColor=COLOR_ASISTIO)
    fill_tardanza = PatternFill("solid", fgColor=COLOR_TARDANZA)
    fill_falta = PatternFill("solid", fgColor=COLOR_FALTA)
    fill_vacante = PatternFill("solid", fgColor=COLOR_VACANTE)
    fill_vacaciones = PatternFill("solid", fgColor=COLOR_VACACIONES)
    fill_editable = PatternFill("solid", fgColor=COLOR_EDITABLE)
    fill_corregido = PatternFill("solid", fgColor=COLOR_CORREGIDO)
    dnis_entrada_corregida = dnis_entrada_corregida or set()
    dnis_salida_corregida = dnis_salida_corregida or set()
    idx_estado = list(df.columns).index("Estado_Entrada") + 1
    cols_editables = {list(df.columns).index("Hora entrada corregida") + 1, list(df.columns).index("Hora salida corregida") + 1}
    idx_ent_prog = list(df.columns).index("Hora entrada prog")
    idx_ent_real = list(df.columns).index("Hora entrada real")
    idx_sal_prog = list(df.columns).index("Hora salida prog anterior")
    idx_sal_real = list(df.columns).index("Hora salida real anterior")
    idx_dni = list(df.columns).index("Dni")
    font_alerta = Font(name=FUENTE, size=10, bold=True, color="C00000")
    borde_fino = Side(style="thin", color="D9D9D9")

    for r, row in enumerate(df.itertuples(index=False), start=fila_header + 1):
        ent_prog, ent_real = row[idx_ent_prog], row[idx_ent_real]
        dni_fila = row[idx_dni]
        dni_fila_str = str(int(dni_fila)) if isinstance(dni_fila, float) else str(dni_fila)
        entrada_tarde = False
        if pd.notna(ent_prog) and pd.notna(ent_real):
            diff = (pd.to_timedelta(str(ent_real)) - pd.to_timedelta(str(ent_prog))).total_seconds() / 60
            entrada_tarde = diff > TOLERANCIA_MIN

        sal_prog, sal_real = row[idx_sal_prog], row[idx_sal_real]
        salida_temprano = False
        if pd.notna(sal_prog) and pd.notna(sal_real):
            diff_salida = (pd.to_timedelta(str(sal_prog)) - pd.to_timedelta(str(sal_real))).total_seconds() / 60
            salida_temprano = diff_salida > TOLERANCIA_MIN

        for c, val in enumerate(row, start=1):
            v = None if pd.isna(val) else val
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FUENTE, size=10)
            cell.border = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
            if c == idx_estado:
                if v == "Asistió":
                    cell.fill = fill_asistio
                elif v == "Tardanza":
                    cell.fill = fill_tardanza
                elif v == "Falta":
                    cell.fill = fill_falta
                elif v == "Vacante":
                    cell.fill = fill_vacante
                elif v == "Vacaciones":
                    cell.fill = fill_vacaciones
                elif v == "Asistió (por confirmar)":
                    cell.fill = fill_editable
            elif c in cols_editables:
                cell.fill = fill_editable
            elif c == idx_ent_real + 1:
                if dni_fila_str in dnis_entrada_corregida:
                    cell.fill = fill_corregido
                if entrada_tarde:
                    cell.font = font_alerta
            elif c == idx_sal_real + 1:
                if dni_fila_str in dnis_salida_corregida:
                    cell.fill = fill_corregido
                if salida_temprano:
                    cell.font = font_alerta

    n_filas = len(df) + 1
    n_cols = len(df.columns)
    ref = f"A{fila_header}:{get_column_letter(n_cols)}{fila_header + n_filas - 1}"
    tabla = Table(displayName="ReporteAsistencia", ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(name=None, showRowStripes=False)
    ws.add_table(tabla)

    for col_oculta in ("Dni", "Fecha"):
        if col_oculta in df.columns:
            letra = get_column_letter(list(df.columns).index(col_oculta) + 1)
            ws.column_dimensions[letra].hidden = True

    anchos = [14, 22, 12, 14, 18, 22, 13, 34, 13, 15, 15, 15, 22, 18, 18, 18, 20, 20, 20]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = f"A{fila_header + 1}"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 80
    wb.save(path)

    # Fase 6: ademas de local (usado por aplicar_correcciones.py DENTRO de la
    # misma corrida), sube por Graph a la ruta de auditoria -- el runner no
    # tiene disco entre corridas, y la version real de esto necesita
    # persistir para que la SIGUIENTE corrida (aplicar_correcciones.py, en
    # otro job) pueda leerlo.
    try:
        nombre_archivo = os.path.basename(path)
        subir_creando_si_no_existe(RUTA_GRAPH_MAC + CARPETA_AUDITORIA_GRAPH + nombre_archivo, wb)
        print(f"Subido a SharePoint (auditoría): {CARPETA_AUDITORIA_GRAPH}{nombre_archivo}")
    except Exception as e:
        print(f"AVISO: no se pudo subir el reporte a SharePoint (auditoría) -- {e}")


if __name__ == "__main__":
    main()
