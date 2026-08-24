# -*- coding: utf-8 -*-
"""
Fase 6: version portable de MAC/export_sharepoint.py -- la parte de datos
(consulta a Postgres, correo_map, vacantes) es identica a la ya migrada hoy
en Fase 3. El cambio real es el de escritura: el original usa
escribir_tabla_in_place(), que necesita disco local para su logica de
"parchear el archivo existente, con Graph como respaldo" -- en un runner sin
disco persistente eso no aplica, asi que esta version sube directo por Graph
con subir_creando_si_no_existe() (crea o reemplaza el contenido preservando
el ID del archivo -- PUT por ruta, no hace falta el paso de resolver-y-parchear
que si necesitaba el original para coexistir con una copia local).

CORTE A PRODUCCION (2026-08-22): tras 1 dia corriendo en paralelo contra la
ruta de auditoria (_nube_Clasificacion_Diaria.xlsx) sin discrepancias, este
script ya escribe el Clasificacion_Diaria.xlsx real y sincroniza la Lista
"ClasificacionDiaria" de SharePoint que lee Power Apps -- ver RUTA_SALIDA
mas abajo.
"""
import os
import sys

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dimension_models import Persona, get_session  # noqa: E402
from excel_safety import texto_seguro_excel  # noqa: E402
from fact_models import ClasificacionDiaria  # noqa: E402
from graph_client import leer_excel, subir_creando_si_no_existe  # noqa: E402

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from historial_cambios import cargar_historial, valor_efectivo  # noqa: E402
import graph_lists  # noqa: E402

RUTA_GRAPH_MAC = "ASISTENCIA/MAC/"
# Ruta real que lee Power Apps -- NO vive bajo ASISTENCIA/MAC/, un nivel
# arriba (ver export_sharepoint.py original, CARPETA_SHAREPOINT).
RUTA_SALIDA = "ASISTENCIA/Clasificacion_Diaria.xlsx"

NOMBRE_CORTO = {
    "40628345": "KARINA", "40817159": "DANIELA CABANILLAS", "80270250": "ANA CARBAJAL",
    "45694774": "ELIZABETH TICONA", "9919446": "EDITH CAMACHO",
    "74943612": "GUERRA MARIA",
}


def escribir_tabla(df, nombre_hoja, nombre_tabla):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        # texto_seguro_excel(): columnas como nombre_completo/comentario_supervisor
        # vienen de texto libre subido por analistas o escrito desde la web
        # -- sin esto, un valor que empiece con =/+/-/@ se interpretaria
        # como formula al abrir este archivo en Excel (CSV/Excel Injection).
        ws.append([None if pd.isna(v) else texto_seguro_excel(v) for v in row])
    n_filas = len(df) + 1
    n_cols = len(df.columns)
    ref = f"A1:{get_column_letter(n_cols)}{max(n_filas, 2)}"
    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tabla)
    for col_idx, col_name in enumerate(df.columns, start=1):
        ancho = max(12, min(40, len(str(col_name)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    # subir_creando_si_no_existe() hace PUT por ruta preservando el ID del
    # archivo si ya existe (que es el caso siempre, salvo la primerisima
    # corrida) -- no rompe la conexion de Power Apps al Excel.
    subir_creando_si_no_existe(RUTA_SALIDA, wb)


def construir_correo_map():
    m = leer_excel("ASISTENCIA/MAC/1_Maestro_Headcount.xlsx", sheet_name="Maestro Headcount", header=3).dropna(how="all")
    m.columns = [str(c).strip() for c in m.columns]
    m = m.rename(columns={m.columns[0]: "DNI"})
    # Mismo gotcha que parseo_headcount.py -- si ALGUNA fila de DNI viene
    # vacía, pandas sube toda la columna a float64 y sobrevive al filtro de
    # abajo, dejando "18074336.0" en vez de "18074336" para todos los DNIs
    # válidos (rompe el match contra ClasificacionDiaria.dni más abajo).
    dni_num = pd.to_numeric(m["DNI"], errors="coerce")
    m = m[dni_num.notna()].copy()
    m["DNI"] = dni_num[dni_num.notna()].astype("int64").astype(str)
    sup = m[m["Rol"] == "SUPERVISORES"]
    correo_por_nombre_corto = {}
    for _, row in sup.iterrows():
        nombre_corto = NOMBRE_CORTO.get(row["DNI"], row["Nombre completo"])
        correo_por_nombre_corto[nombre_corto] = row["Correo corporativo"]
    return m, correo_por_nombre_corto


def generar_clasificacion_diaria():
    hoy = pd.Timestamp.today().normalize()

    session = get_session()
    try:
        filas = (
            session.query(ClasificacionDiaria.dni, Persona.nombre_completo, ClasificacionDiaria.fecha,
                           ClasificacionDiaria.canal_esperado, ClasificacionDiaria.canales_marcados,
                           ClasificacionDiaria.entrada_esperada, ClasificacionDiaria.entrada_real,
                           ClasificacionDiaria.salida_esperada, ClasificacionDiaria.salida_real,
                           ClasificacionDiaria.estado, ClasificacionDiaria.salida_anticipada_min,
                           ClasificacionDiaria.trabajo_otro_canal, ClasificacionDiaria.alerta_analista,
                           ClasificacionDiaria.fuente_dato)
            .join(Persona, Persona.dni == ClasificacionDiaria.dni)
            .filter(ClasificacionDiaria.fecha == hoy.date())
            .all()
        )
    finally:
        session.close()

    r_hoy = pd.DataFrame(filas, columns=[
        "DNI", "Nombre", "Fecha", "Canal esperado (Patrón)", "Canal(es) marcado(s)",
        "Entrada esperada", "Entrada real", "Salida esperada", "Salida real", "Estado",
        "Salida anticipada (min)", "Trabajó para otro canal", "Alerta para analista", "Fuente del dato",
    ])
    r_hoy["Trabajó para otro canal"] = r_hoy["Trabajó para otro canal"].map({True: "SÍ", False: "NO"})
    r_hoy["Alerta para analista"] = r_hoy["Alerta para analista"].map({True: "SÍ", False: "NO"})
    r_hoy["Fecha"] = pd.to_datetime(r_hoy["Fecha"])

    m, correo_map = construir_correo_map()
    sup_asignado = m.set_index("DNI")["Supervisor asignado"].str.strip().to_dict()
    idx_historial = cargar_historial()
    r_hoy["_supervisor_nombre"] = r_hoy["DNI"].apply(
        lambda d: valor_efectivo(idx_historial, d, "Supervisor", hoy, sup_asignado.get(d))
    )
    r_hoy["Correo supervisor"] = r_hoy["_supervisor_nombre"].map(correo_map)
    ciudad_map = m.set_index("DNI")["Ciudad / Mercado"].to_dict()
    r_hoy["Ciudad"] = r_hoy["DNI"].map(ciudad_map)

    r_hoy["Estado base"] = r_hoy["Estado"].str.split(" (", regex=False).str[0]

    cols = [
        "DNI", "Nombre", "Fecha", "Correo supervisor", "Ciudad", "Canal esperado (Patrón)",
        "Canal(es) marcado(s)", "Entrada esperada", "Entrada real", "Salida esperada",
        "Salida real", "Estado", "Estado base", "Salida anticipada (min)", "Trabajó para otro canal",
        "Alerta para analista", "Fuente del dato",
    ]
    out = r_hoy[[c for c in cols if c in r_hoy.columns]].copy()
    out["Fecha"] = out["Fecha"].dt.strftime("%Y-%m-%d")

    out = out.rename(columns={
        "Correo supervisor": "CorreoSupervisor",
        "Canal esperado (Patrón)": "CanalEsperado",
        "Canal(es) marcado(s)": "CanalMarcado",
        "Entrada esperada": "EntradaEsperada",
        "Entrada real": "EntradaReal",
        "Salida esperada": "SalidaEsperada",
        "Salida real": "SalidaReal",
        "Estado base": "EstadoBase",
        "Salida anticipada (min)": "SalidaAnticipadaMin",
        "Trabajó para otro canal": "TrabajoOtroCanal",
        "Alerta para analista": "AlertaAnalista",
        "Fuente del dato": "FuenteDato",
    })

    vacantes = m[m["Estado"].astype(str).str.strip() == "Vacante"].copy()
    # Si un DNI ya tiene una fila real de ClasificacionDiaria hoy (ej.
    # trabajó antes de quedar Vacante recién hoy en el Maestro), no se
    # agrega también la fila sintética "VACANTE SIN CUBRIR" -- 2 filas con
    # el mismo DNI el mismo día hacían que reemplazar_todos_los_items()
    # (indexa por DNI en la Lista de SharePoint que lee Power Apps) se
    # quedara solo con una de las 2 en silencio.
    dnis_ya_en_out = set(out["DNI"].astype(str))
    vacantes = vacantes[~vacantes["DNI"].astype(str).isin(dnis_ya_en_out)]
    if len(vacantes):
        filas_vacante = pd.DataFrame({
            "DNI": vacantes["DNI"],
            "Nombre": vacantes["Nombre completo"].astype(str),
            "Fecha": hoy.strftime("%Y-%m-%d"),
            "CorreoSupervisor": vacantes["DNI"].map(sup_asignado).map(correo_map),
            "Ciudad": vacantes["DNI"].map(ciudad_map),
            "CanalEsperado": None, "CanalMarcado": None,
            "EntradaEsperada": None, "EntradaReal": None,
            "SalidaEsperada": None, "SalidaReal": None,
            "Estado": "VACANTE SIN CUBRIR", "EstadoBase": "VACANTE",
            "SalidaAnticipadaMin": None, "TrabajoOtroCanal": "NO",
            "AlertaAnalista": "SÍ", "FuenteDato": "Vacante (Maestro)",
        })
        out = pd.concat([out, filas_vacante], ignore_index=True)
        print(f"Vacantes sin cubrir agregadas: {len(filas_vacante)}")

    sin_correo = out["CorreoSupervisor"].isna().sum()
    if sin_correo:
        print(f"Aviso: {sin_correo} filas sin correo de supervisor (Supervisor asignado sin match)")

    escribir_tabla(out, "Clasificacion Diaria", "ClasificacionDiaria")
    print(f"Guardado: {RUTA_SALIDA} ({len(out)} filas, hoy={hoy.date()})")

    resultado = graph_lists.reemplazar_todos_los_items("ClasificacionDiaria", out.to_dict("records"))
    fallidos = resultado["fallidos"]
    print(f"Lista SharePoint 'ClasificacionDiaria' sincronizada: "
          f"{resultado['creados']} creados, {resultado['actualizados']} actualizados, {resultado['borrados']} borrados.")
    if fallidos:
        # Igual que la version local: si algo no sincronizo, se lanza para
        # que pipeline_completo.py lo marque como FALLO -- no se traga en
        # silencio, la Lista quedaria incompleta sin que nadie se entere.
        detalle = "; ".join(f"{f['accion']} DNI {f['dni']}: {f['error']}" for f in fallidos)
        raise RuntimeError(
            f"{len(fallidos)} de {len(out)} filas no se sincronizaron en la Lista 'ClasificacionDiaria' -- {detalle}"
        )


if __name__ == "__main__":
    generar_clasificacion_diaria()
