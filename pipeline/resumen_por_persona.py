# -*- coding: utf-8 -*-
"""
Fase 6: version portable de MAC/resumen_por_persona.py -- ya era 100%
Postgres (Fase 3), solo cambia el import a ruta relativa. Se agrega al
pipeline de la nube (no estaba en pipeline_completo.py original) para que
9_Resumen_Por_Persona.xlsx -- del que depende export_dashboard_data.py --
este siempre fresco, en vez de quedar potencialmente stale como en la
laptop (ahi ese archivo no se regeneraba solo, habia que correrlo a mano).
"""
import os
import sys

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dimension_models import Persona, get_session  # noqa: E402
from fact_models import ClasificacionDiaria  # noqa: E402

session = get_session()
try:
    filas = (
        session.query(ClasificacionDiaria.dni, Persona.nombre_completo, ClasificacionDiaria.estado,
                       ClasificacionDiaria.trabajo_otro_canal)
        .join(Persona, Persona.dni == ClasificacionDiaria.dni)
        .all()
    )
finally:
    session.close()

r = pd.DataFrame(filas, columns=["DNI", "Nombre", "Estado", "Trabajó para otro canal"])
r["Trabajó para otro canal"] = r["Trabajó para otro canal"].map({True: "SÍ", False: "NO"})
r["Estado_base"] = r["Estado"].apply(lambda s: s.split(" (")[0])

resumen = r.groupby(["DNI", "Nombre"]).apply(lambda g: pd.Series({
    "Días evaluados": len(g),
    "Asistió a tiempo": (g["Estado_base"] == "ASISTIÓ A TIEMPO").sum(),
    "Tardanza": (g["Estado_base"] == "TARDANZA").sum(),
    "Falta": (g["Estado_base"] == "FALTA").sum(),
    "Alerta geofence": (g["Estado_base"].str.startswith("ALERTA")).sum(),
    "Trabajó otro canal": (g["Trabajó para otro canal"] == "SÍ").sum(),
}), include_groups=False).reset_index()

resumen["% Falta"] = (resumen["Falta"] / resumen["Días evaluados"] * 100).round(1)
resumen["% Asistencia efectiva"] = ((resumen["Asistió a tiempo"] + resumen["Tardanza"]) / resumen["Días evaluados"] * 100).round(1)
resumen = resumen.sort_values(["Falta", "Tardanza"], ascending=False)

print(f"Total personas: {len(resumen)}")

TITLE_FONT = Font(bold=True, color="1F4E78", size=13)
DESC_FONT = Font(color="595959", size=10)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resumen por persona"
ws.cell(row=1, column=1, value="Tabla 9 – Resumen de clasificación diaria por persona").font = TITLE_FONT
ws.cell(row=2, column=1, value="Regenerado automáticamente (nube) desde Postgres.").font = DESC_FONT
headers = list(resumen.columns)
for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BORDER
ws.freeze_panes = "A5"
for r_idx, (_, row) in enumerate(resumen.iterrows(), start=5):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val).border = BORDER
widths = [12, 28, 14, 16, 10, 8, 14, 16, 10, 20]
for col_idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w

wb.save("9_Resumen_Por_Persona.xlsx")
print("Guardado: 9_Resumen_Por_Persona.xlsx")
