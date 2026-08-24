# -*- coding: utf-8 -*-
"""
Fase 6: version portable de MAC/aplicar_correcciones.py -- misma logica de
negocio, pero el reporte de las 9am (donde el analista escribe las
correcciones a mano) se lee/escribe por Graph, no por disco local. En la
laptop el disco persiste entre corridas (asi lo lee el original,
openpyxl.load_workbook(reporte_path) puro); en un runner de GitHub Actions
cada corrida arranca en blanco, asi que la unica forma de ver lo que el
analista corrigio es descargar la version mas reciente de SharePoint (donde
ya esta, porque OneDrive sincroniza sola la edicion local del analista hacia
SharePoint -- este script no necesita saber nada de eso, solo lee la copia
de SharePoint) -- ver reporte_diario_9am.py (Fase 6) para el lado que sube
esa copia.

Uso: python aplicar_correcciones.py [YYYY-MM-DD]
"""
import io
import os
import subprocess
import sys

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from graph_client import descargar, leer_excel, subir_creando_si_no_existe, subir_in_place  # noqa: E402

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from reporte_diario_9am import (  # noqa: E402
    dia_habil_anterior, cargar_feriados, comentarios_supervisor_dia, comentario_historico_dia, CARPETA_SALIDA,
    limpiar_motivo, RUTA_GRAPH_MAC, CARPETA_AUDITORIA_GRAPH,
)

TABLA3_PATH = "3_Registro_Diario_Supervisor.xlsx"
TABLA3_RUTA_GRAPH = f"{RUTA_GRAPH_MAC}{TABLA3_PATH}"

TEXTO_PENDIENTE_ENTRADA = "Asistió (reportado por supervisor -- confirmar en \"Hora entrada corregida\")"
TEXTO_PENDIENTE_SALIDA = "Asistió (reportado por supervisor -- confirmar en \"Hora salida corregida\")"


def _valor_auto_entrada(comentario_crudo):
    if comentario_crudo is None:
        return None
    texto_upper = str(comentario_crudo).upper()
    if "VACANTE" in texto_upper or "VACACIONES" in texto_upper:
        return None
    if "ASISTIÓ" in texto_upper:
        return TEXTO_PENDIENTE_ENTRADA
    return limpiar_motivo(comentario_crudo, "Falta")


def _valor_auto_salida(comentario_crudo):
    if comentario_crudo is None:
        return None
    if "ASISTIÓ" in str(comentario_crudo).upper():
        return TEXTO_PENDIENTE_SALIDA
    return comentario_crudo


def _valor_auto_entrada_resuelto(comentario_crudo):
    if comentario_crudo is not None and str(comentario_crudo).strip().upper() == "ASISTIÓ":
        return None
    return comentario_crudo


def _valor_auto_salida_resuelto(comentario_crudo):
    if comentario_crudo is not None and str(comentario_crudo).strip().upper() in (
        "ASISTIÓ", "CORRECCIÓN DESDE REPORTE DIARIO 9AM (SALIDA DÍA ANTERIOR)",
    ):
        return None
    return comentario_crudo


FILA_HEADER_REPORTE = 5


def _agregar_fila_t3(ws_t3, fila_libre, dni, fecha, comentario, hora_ent=None, hora_sal=None):
    ws_t3.cell(row=fila_libre, column=1, value=str(dni))
    ws_t3.cell(row=fila_libre, column=2, value=str(fecha))
    ws_t3.cell(row=fila_libre, column=4, value=comentario)
    ws_t3.cell(row=fila_libre, column=5, value="Reporte 9am")
    ws_t3.cell(row=fila_libre, column=6, value=str(pd.Timestamp.now().strftime("%H:%M")))
    if hora_ent:
        ws_t3.cell(row=fila_libre, column=7, value=str(hora_ent))
    if hora_sal:
        ws_t3.cell(row=fila_libre, column=8, value=str(hora_sal))


def main():
    hoy = pd.Timestamp.today().normalize()
    if len(sys.argv) > 1:
        hoy = pd.Timestamp(sys.argv[1])
    anterior = dia_habil_anterior(hoy, cargar_feriados())
    comentarios_ent_auto = comentarios_supervisor_dia(hoy)
    comentarios_sal_auto = comentarios_supervisor_dia(anterior)
    comentario_hist_auto = comentario_historico_dia(anterior)

    m = leer_excel("ASISTENCIA/MAC/1_Maestro_Headcount.xlsx", sheet_name="Maestro Headcount", header=3).dropna(how="all")
    m.columns = [str(c).strip() for c in m.columns]
    m = m.rename(columns={m.columns[0]: "DNI"})
    # Mismo gotcha que parseo_headcount.py -- si ALGUNA fila de DNI viene
    # vacía, pandas sube toda la columna a float64 y sobrevive al filtro,
    # dejando "18074336.0" en vez de "18074336" para todos los DNIs válidos.
    dni_num = pd.to_numeric(m["DNI"], errors="coerce")
    m = m[dni_num.notna()].copy()
    m["DNI"] = dni_num[dni_num.notna()].astype("int64").astype(str)
    dnis_vacante = set(m.loc[m["Estado"].astype(str).str.strip() == "Vacante", "DNI"])

    nombre_archivo = f"Reporte_Asistencia_{hoy.date()}.xlsx"
    try:
        contenido = descargar(RUTA_GRAPH_MAC + CARPETA_AUDITORIA_GRAPH + nombre_archivo)
    except Exception:
        print(f"{nombre_archivo} todavía no existe en SharePoint (auditoría) -- nada que aplicar todavía.")
        return
    wb_rep = openpyxl.load_workbook(io.BytesIO(contenido))
    ws_rep = wb_rep.active

    headers = [c.value for c in ws_rep[FILA_HEADER_REPORTE]]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    wb_t3 = openpyxl.load_workbook(io.BytesIO(descargar(TABLA3_RUTA_GRAPH)))
    ws_t3 = wb_t3["Registro diario supervisor"]
    fila_libre = ws_t3.max_row + 1

    filas_t3_existentes = set()
    # Dedup de correcciones de SOLO hora (sin comentario) -- el guard de
    # arriba (filas_t3_existentes, por texto de comentario) no cubre este
    # caso. Si subir_in_place(T3) sale bien pero la subida del reporte
    # falla justo después, la próxima corrida no se entera de que la hora
    # ya se aplicó (el reporte sigue con la celda sin limpiar) y la
    # reagrega duplicada a T3.
    filas_t3_horas_ent = set()
    filas_t3_horas_sal = set()
    for row in ws_t3.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        row_dni = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0]).strip()
        filas_t3_existentes.add((row_dni, str(row[1]), row[3]))
        if len(row) > 6 and row[6]:
            filas_t3_horas_ent.add((row_dni, str(row[1]), str(row[6])))
        if len(row) > 7 and row[7]:
            filas_t3_horas_sal.add((row_dni, str(row[1]), str(row[7])))

    n_aplicadas = 0
    for r in range(FILA_HEADER_REPORTE + 1, ws_rep.max_row + 1):
        dni = ws_rep.cell(row=r, column=idx["Dni"]).value
        if not dni:
            continue
        hora_ent_corr = ws_rep.cell(row=r, column=idx["Hora entrada corregida"]).value
        hora_sal_corr = ws_rep.cell(row=r, column=idx["Hora salida corregida"]).value
        comentario_ent = ws_rep.cell(row=r, column=idx["Comentario Entrada"]).value
        comentario_sal = ws_rep.cell(row=r, column=idx["Comentario salida"]).value
        tiene_entrada_real = ws_rep.cell(row=r, column=idx["Hora entrada real"]).value is not None
        tiene_salida_real = ws_rep.cell(row=r, column=idx["Hora salida real anterior"]).value is not None

        dni_str = str(int(dni)) if isinstance(dni, float) else str(dni)
        es_vacante = dni_str in dnis_vacante
        if es_vacante:
            auto_ent = None
        elif tiene_entrada_real:
            auto_ent = _valor_auto_entrada_resuelto(comentarios_ent_auto.get(dni_str))
        else:
            auto_ent = _valor_auto_entrada(comentarios_ent_auto.get(dni_str))
        crudo_sal = comentario_hist_auto.get(dni_str) or comentarios_sal_auto.get(dni_str)
        if es_vacante:
            auto_sal = "Vacante"
        elif tiene_salida_real:
            auto_sal = _valor_auto_salida_resuelto(crudo_sal)
        else:
            auto_sal = _valor_auto_salida(crudo_sal)
        if comentario_ent == auto_ent and (dni_str, str(hoy.date()), comentario_ent) in filas_t3_existentes:
            comentario_ent = None
        if comentario_sal == "Faltó (sin motivo registrado)":
            comentario_sal = None
        elif comentario_sal == auto_sal and (dni_str, str(anterior.date()), comentario_sal) in filas_t3_existentes:
            comentario_sal = None

        motivo_vac_vacac = None
        crudo_ent = comentarios_ent_auto.get(dni_str)
        if crudo_ent and not tiene_entrada_real:
            texto_up = str(crudo_ent).upper()
            if "VACANTE" in texto_up:
                motivo_vac_vacac = "Vacante"
            elif "VACACIONES" in texto_up:
                motivo_vac_vacac = "Vacaciones"

        if hora_ent_corr or comentario_ent:
            ya_en_t3 = (
                hora_ent_corr and not comentario_ent
                and (dni_str, str(hoy.date()), str(hora_ent_corr)) in filas_t3_horas_ent
            )
            if not ya_en_t3:
                texto = comentario_ent or "Corrección desde reporte diario 9am (entrada)"
                _agregar_fila_t3(ws_t3, fila_libre, dni, hoy.date(), texto, hora_ent=hora_ent_corr)
                fila_libre += 1
                n_aplicadas += 1
            ws_rep.cell(row=r, column=idx["Hora entrada corregida"]).value = None
            ws_rep.cell(row=r, column=idx["Comentario Entrada"]).value = None
        elif motivo_vac_vacac and (dni_str, str(hoy.date()), motivo_vac_vacac) not in filas_t3_existentes:
            _agregar_fila_t3(ws_t3, fila_libre, dni, hoy.date(), motivo_vac_vacac)
            fila_libre += 1
            n_aplicadas += 1

        if hora_sal_corr or comentario_sal:
            ya_en_t3_sal = (
                hora_sal_corr and not comentario_sal
                and (dni_str, str(anterior.date()), str(hora_sal_corr)) in filas_t3_horas_sal
            )
            if not ya_en_t3_sal:
                texto = comentario_sal or "Corrección desde reporte diario 9am (salida día anterior)"
                _agregar_fila_t3(ws_t3, fila_libre, dni, anterior.date(), texto, hora_sal=hora_sal_corr)
                fila_libre += 1
                n_aplicadas += 1
            ws_rep.cell(row=r, column=idx["Hora salida corregida"]).value = None
            ws_rep.cell(row=r, column=idx["Comentario salida"]).value = None

    if n_aplicadas == 0:
        print("No había correcciones/comentarios nuevos que aplicar.")
        return

    subir_in_place(TABLA3_RUTA_GRAPH, wb_t3)
    subir_creando_si_no_existe(RUTA_GRAPH_MAC + CARPETA_AUDITORIA_GRAPH + nombre_archivo, wb_rep)
    print(f"{n_aplicadas} fila(s) agregada(s) a {TABLA3_PATH}.")

    aqui = os.path.dirname(os.path.abspath(__file__))
    print("\nCorriendo motor_clasificacion.py para aplicar al histórico...")
    r1 = subprocess.run([sys.executable, os.path.join(aqui, "motor_clasificacion.py")])
    if r1.returncode != 0:
        print("\nAviso: motor_clasificacion.py falló -- la corrección ya quedó en Tabla 3, "
              "pero no se aplicó al histórico ni se regeneró el reporte. Revisa el error de arriba.")
        return

    print("\nRegenerando el reporte diario...")
    r2 = subprocess.run([sys.executable, os.path.join(aqui, "reporte_diario_9am.py"), str(hoy.date())])
    if r2.returncode != 0:
        print("\nAviso: reporte_diario_9am.py falló al regenerar el reporte. Revisa el error de arriba.")


if __name__ == "__main__":
    main()
