# -*- coding: utf-8 -*-
"""
Fase 5: mismo "puente" Postgres -> Excel que MAC/exportar_dimensiones_excel.py
(Fase 2), pero pensado para correr en GitHub Actions (sin laptop) en vez de
local -- por eso vive DENTRO del repo mac-cloud (un workflow de Actions solo
puede correr lo que está en el repo) e importa `graph_client` (credenciales
por variable de entorno) en vez de `graph_excel` (que lee un archivo local
graph_credentials.local.env que no existe en el runner).

MAC/exportar_dimensiones_excel.py (el original) se deja intacto como
respaldo local -- ver ese archivo para la nota completa sobre el problema de
sincronización de OneDrive que motivó escribir por Graph en vez de solo local.

Uso: python exportar_dimensiones.py
(usa DATABASE_URL, TENANT_ID, CLIENT_ID, CLIENT_SECRET del entorno)
"""
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import text

from dimension_models import (
    CatalogoMotivo, Feriado, HistorialCambio, Persona, PatronRecurrente, get_session,
)
from excel_safety import texto_seguro_excel
from graph_client import subir_in_place

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "pipeline"))
from db_lock import adquirir_lock, liberar_lock  # noqa: E402

NOMBRE_LOCK = "exportar_dimensiones"
LOCK_MAX_MINUTOS = 10

RUTA_GRAPH_MAC = "ASISTENCIA/MAC/"

# Mismo alias que migrar_dimensiones_a_postgres.py -- mantener sincronizado si cambia.
NOMBRE_CORTO = {
    "40628345": "KARINA", "40817159": "DANIELA CABANILLAS", "80270250": "ANA CARBAJAL",
    "45694774": "ELIZABETH TICONA", "9919446": "EDITH CAMACHO", "74943612": "GUERRA MARIA",
}

DIAS_ORDEN = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def _escribir_tabla_dimension(path, hoja, nombre_tabla, columnas, filas, titulo, descripcion):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = descripcion
    ws["A2"].font = Font(italic=True, size=9, color="6B6154")

    header_row = 4
    for col_idx, nombre_col in enumerate(columnas, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=nombre_col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")

    for r_idx, fila in enumerate(filas, start=header_row + 1):
        for c_idx, valor in enumerate(fila, start=1):
            # texto_seguro_excel(): estas filas terminan con nombres/
            # motivos/comentarios que alguien tipeó libremente (Cargar
            # Headcount, Dar de baja, Historial de cambios) -- sin esto,
            # un valor que arranca con "=" se ejecuta como fórmula al
            # abrir el Excel (inyección de fórmulas, hallazgo 2026-08-24).
            ws.cell(row=r_idx, column=c_idx, value=texto_seguro_excel(valor))

    n_filas = len(filas) + 1
    ref = f"A{header_row}:{get_column_letter(len(columnas))}{max(header_row + n_filas - 1, header_row + 1)}"
    try:
        tabla = Table(displayName=nombre_tabla, ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tabla)
    except Exception:
        pass

    for col_idx, nombre_col in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, len(str(nombre_col)) + 4))
    ws.freeze_panes = f"A{header_row + 1}"

    subir_in_place(RUTA_GRAPH_MAC + path, wb)  # sin fallback local -- en un runner de GitHub Actions no tiene sentido (se descarta al terminar el job); si Graph falla, que falle el paso y se vea en el log del workflow


def _nombre_supervisor(personas_por_dni, supervisor_dni):
    if not supervisor_dni:
        return None
    sup = personas_por_dni.get(supervisor_dni)
    if sup is None:
        return None
    return NOMBRE_CORTO.get(supervisor_dni, sup.nombre_completo)


def exportar_maestro(session, personas_por_dni):
    columnas = [
        "DNI", "Nombre completo", "Fecha de ingreso", "Fecha de baja", "Estado",
        "Reemplaza_a (DNI)", "Es reingreso (Sí/No)", "Rol", "Canal", "Región",
        "Ciudad / Mercado", "Zona / Ruta asignada", "Supervisor asignado",
        "Correo corporativo", "Motivo de baja", "Dado de baja por", "Registrado por", "Fecha de registro",
    ]
    filas = []
    for p in session.query(Persona).order_by(Persona.dni).all():
        filas.append([
            int(p.dni) if p.dni.isdigit() else p.dni, p.nombre_completo, p.fecha_ingreso, p.fecha_baja,
            p.estado, (int(p.reemplaza_a_dni) if p.reemplaza_a_dni and p.reemplaza_a_dni.isdigit() else p.reemplaza_a_dni),
            "Sí" if p.es_reingreso else "No", p.rol, p.canal, p.region, p.ciudad, p.zona,
            _nombre_supervisor(personas_por_dni, p.supervisor_dni), p.correo, p.motivo_baja,
            p.dado_de_baja_por, p.registrado_por, p.fecha_registro,
        ])
    _escribir_tabla_dimension(
        "1_Maestro_Headcount.xlsx", "Maestro Headcount", "MaestroHeadcount", columnas, filas,
        "Maestro Headcount", "Regenerado automáticamente desde Postgres -- no editar a mano (usar agregar_reemplazo.py o el panel de mac_cloud).",
    )
    return len(filas)


def exportar_patron(session):
    columnas = ["DNI", "Día de la semana", "Hora entrada programada", "Hora salida programada", "Canal del día", "Refrigerio"]
    filas = []
    orden = {d: i for i, d in enumerate(DIAS_ORDEN)}
    registros = session.query(PatronRecurrente).all()
    registros.sort(key=lambda r: (r.dni, orden.get(r.dia_semana, 99)))
    for r in registros:
        dni_int = int(r.dni) if r.dni.isdigit() else r.dni
        filas.append([dni_int, r.dia_semana, r.hora_entrada_prog, r.hora_salida_prog, r.canal_dia, r.refrigerio])
    _escribir_tabla_dimension(
        "2A_Patron_Recurrente.xlsx", "Patrón recurrente", "PatronRecurrente", columnas, filas,
        "Patrón Recurrente", "Regenerado automáticamente desde Postgres -- no editar a mano.",
    )
    return len(filas)


def exportar_vacantes(session, personas_por_dni):
    columnas = [
        "Zona / Ruta", "Rol", "Canal", "Región / Ciudad", "Supervisor asignado",
        "Fecha desde que está vacante", "Motivo de la baja previa", "Prioridad de cobertura", "Estado",
    ]
    filas = []
    resultado = session.execute(text("SELECT * FROM vacantes ORDER BY dni")).mappings().all()
    for v in resultado:
        region_ciudad = " / ".join(x for x in (v["region"], v["ciudad"]) if x)
        filas.append([
            v["zona"], v["rol"], v["canal"], region_ciudad or None,
            _nombre_supervisor(personas_por_dni, v["supervisor_dni"]),
            v["fecha_baja"].strftime("%d/%m/%Y") if v["fecha_baja"] else None,
            v["motivo_baja"], v["prioridad_cobertura"], v["estado_cobertura"],
        ])
    _escribir_tabla_dimension(
        "6_Vacantes.xlsx", "Vacantes", "Vacantes", columnas, filas,
        "Vacantes", "Regenerado automáticamente desde Postgres (vista `vacantes` = personas con estado='Vacante') -- no editar a mano.",
    )
    return len(filas)


def exportar_catalogo(session):
    columnas = ["Motivo", "Categoría", "Requiere sustento", "Efecto en indicador de asistencia", "Efecto en horas trabajadas"]
    filas = [
        [c.motivo, c.categoria, c.requiere_sustento, c.efecto_indicador, c.efecto_horas]
        for c in session.query(CatalogoMotivo).order_by(CatalogoMotivo.id).all()
    ]
    _escribir_tabla_dimension(
        "4_Catalogo_Motivos.xlsx", "Catalogo Motivos", "CatalogoMotivos", columnas, filas,
        "Catálogo de Motivos", "Regenerado automáticamente desde Postgres -- no editar a mano.",
    )
    return len(filas)


def exportar_historial(session):
    columnas = ["DNI", "Campo", "Valor nuevo", "Fecha desde", "Fecha hasta (opcional)", "Comentario", "Día de la semana (opcional)"]
    filas = []
    for h in session.query(HistorialCambio).order_by(HistorialCambio.dni, HistorialCambio.fecha_desde).all():
        dni_int = int(h.dni) if h.dni.isdigit() else h.dni
        filas.append([dni_int, h.campo, h.valor_nuevo, h.fecha_desde, h.fecha_hasta, h.comentario, h.dia_semana])
    _escribir_tabla_dimension(
        "8_Historial_Cambios.xlsx", "Historial de cambios", "HistorialCambios", columnas, filas,
        "Historial de Cambios", "Regenerado automáticamente desde Postgres -- editar vía Postgres/mac_cloud, no acá.",
    )
    return len(filas)


def exportar_feriados(session):
    columnas = ["Fecha", "Día de la semana", "Motivo", "Feriado nacional oficial"]
    filas = [
        [f.fecha, f.dia_semana, f.motivo, "Sí" if f.feriado_nacional else "No"]
        for f in session.query(Feriado).order_by(Feriado.fecha).all()
    ]
    _escribir_tabla_dimension(
        "5_Feriados_2026.xlsx", "Feriados 2026", "Feriados2026", columnas, filas,
        "Feriados 2026", "Regenerado automáticamente desde Postgres -- editar vía Postgres/mac_cloud, no acá.",
    )
    return len(filas)


def main():
    session = get_session()
    try:
        personas_por_dni = {p.dni: p for p in session.query(Persona).all()}
        n_m = exportar_maestro(session, personas_por_dni)
        n_p = exportar_patron(session)
        n_v = exportar_vacantes(session, personas_por_dni)
        n_c = exportar_catalogo(session)
        n_h = exportar_historial(session)
        n_f = exportar_feriados(session)
        print(f"Maestro: {n_m} | Patrón: {n_p} | Vacantes: {n_v} | Catálogo: {n_c} | Historial: {n_h} | Feriados: {n_f}")
        print("6 archivos Excel regenerados desde Postgres (via GitHub Actions).")
    finally:
        session.close()


if __name__ == "__main__":
    # Corre cada 15 min via GitHub Actions -- sin candado, una corrida lenta
    # (latencia de Graph en los 6 uploads) puede superponerse con la
    # siguiente y dejar los 6 Excel reflejando 2 estados distintos de
    # Postgres entre sí (mismo patrón que pipeline_completo.py, ver
    # pipeline/db_lock.py).
    ok, motivo = adquirir_lock(NOMBRE_LOCK, "github-actions", max_minutos=LOCK_MAX_MINUTOS)
    if not ok:
        print(f"exportar_dimensiones no iniciado: {motivo}")
        sys.exit(0)
    try:
        main()
    finally:
        liberar_lock(NOMBRE_LOCK)
