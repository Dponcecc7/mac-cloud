# -*- coding: utf-8 -*-
"""
Fase 4: carga de Headcount (Maestro + Patrón) por otros analistas, no solo
Davor -- cada quien sube su propio equipo. Los datos quedan separados por
`Persona.analista_propietario` (el correo de quien los subió): un DNI nuevo
se crea a nombre del que lo sube; un DNI que ya existe y es de OTRO analista
nunca se pisa en silencio -- se reporta como conflicto para revisión manual
(probablemente un DNI mal tipeado).
"""
import datetime as dt
import io

import openpyxl
import pandas as pd
from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user
from openpyxl.styles import Font

from dimension_models import Persona, PatronRecurrente, PersonaSupervisorCanal, PersonaZonaCanal, get_session
from github_actions import disparar_workflow
from permisos import requiere_pagina
from scoping import canonizar_canal
from parseo_headcount import (
    COLUMNAS_MAESTRO_ESPERADAS, COLUMNAS_PATRON_ESPERADAS,
    parsear_maestro, parsear_patron, validar_columnas,
)

bp = Blueprint("cargas", __name__, url_prefix="/cargas")


def _dni(valor):
    if pd.isna(valor):
        return None
    if isinstance(valor, float):
        return str(int(valor))
    return str(valor).strip()


def _texto(valor):
    if pd.isna(valor):
        return None
    return str(valor).strip() or None


def _fecha(valor):
    if pd.isna(valor):
        return None
    if isinstance(valor, dt.date) and not isinstance(valor, dt.datetime):
        return valor
    return pd.Timestamp(valor).date()


def _si_no_a_bool(valor):
    return str(valor).strip().upper() == "SÍ" if pd.notna(valor) else False


def _normalizar_nombre(nombre):
    # .split() sin args colapsa CUALQUIER espacio en blanco Unicode,
    # incluido \xa0 (espacio de no separación, típico al copiar/pegar desde
    # Word/Excel) -- sin esto, "MARIA ... DE\xa0LA\xa0CRUZ" (como quedó
    # guardado el nombre de Maria Plasencia) nunca matchea contra el mismo
    # nombre tipeado con espacios normales, en silencio (Davor, 2026-09-01).
    return " ".join(str(nombre).split()).upper() if nombre else ""


def _hora(valor):
    return valor if pd.notna(valor) else None


# Columnas OPCIONALES del Maestro Headcount (Davor, 2026-08-29) -- para
# mercaderistas compartidos entre canales con un supervisor real distinto
# por Tradicional que por Farmacia/AU (fijo por canal, no varía según qué
# canal_dia le toque trabajar ese día puntual -- ver
# scoping.overrides_supervisor_canal() y dimension_models.PersonaSupervisorCanal).
# No están en COLUMNAS_MAESTRO_ESPERADAS a propósito -- son opcionales,
# el resto de los analistas sigue subiendo su Excel de siempre sin esto.
COL_SUP_FARMACIA_AU = "Supervisor Farmacia/AU"
COL_SUP_TRADICIONAL = "Supervisor Tradicional"
# Mismo mecanismo para Zona (Davor, 2026-09-04: "Misma lógica para zonas,
# ya que tienen zonas por canal también") -- ver dimension_models.PersonaZonaCanal.
COL_ZONA_FARMACIA_AU = "Zona Farmacia/AU"
COL_ZONA_TRADICIONAL = "Zona Tradicional"


def _resolver_supervisor_por_nombre(supervisores, nombre):
    """Busca un supervisor por nombre en `supervisores` (dict nombre
    normalizado -> dni, ya construido para TODA la tabla de personas, no
    solo las del analista que sube el archivo) -- el caso de uso es
    justamente referenciar a un supervisor de OTRO canal/analista (ej.
    Diego, Farmacia, apuntando al supervisor de Tradicional de un
    mercaderista compartido)."""
    if not nombre:
        return None
    return supervisores.get(_normalizar_nombre(nombre))


def crear_persona_individual(
    propietario, dni, nombre, rol, canal, region, ciudad, zona,
    supervisor_dni, correo, fecha_ingreso, patron_dias,
):
    """Alta de UNA persona nueva a Headcount sin pasar por el Excel de
    "Cargar Headcount" -- Davor, 2026-08-25: "si quiero agregar un
    mercaderista nuevo a mi personal, como lo haria... donde dice agregar
    reemplazo, debe haber agregar nuevo headcount". A diferencia de
    "Agregar reemplazo" (reemplazos.py), acá NO hace falta un DNI de
    vacante existente del cual heredar datos -- es headcount genuinamente
    nuevo, no la cobertura de una posición que ya existía.

    `patron_dias`: lista de dicts {dia_semana, hora_entrada, hora_salida,
    canal_dia, refrigerio} -- se ignoran los días sin hora de entrada Y
    salida (persona sin ese día en su semana laboral, mismo criterio que
    "sin fila = no trabaja ese día" que ya usa todo el pipeline)."""
    session = get_session()
    try:
        existente = session.get(Persona, dni)
        if existente and existente.analista_propietario and existente.analista_propietario != propietario:
            raise ValueError(f"El DNI {dni} ya existe y pertenece a otro analista ({existente.analista_propietario}).")
        if existente and existente.estado == "Activo":
            raise ValueError(f"El DNI {dni} ya está Activo en el sistema -- usá Historial de cambios para modificar sus datos.")

        es_reingreso = existente is not None
        persona = existente or Persona(dni=dni)
        if not es_reingreso:
            session.add(persona)
        persona.nombre_completo = nombre
        persona.rol = rol
        persona.canal = canonizar_canal(canal)
        persona.region = region
        persona.ciudad = ciudad
        persona.zona = zona
        persona.supervisor_dni = supervisor_dni
        persona.correo = correo
        persona.fecha_ingreso = fecha_ingreso
        persona.fecha_baja = None
        persona.estado = "Activo"
        persona.es_reingreso = es_reingreso
        persona.motivo_baja = None
        persona.registrado_por = propietario
        persona.fecha_registro = dt.date.today()
        persona.analista_propietario = propietario
        session.flush()  # la persona ya debe existir en la sesion antes de tocar Patron (FK)

        session.query(PatronRecurrente).filter_by(dni=dni).delete()
        session.flush()
        n_patron = 0
        for fila in patron_dias:
            if not fila.get("hora_entrada") or not fila.get("hora_salida"):
                continue
            session.add(PatronRecurrente(
                dni=dni, dia_semana=fila["dia_semana"],
                hora_entrada_prog=fila["hora_entrada"], hora_salida_prog=fila["hora_salida"],
                canal_dia=fila.get("canal_dia") or canal, refrigerio=fila.get("refrigerio"),
            ))
            n_patron += 1

        session.commit()
        return es_reingreso, n_patron
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _plantilla_excel(nombre_hoja, titulo, columnas, fila_ejemplo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Plantilla en blanco -- completá una fila por persona/día y subila en \"Cargar Headcount\".")
    for i, col in enumerate(columnas, start=1):
        celda = ws.cell(row=4, column=i, value=col)
        celda.font = Font(bold=True)
        ws.column_dimensions[celda.column_letter].width = max(14, len(col) + 2)
    for i, valor in enumerate(fila_ejemplo, start=1):
        celda = ws.cell(row=5, column=i, value=valor)
        celda.font = Font(italic=True, color="9C6B13")
    ws.cell(row=6, column=1, value="↑ borrá esta fila de ejemplo antes de subir el archivo").font = Font(italic=True, size=9, color="9C6B13")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.get("/plantilla/maestro.xlsx")
@requiere_pagina("cargar_headcount")
def plantilla_maestro():
    # DNI de ejemplo = "EJEMPLO" (no numérico) a propósito -- parsear_maestro()
    # descarta filas cuyo DNI no es numérico, así que si alguien sube el
    # archivo sin borrar la fila de ejemplo, no crea una persona fantasma.
    buf = _plantilla_excel(
        "Maestro Headcount", "Maestro Headcount", COLUMNAS_MAESTRO_ESPERADAS + [COL_SUP_FARMACIA_AU, COL_SUP_TRADICIONAL],
        ["EJEMPLO", "APELLIDOS NOMBRES", dt.date.today(), None, "Activo", None, "No",
         "MERCADERISTAS", "FARMACIA", "LIMA", "LIMA", "NOMBRE DE ZONA", "NOMBRE DEL SUPERVISOR",
         None, None, "Analista MAC", dt.date.today(), None, None],
    )
    return send_file(buf, as_attachment=True, download_name="Plantilla_Maestro_Headcount.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.get("/plantilla/patron.xlsx")
@requiere_pagina("cargar_headcount")
def plantilla_patron():
    buf = _plantilla_excel(
        "Patrón recurrente", "Patrón Recurrente", COLUMNAS_PATRON_ESPERADAS,
        ["EJEMPLO", "lunes", dt.time(8, 0), dt.time(17, 30), "Farmacia", "Con refrigerio"],
    )
    return send_file(buf, as_attachment=True, download_name="Plantilla_Patron_Recurrente.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.get("/headcount")
@requiere_pagina("cargar_headcount")
def headcount_form():
    return render_template("cargas_headcount.html", usuario=current_user, resultado=False)


@bp.post("/headcount")
@requiere_pagina("cargar_headcount")
def headcount_submit():
    archivo_maestro = request.files.get("maestro")
    archivo_patron = request.files.get("patron")
    if not archivo_maestro or archivo_maestro.filename == "":
        flash("Tenés que subir el archivo de Maestro Headcount.", "error")
        return redirect(url_for("cargas.headcount_form"))

    propietario = current_user.email

    try:
        m = parsear_maestro(io.BytesIO(archivo_maestro.read()))
        validar_columnas(m, COLUMNAS_MAESTRO_ESPERADAS, archivo_maestro.filename)
    except Exception as e:
        flash(f"Error leyendo el Maestro Headcount: {e}", "error")
        return redirect(url_for("cargas.headcount_form"))

    p = None
    if archivo_patron and archivo_patron.filename:
        try:
            p = parsear_patron(io.BytesIO(archivo_patron.read()))
            validar_columnas(p, COLUMNAS_PATRON_ESPERADAS, archivo_patron.filename)
        except Exception as e:
            flash(f"Error leyendo el Patrón Recurrente: {e}", "error")
            return redirect(url_for("cargas.headcount_form"))

    canal_propio = (getattr(current_user, "canal_asignado", None) or "").strip().upper() or None

    session = get_session()
    try:
        nuevas, actualizadas, conflictos, compartidos = [], [], [], []
        # startswith("SUPERVISOR") en vez de == "SUPERVISORES" a propósito
        # (Davor, 2026-09-01: cargó "Supervisor asignado" para 82 personas
        # de Autoservicios/Multicanal y quedó vacío en TODAS, sin ningún
        # error -- encontramos que 2 supervisores de Autoservicios estaban
        # guardados con Rol="SUPERVISOR" en singular, así que el match
        # exacto contra "SUPERVISORES" nunca los encontraba y el DNI
        # simplemente no se aplicaba, en silencio). Sigue sin matchear
        # "Asesora"/"Mercaderistas"/etc, solo tolera la variante singular.
        #
        # Sin filtro por analista_propietario a propósito (Davor, 2026-09-01:
        # Jesus Marquez y Maria Plasencia De La Cruz, ambos de Diego, también
        # supervisan gente de Autoservicios de Davor -- "permitir que estén
        # en ambas cargas"). Un supervisor puede tener gente a cargo en el
        # headcount de MÁS DE UN analista, así que el universo de supervisores
        # resolubles es TODA la tabla Persona, no solo los del uploader.
        # _normalizar_nombre (no .strip().upper() a secas) en las claves:
        # Maria Plasencia quedó guardada con \xa0 en vez de espacio normal
        # entre "DE"/"LA"/"CRUZ" (típico de copiar/pegar desde Word) y sin
        # esto el nombre tipeado en el Excel (con espacios normales) nunca
        # matcheaba, en silencio (Davor, 2026-09-01).
        supervisores_propios = {
            _normalizar_nombre(per.nombre_completo): per.dni
            for per in session.query(Persona).all()
            if (per.rol or "").strip().upper().startswith("SUPERVISOR")
        }
        # Los supervisores de ESTE MISMO archivo se resuelven antes del loop
        # principal, sin importar en qué fila aparezcan -- antes, si el Excel
        # listaba a los supervisores DESPUÉS de sus mercaderistas (orden
        # común: armar la planilla por equipo y agregar el supervisor al
        # final), esas filas quedaban con supervisor_dni=None porque el
        # supervisor todavía no se había "visto" en el loop (Davor,
        # 2026-08-29: "le colocó supervisor pero no se visualiza" -- Diego
        # cargó a Chuchon y María como filas SUPERVISORES al final).
        for _, fila_sup in m[m["Rol"].astype(str).str.strip().str.upper().str.startswith("SUPERVISOR")].iterrows():
            nombre_sup, dni_sup = _texto(fila_sup["Nombre completo"]), _dni(fila_sup["DNI"])
            if nombre_sup and dni_sup:
                supervisores_propios[_normalizar_nombre(nombre_sup)] = dni_sup

        pendientes_supervisor = []  # [(dni, supervisor_dni), ...] -- ver comentario mas abajo
        overrides_canal = []  # [(dni, canal, supervisor_dni), ...] -- ver COL_SUP_FARMACIA_AU/TRADICIONAL
        overrides_canal_zona = []  # [(dni, canal, zona), ...] -- ver COL_ZONA_FARMACIA_AU/TRADICIONAL
        supervisores_no_encontrados = set()
        for _, row in m.iterrows():
            dni = _dni(row["DNI"])
            existente = session.get(Persona, dni)
            # Mercaderista compartido entre canales (Davor, 2026-08-27): si
            # el dueño actual es OTRO analista pero de un canal DISTINTO al
            # mío, no es un error de tipeo -- es la misma persona trabajando
            # ambos canales en días distintos (ver PatronRecurrente.canal_dia
            # mas abajo). No le piso sus datos base (nombre/rol/etc, siguen
            # siendo del dueño original); solo se van a fusionar mis días de
            # Patrón para este DNI. Si el canal es el MISMO, sigue siendo
            # conflicto real (probable DNI mal tipeado).
            es_compartido = (
                existente is not None and existente.analista_propietario != propietario
                and canal_propio and (existente.canal or "").strip().upper() != canal_propio
            )
            if existente and existente.analista_propietario != propietario and not es_compartido:
                conflictos.append({
                    "dni": dni, "nombre": _texto(row["Nombre completo"]),
                    "dueño_actual": existente.analista_propietario,
                })
                continue

            # Overrides de supervisor por canal -- columnas opcionales del
            # Excel (Davor, 2026-08-29). Se procesan para cualquier fila que
            # no sea un conflicto real (propia o compartida) porque es
            # justamente en los mercaderistas compartidos entre canales
            # donde hace falta un supervisor distinto por canal.
            for columna, canales in ((COL_SUP_FARMACIA_AU, ("FARMACIA", "AUTOSERVICIO")), (COL_SUP_TRADICIONAL, ("TRADICIONAL",))):
                if columna not in m.columns:
                    continue
                nombre_sup = _texto(row[columna])
                if not nombre_sup:
                    continue
                dni_sup = _resolver_supervisor_por_nombre(supervisores_propios, nombre_sup)
                if dni_sup:
                    overrides_canal.extend((dni, canal, dni_sup) for canal in canales)
                else:
                    supervisores_no_encontrados.add(nombre_sup)

            # Overrides de zona por canal -- mismo mecanismo, columnas
            # opcionales aparte (Davor, 2026-09-04).
            for columna, canales in ((COL_ZONA_FARMACIA_AU, ("FARMACIA", "AUTOSERVICIO")), (COL_ZONA_TRADICIONAL, ("TRADICIONAL",))):
                if columna not in m.columns:
                    continue
                zona_canal = _texto(row[columna])
                if zona_canal:
                    overrides_canal_zona.extend((dni, canal, zona_canal) for canal in canales)

            if es_compartido:
                compartidos.append(dni)
                continue

            sup_texto = _texto(row["Supervisor asignado"])
            supervisor_dni = supervisores_propios.get(_normalizar_nombre(sup_texto)) if sup_texto else None
            nombre = _texto(row["Nombre completo"]) or "(sin nombre)"
            rol = _texto(row["Rol"])
            # Canal forzado al del analista (Davor, 2026-08-27): antes se
            # confiaba en lo que decía la columna "Canal" del Excel -- así
            # terminó "todo cargado como Tradicional" pase lo que pase que
            # tipeara cada quien. Si el analista tiene canal_asignado (Diego,
            # Yeny), se ignora la columna del Excel y se usa el suyo; si no
            # (Kevin, Davor) se sigue respetando lo tipeado, sin cambio de
            # comportamiento.
            # canonizar_canal() (Davor, 2026-09-04: "Homologa autoservicio y
            # autoservicios"): el Excel trae texto libre -- "Autoservicios"
            # (plural) y "Autoservicio" (singular) son el MISMO canal pero
            # sin esto quedaban guardados como 2 valores distintos.
            canal = canonizar_canal(canal_propio or _texto(row["Canal"]))

            datos = dict(
                nombre_completo=nombre, rol=rol, canal=canal, region=_texto(row["Región"]),
                ciudad=_texto(row["Ciudad / Mercado"]), zona=_texto(row["Zona / Ruta asignada"]),
                # supervisor_dni se aplica DESPUÉS del flush de abajo, no
                # acá -- si el supervisor está en ESTE MISMO archivo pero en
                # una fila posterior (caso real, Davor 2026-08-29: Diego
                # cargó a sus mercaderistas primero y a Chuchon/María como
                # SUPERVISORES al final), insertar la fila del mercaderista
                # con supervisor_dni ya apuntando a un DNI que TODAVÍA no
                # existe como fila en la tabla revienta la FK
                # (personas_supervisor_dni_fkey) -- SQLAlchemy no sabe
                # reordenar el INSERT porque supervisor_dni es un DNI suelto,
                # no una relationship() que le indique la dependencia.
                supervisor_dni=None, correo=_texto(row["Correo corporativo"]),
                fecha_ingreso=_fecha(row["Fecha de ingreso"]), fecha_baja=_fecha(row["Fecha de baja"]),
                estado=(_texto(row["Estado"]) or "Inactivo"),
                es_reingreso=_si_no_a_bool(row["Es reingreso (Sí/No)"]),
                motivo_baja=_texto(row["Motivo de baja"]), registrado_por=_texto(row["Registrado por"]) or propietario,
                fecha_registro=_fecha(row["Fecha de registro"]) or dt.date.today(),
                analista_propietario=propietario,
            )
            if existente:
                for k, v in datos.items():
                    setattr(existente, k, v)
                actualizadas.append(dni)
            else:
                session.add(Persona(dni=dni, **datos))
                nuevas.append(dni)
            if supervisor_dni:
                pendientes_supervisor.append((dni, supervisor_dni))

        session.flush()  # las personas nuevas ya deben existir en la sesion antes de aplicar supervisor_dni/tocar Patron (FK)

        # Ahora sí -- toda fila de este archivo ya existe como fila real en
        # la tabla, así que apuntar supervisor_dni a cualquiera de ellas
        # (sin importar el orden en que aparecían en el Excel) ya no puede
        # violar la FK.
        for dni, supervisor_dni in pendientes_supervisor:
            session.get(Persona, dni).supervisor_dni = supervisor_dni

        n_overrides = 0
        for dni, canal, supervisor_dni in overrides_canal:
            existente_ov = session.query(PersonaSupervisorCanal).filter_by(dni=dni, canal=canal).first()
            if existente_ov:
                existente_ov.supervisor_dni = supervisor_dni
            else:
                session.add(PersonaSupervisorCanal(dni=dni, canal=canal, supervisor_dni=supervisor_dni))
            n_overrides += 1

        for dni, canal, zona_canal in overrides_canal_zona:
            existente_zc = session.query(PersonaZonaCanal).filter_by(dni=dni, canal=canal).first()
            if existente_zc:
                existente_zc.zona = zona_canal
            else:
                session.add(PersonaZonaCanal(dni=dni, canal=canal, zona=zona_canal))

        n_patron = 0
        if p is not None:
            dnis_propios = set(nuevas) | set(actualizadas) | set(compartidos)
            filas_patron = [
                (dni, _texto(row["Día de la semana"]),
                 _hora(row["Hora entrada programada"]), _hora(row["Hora salida programada"]),
                 canal_propio or _texto(row["Canal del día"]), _texto(row["Refrigerio"]))
                for dni, row in ((_dni(r["DNI"]), r) for _, r in p.iterrows())
                if dni in dnis_propios
            ]
            # Se borra SOLO (dni, dia_semana) que se va a re-insertar, no
            # todo el patron del dni -- antes un upload PARCIAL (ej. Yeny
            # sube solo lunes/miércoles/viernes de un compartido) borraba
            # también los días de Diego que no venían en su archivo. Con
            # esto, cada analista solo toca los días que trae su propio
            # Excel, sin importar de quién sean los demás días.
            for dni, dia, *_r in filas_patron:
                session.query(PatronRecurrente).filter_by(dni=dni, dia_semana=dia).delete()
            session.flush()
            for dni, dia, hora_ent, hora_sal, canal_dia, refrigerio in filas_patron:
                session.add(PatronRecurrente(
                    dni=dni, dia_semana=dia, hora_entrada_prog=hora_ent, hora_salida_prog=hora_sal,
                    canal_dia=canal_dia, refrigerio=refrigerio,
                ))
                n_patron += 1

        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()

    mensaje = (
        f"Carga completa: {len(nuevas)} personas nuevas, {len(actualizadas)} actualizadas, "
        f"{len(compartidos)} compartidas con otro canal (se fusionó su Patrón, sin tocar sus datos base), "
        f"{len(conflictos)} en conflicto, {n_patron} filas de Patrón agregadas."
    )
    if n_overrides:
        mensaje += f" {n_overrides} overrides de supervisor por canal guardados."
    if supervisores_no_encontrados:
        mensaje += (
            f" ADVERTENCIA: no se encontró ningún supervisor (Rol=SUPERVISORES) con el nombre exacto: "
            f"{', '.join(sorted(supervisores_no_encontrados))}."
        )
    # El motor de clasificación lee el Excel Maestro/Patrón (exportado desde
    # Postgres cada 15 min por exportar_dimensiones.yml), no Postgres
    # directo -- sin este disparo, lo recién cargado quedaba invisible en
    # Asistencia diaria hasta el próximo ciclo del cron (Davor, 2026-08-28:
    # "Diego ya cargó su headcount, pero no le parece en asistencia" -- sus
    # 12 personas nunca habían sido clasificadas ni una vez, porque este
    # endpoint nunca disparaba el export, a diferencia de "Agregar
    # headcount"/"Agregar reemplazo" que sí lo hacen).
    if nuevas or actualizadas or compartidos:
        ok_disparo, _ = disparar_workflow("exportar_dimensiones.yml")
        mensaje += (
            " Se va a reflejar en Asistencia diaria en unos minutos."
            if ok_disparo else
            " No se pudo iniciar la actualización automática -- probá \"Actualizar ahora\" desde Asistencia diaria en unos minutos."
        )
    flash(mensaje, "ok" if not conflictos else "error")
    return render_template(
        "cargas_headcount.html", usuario=current_user, resultado=True,
        nuevas=nuevas, actualizadas=actualizadas, conflictos=conflictos, compartidos=compartidos, n_patron=n_patron,
    )
