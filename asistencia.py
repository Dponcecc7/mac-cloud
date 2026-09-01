# -*- coding: utf-8 -*-
"""
Blueprint de corrección diaria de asistencia -- version web de
C:\\Archivos Python\\One Page\\asistencia_app\\app.py, para que varias
personas lo usen en vez de solo desde la laptop de Davor.

Diferencia de diseño clave respecto al original: la app local lee/escribe
Reportes_Diarios/Reporte_Asistencia_*.xlsx en disco y llama a
reporte_diario_9am.py/aplicar_correcciones.py como subprocesos -- nada de
eso existe en Render. Acá se lee directo de Postgres (clasificacion_diaria +
personas, ya mantenido fresco por el pipeline en la nube, Fase 6) y las
operaciones largas (Athena) se disparan como workflow de GitHub Actions
(github_actions.py) en vez de correrlas sincronicas en el request.

Guardar una corrección escribe DOBLE (decision del usuario, 2026-08-21):
Tabla 3 en SharePoint via Graph (la fuente REAL que lee el motor de
clasificacion, local y nube -- cero riesgo de romper el flujo que ya
funciona) y ademas una copia en Postgres (correcciones_web, solo auditoria).
"""
import datetime as dt
import io
import os
import re
import sys
from zoneinfo import ZoneInfo

import openpyxl
import requests
from flask import Blueprint, flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required

from dimension_models import CatalogoMotivo, CorreccionWeb, Feriado, Persona, Visita, get_session
from excel_safety import texto_seguro_excel
from fact_models import ClasificacionDiaria
from graph_client import descargar, subir_in_place
from github_actions import disparar_workflow, estado_ultima_corrida
from permisos import requiere_analista_admin, requiere_pagina
from scoping import CANALES_FILTRABLES, aplicar_filtros_extra, condicion_scope, overrides_supervisor_canal
from sqlalchemy.orm import aliased

_AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_AQUI, "pipeline"))
from db_lock import adquirir_lock, liberar_lock  # noqa: E402
import reporte_diario_9am as r9am  # noqa: E402 -- reusa reemplazos_pendientes()/dnis_alguna_vez_en_maestro(), no se duplica

bp = Blueprint("asistencia", __name__, url_prefix="/asistencia")

# "Agregar reemplazo"/"Agregar headcount"/"Dar de baja" dan de alta/baja
# gente de verdad en Postgres (no es solo ver o marcar asistencia), asi
# que quedan reservadas a analista/admin -- ver permisos.py para el
# porque de redirigir_a en vez del default "dashboard".
_analista_requerido = requiere_analista_admin(redirigir_a="asistencia.reporte")


RUTA_GRAPH_MAC = "ASISTENCIA/MAC/"
TABLA3_RUTA_GRAPH = f"{RUTA_GRAPH_MAC}3_Registro_Diario_Supervisor.xlsx"
SALIDA_ANTICIPADA_MIN = 10  # mismo umbral que usaba la app local para resaltar "salida temprana"

PERU_TZ = ZoneInfo("America/Lima")  # sin horario de verano -- offset fijo UTC-5

# aplicar_correcciones.py escribe este texto exacto en el comentario cuando
# un supervisor reporta "Asistió" desde la app movil pero todavia nadie
# confirmo la hora real en "Entrada/Salida corregida" -- hay que poder
# encontrar estos casos facil para no perderlos (ver TEXTO_PENDIENTE_ENTRADA/
# TEXTO_PENDIENTE_SALIDA en pipeline/aplicar_correcciones.py).
MARCADOR_PENDIENTE = "reportado por supervisor -- confirmar en"

# motor_clasificacion.py escribe este sufijo exacto en el Estado cuando el
# dato viene del botón "Marcar asistencia" de acá (Asistió/Vacante sin
# marcación real de GPS -- "Apoyo zona" tenía este mismo problema, por eso
# se reemplazó por "Día de descanso", ver ACCIONES_MARCAR más abajo) --
# mismo caso que MARCADOR_PENDIENTE pero para el flujo nuevo: acá el aviso
# vive en el propio Estado, no en el comentario del supervisor, asi que hay
# que buscarlo aparte.
MARCADOR_SIN_APP = "según supervisor, sin marcación app"

# Tabla 3 es append-only -- el motor arma cada (DNI, Fecha) fusionando todas
# sus filas y quedandose con el ultimo valor NO VACÍO de cada columna (ver
# motor_clasificacion.py::main(), el groupby de "sup"). Por eso no existe
# forma de "vaciar" un comentario escribiendo un texto vacío -- una celda
# vacía simplemente se ignora en esa fusión y el comentario viejo (equivocado)
# sigue ganando. La solución es escribir ESTE texto puntual: no vacío (así sí
# gana la fusión) pero tampoco empieza con "FALTA" ni contiene
# "VACANTE"/"VACACIONES" ni llena "Estado reportado" -- el motor no lo
# reconoce como ninguna de sus palabras clave, así que no fuerza ningún
# estado y la persona vuelve a clasificarse solo por su marcación real (o
# "Falta (sin marcación)" si de verdad no hay ninguna).
MARCADOR_BORRADO = "(comentario borrado desde mac_cloud)"

# Solo Asistió/Apoyo zona implican una hora real pendiente de escribir --
# Vacante/Falta no tienen entrada/salida que confirmar, asi que no deben
# mostrar el aviso aunque tambien vengan de "Marcar asistencia".
_ESTADOS_CON_HORA_PENDIENTE = ("ASISTIÓ", "APOYO ZONA")


def _estado_base(estado):
    return (estado or "").split(" (")[0]


def _canal_para_mostrar(marcado, canal_asignado):
    """"Canal hoy"/"Canal ayer" salen de canales_marcados (el canal REAL de
    la visita GPS) -- si la persona marco en un punto administrativo/censo
    (ver UBICACIONES_ADMINISTRATIVAS en pipeline/athena_client.py), el
    motor lo clasifica como "Otro" a proposito (no es Tradicional/Farmacia/
    Autoservicio de verdad). Mostrar "Otro" tal cual parece un dato
    faltante -- Davor, 2026-08-31: "igual deberia aparecer el canal, no
    debe salir Otro". Se usa el canal asignado de la persona en su lugar
    cuando no hay nada mas util que mostrar."""
    if not marcado or marcado.strip() == "Otro":
        return canal_asignado or marcado or ""
    return marcado

_RE_PREFIJO_FALTA = re.compile(r"^falta\s*[-–]?\s*", re.IGNORECASE)


def _homologar_motivo(texto):
    """Version tolerante de reporte_diario_9am.py::limpiar_motivo() solo
    para lo que se muestra/edita acá -- saca el prefijo "Falta" repetido con
    cualquier variante de formato (con/sin guion, con/sin espacios,
    mayúsculas), no solo el "Falta - " exacto que espera el motor. Ej.:
    "Falta-Falta Injustificada" -> "Injustificada". No toca limpiar_motivo()
    en sí (esa la usa el motor de clasificación real, no se arriesga)."""
    if not texto:
        return texto
    texto = str(texto).strip()
    while True:
        m = _RE_PREFIJO_FALTA.match(texto)
        if not m or not m.group(0) or m.end() == 0:
            break
        resto = texto[m.end():].strip()
        if resto == texto:
            break
        texto = resto
    # Recorta cualquier paréntesis final -- de paso, esto es lo que hace que
    # MARCADOR_BORRADO (que es 100% un paréntesis) se muestre vacío en el
    # reporte sin necesitar un caso especial acá.
    texto = re.sub(r"\s*\([^)]*\)\s*$", "", texto).strip()
    if not texto:
        return None
    return texto[0].upper() + texto[1:]


def _cargar_feriados():
    session = get_session()
    try:
        filas = session.query(Feriado.fecha).all()
    finally:
        session.close()
    return {f for (f,) in filas}


def _dia_habil_anterior(fecha, feriados_set):
    d = fecha - dt.timedelta(days=1)
    while d.weekday() == 6 or d in feriados_set:
        d -= dt.timedelta(days=1)
    return d


def _cargar_reporte(fecha, usuario_actual=None, rol_filtro=None, region_filtro=None, supervisor_filtro=None, ciudad_filtro=None, canal_filtro=None, salida_mismo_dia=False, dni_filtro=None):
    """Devuelve (resumen_dict, filas, frescura) para `fecha` -- None, None, None
    si no hay ninguna fila ese día (motor todavía no corrió para esa fecha).

    `dni_filtro` (Davor, 2026-09-01, "Histórico diario" -- ver día por día
    de UN mercaderista en vez de todos): acota a una sola persona, además
    del scope/filtros normales -- se usa llamando esta función una vez por
    día del rango elegido.

    `usuario_actual`: si se pasa, acota a las Personas visibles para ese
    usuario (ver scoping.py) -- sin esto, un analista de otro cliente veía
    nombres/DNI de todos los clientes, y un supervisor veía todo el equipo
    en vez de solo el suyo. `rol_filtro`/`region_filtro`/`supervisor_filtro`/
    `ciudad_filtro`: filtros adicionales de Marcar asistencia (solo
    admin/analista, ver _filtros_marcar()), encima del scope de acceso, no
    en vez de. `canal_filtro` (Davor, 2026-08-29) -- SOLO admin: "debo
    tener un filtro para ver Tradicional, Farmacia y AU" -- un analista de
    canal ya está acotado por condicion_scope(), no lo necesita.

    `salida_mismo_dia` (Davor, 2026-08-29) -- False (default, sin cambios
    de comportamiento): la Salida prog./real que se arma es la de AYER a
    propósito -- "Reporte diario"/"Marcar asistencia" se generan a media
    tarde, ANTES de que el turno de hoy cierre, entonces la salida de HOY
    todavía estaría vacía; se completa la fila con la salida YA CERRADA de
    ayer para que el reporte que se manda al cliente nunca tenga una
    columna en blanco. True: usa la salida del MISMO día que la entrada --
    para "Histórico de asistencia diaria" (Reportes), que mira un día
    puntual del pasado y necesita ver SU PROPIA salida, no la de ayer
    (Davor: "ahi me saldrá la información del día seleccionado su hora
    entrada y salida, no como esta vista que la salida es del día anterior")."""
    feriados = _cargar_feriados()
    ayer = _dia_habil_anterior(fecha, feriados)

    session = get_session()
    try:
        query_personas = session.query(Persona)
        cond_scope = condicion_scope(Persona, usuario_actual) if usuario_actual else None
        if cond_scope is not None:
            query_personas = query_personas.filter(cond_scope)
        query_personas = aplicar_filtros_extra(query_personas, Persona, rol_filtro, region_filtro, supervisor_filtro, ciudad_filtro, canal_filtro)
        if dni_filtro:
            query_personas = query_personas.filter(Persona.dni == dni_filtro)
        personas = {p.dni: p for p in query_personas.all()}
        nombre_de = {dni: p.nombre_completo for dni, p in personas.items()}
        # Nombre del SUPERVISOR -- consulta aparte, sin condicion_scope().
        # Antes se resolvía contra nombre_de (ya acotado al scope del
        # visor), así que un supervisor de OTRO canal/cliente (el
        # supervisor real de alguien "compartido" entre canales, ej.
        # MULTICANAL) salía en blanco -- el visor tiene acceso a ver a ESA
        # persona (ya pasó el scope), no hace falta que su supervisor
        # TAMBIÉN esté en su mismo scope para mostrar el nombre (Davor,
        # 2026-08-29: "No le aparece supervisor asignado, pero si les cargó
        # en el archivo headcount").
        # Override de supervisor por canal (Davor, 2026-08-29) -- caso
        # puntual: un mercaderista compartido entre canales puede tener un
        # supervisor real distinto por Tradicional que por Farmacia/AU, fijo
        # por canal (no varía según qué canal le toque trabajar ese día
        # puntual). Ver scoping.overrides_supervisor_canal().
        overrides_sup = overrides_supervisor_canal(session, list(personas.keys()), usuario_actual)
        dnis_supervisores = {p.supervisor_dni for p in personas.values() if p.supervisor_dni}
        dnis_supervisores |= set(overrides_sup.values())
        nombre_supervisor_de = {}
        if dnis_supervisores:
            nombre_supervisor_de = dict(
                session.query(Persona.dni, Persona.nombre_completo)
                .filter(Persona.dni.in_(dnis_supervisores)).all()
            )

        filas_hoy = (
            session.query(ClasificacionDiaria)
            .filter(ClasificacionDiaria.fecha == fecha, ClasificacionDiaria.dni.in_(personas.keys()))
            .all()
        )
        filas_ayer = {
            c.dni: c for c in
            session.query(ClasificacionDiaria)
            .filter(ClasificacionDiaria.fecha == ayer, ClasificacionDiaria.dni.in_(personas.keys()))
            .all()
        }
        # DNIs ya marcados hoy desde "Marcar asistencia" (correcciones_web) --
        # la Tabla 3 ya tiene su fila, pero clasificacion_diaria.comentario_supervisor
        # recien se actualiza cuando el motor vuelve a correr. Sin esto, la
        # persona seguia apareciendo en "Pendientes" varios minutos despues
        # de haberla marcado.
        dnis_marcados_hoy = {
            dni for (dni,) in session.query(CorreccionWeb.dni).filter(CorreccionWeb.fecha == fecha).all()
        }
        # Motivo/comentario recién guardado -- MISMO gap que dnis_marcados_hoy
        # pero para el TEXTO del motivo, no solo si ya se marcó (Davor,
        # 2026-08-29: "colocó los motivos de asistencia, aparecen todos falta
        # y luego de un rato recién aparece el motivo... cuando Diego
        # registra sus motivos, se demora en aparecerme"). guardar() escribe
        # esta tabla AL INSTANTE (antes incluso de subir a Tabla 3), así que
        # se puede mostrar de una en vez de esperar la próxima corrida del
        # motor -- mismo texto que el motor va a terminar guardando en
        # clasificacion_diaria.comentario_supervisor de todos modos, ya que
        # ambos salen de la misma fila que guardar() sube a Tabla 3.
        override_entrada, override_salida = {}, {}
        fecha_fuente_salida = fecha if salida_mismo_dia else ayer
        correcciones_recientes = (
            session.query(CorreccionWeb)
            .filter(CorreccionWeb.dni.in_(personas.keys()), CorreccionWeb.fecha.in_([fecha, fecha_fuente_salida]))
            .order_by(CorreccionWeb.fecha_registro.desc())
            .all()
        )
        for corr in correcciones_recientes:
            if corr.fecha == fecha and corr.comentario_entrada and corr.dni not in override_entrada:
                override_entrada[corr.dni] = corr.comentario_entrada
            if corr.fecha == fecha_fuente_salida and corr.comentario_salida and corr.dni not in override_salida:
                override_salida[corr.dni] = corr.comentario_salida
    finally:
        session.close()

    if not filas_hoy:
        return None, None, None

    filas = []
    ultima_sync = None
    for c in filas_hoy:
        p = personas.get(c.dni)
        # Davor, 2026-09-01: dar de baja a alguien con fecha de baja HOY (o
        # antes) no reprocesa la fila de ClasificacionDiaria de esa fecha
        # -- el motor la calculo ANTES de la baja (mientras todavia estaba
        # Activo en el Maestro) y no vuelve a tocar a alguien que ya no
        # esta Activo. Sin este filtro, la persona seguia apareciendo en
        # Reporte diario/Marcar asistencia como "Falta" el mismo dia que se
        # le dio de baja. "No habria porque tenerla ese dia" (Davor) -- para
        # agregar su reemplazo esta "Agregar reemplazo" (tiene su propia
        # lista de vacantes pendientes, independiente de este reporte).
        if p is not None and p.fecha_baja is not None and p.fecha_baja <= fecha:
            continue
        ayer_c = filas_ayer.get(c.dni)
        fuente_salida = c if salida_mismo_dia else ayer_c
        supervisor_dni_efectivo = overrides_sup.get(c.dni, p.supervisor_dni) if p else None
        supervisor_nombre = nombre_supervisor_de.get(supervisor_dni_efectivo) if supervisor_dni_efectivo else None
        salida_anticipada = fuente_salida.salida_anticipada_min if fuente_salida else None
        filas.append({
            "dni": c.dni,
            "mercaderista": nombre_de.get(c.dni, c.dni),
            "supervisor": supervisor_nombre,
            "region": p.region if p else None,
            "ciudad": p.ciudad if p else None,
            "canal": p.canal if p else None,
            "rol": p.rol if p else None,
            "mercado": p.zona if p else None,
            # OJO: canal_hoy se usa como CHEQUEO DE VERDAD/FALSO en otros
            # lados (_ya_marcaron(), _marcaciones_reales_hoy() -- "¿hay
            # visita real hoy?"), no solo para mostrar -- por eso el
            # reemplazo por el canal asignado cuando cae en "Otro" va en un
            # campo aparte (canal_hoy_mostrar), sin tocar este.
            "canal_hoy": c.canales_marcados or "",
            "canal_hoy_mostrar": _canal_para_mostrar(c.canales_marcados, p.canal if p else None),
            "entrada_prog": c.entrada_esperada,
            "entrada_real": c.entrada_real,
            "entrada_corregida": c.fuente_dato == "Corregido manualmente (Tabla 3)",
            "estado": c.estado,
            # _homologar_motivo() saca el prefijo redundante "Falta" (el
            # badge de Estado ya dice "Falta"), tolerando variantes de
            # formato -- ver docstring de la función.
            "comentario_entrada": _homologar_motivo(override_entrada.get(c.dni, c.comentario_supervisor)) or "",
            "salida_prog": fuente_salida.salida_esperada if fuente_salida else None,
            "salida_real": fuente_salida.salida_real if fuente_salida else None,
            "salida_corregida": bool(fuente_salida and fuente_salida.fuente_dato == "Corregido manualmente (Tabla 3)"),
            "salida_temprana": bool(salida_anticipada and salida_anticipada > SALIDA_ANTICIPADA_MIN),
            "canal_ayer": _canal_para_mostrar(ayer_c.canales_marcados if ayer_c else "", p.canal if p else None),
            "comentario_salida": _homologar_motivo(override_salida.get(c.dni, fuente_salida.comentario_supervisor if fuente_salida else None)) or "",
            "entrada_pendiente": (
                MARCADOR_PENDIENTE in (c.comentario_supervisor or "")
                or (
                    MARCADOR_SIN_APP in (c.estado or "")
                    and _estado_base(c.estado) in _ESTADOS_CON_HORA_PENDIENTE
                )
            ),
            "salida_pendiente": (
                MARCADOR_PENDIENTE in ((fuente_salida.comentario_supervisor or "") if fuente_salida else "")
                or (
                    MARCADOR_SIN_APP in ((fuente_salida.estado or "") if fuente_salida else "")
                    and _estado_base(fuente_salida.estado if fuente_salida else None) in _ESTADOS_CON_HORA_PENDIENTE
                )
            ),
            "marcado_web": c.dni in dnis_marcados_hoy,
        })
        if c.procesado_en and (ultima_sync is None or c.procesado_en > ultima_sync):
            ultima_sync = c.procesado_en

    resumen = {
        "asistio": sum(1 for f in filas if _estado_base(f["estado"]) == "ASISTIÓ A TIEMPO"),
        "tardanza": sum(1 for f in filas if _estado_base(f["estado"]) == "TARDANZA"),
        "falta": sum(1 for f in filas if _estado_base(f["estado"]) == "FALTA"),
        "total": len(filas),
    }
    # procesado_en se guarda con func.now() de Postgres -- en UTC, no hora
    # Peru. Sin esto, "Datos hasta las X" mostraba la hora UTC directo (5
    # horas atrás de la hora real de Peru), pareciendo desactualizado.
    ultima_sync_peru = (
        ultima_sync.replace(tzinfo=ZoneInfo("UTC")).astimezone(PERU_TZ) if ultima_sync else None
    )
    frescura = {"ultima_sync": ultima_sync_peru, "fecha": fecha, "ayer": ayer}
    return resumen, filas, frescura


def historico_persona(dni_filtro, desde, hasta, usuario_actual=None):
    """Como _cargar_reporte(..., salida_mismo_dia=True) pero para UNA sola
    persona a lo largo de un rango de fechas, en una sola tanda de
    consultas (Davor, 2026-09-01, "Histórico diario" con rango + filtro de
    mercaderista). Llamar a _cargar_reporte() una vez POR DÍA tardaba ~2.5s
    por día (7+ consultas cada vez, incluidas cosas que no cambian entre
    llamadas como feriados/supervisor) -- para un rango de varias semanas
    eso pasaba largamente el timeout del worker en Render. Acá se trae
    TODO el rango de una, y se arma cada fila en memoria.

    No incluye supervisor/región/etc -- son constantes para una sola
    persona, no hace falta resolverlos por fila (a diferencia de
    _cargar_reporte(), que muestra varias personas a la vez)."""
    session = get_session()
    try:
        query = session.query(Persona).filter(Persona.dni == dni_filtro)
        cond_scope = condicion_scope(Persona, usuario_actual) if usuario_actual else None
        if cond_scope is not None:
            query = query.filter(cond_scope)
        persona = query.first()
        if persona is None:
            return [], None

        filas_rango = (
            session.query(ClasificacionDiaria)
            .filter(ClasificacionDiaria.dni == dni_filtro, ClasificacionDiaria.fecha >= desde, ClasificacionDiaria.fecha <= hasta)
            .order_by(ClasificacionDiaria.fecha.desc())
            .all()
        )
        if not filas_rango:
            return [], None

        # Mismo criterio que correcciones_recientes en _cargar_reporte(),
        # pero para TODO el rango de una -- el motivo/comentario recién
        # guardado desde la web aparece de una, sin esperar a que el motor
        # vuelva a correr (ver ese comentario más arriba para el porqué).
        correcciones = (
            session.query(CorreccionWeb)
            .filter(CorreccionWeb.dni == dni_filtro, CorreccionWeb.fecha >= desde, CorreccionWeb.fecha <= hasta)
            .order_by(CorreccionWeb.fecha_registro.desc())
            .all()
        )
        override_entrada_por_fecha, override_salida_por_fecha = {}, {}
        for corr in correcciones:
            if corr.comentario_entrada and corr.fecha not in override_entrada_por_fecha:
                override_entrada_por_fecha[corr.fecha] = corr.comentario_entrada
            if corr.comentario_salida and corr.fecha not in override_salida_por_fecha:
                override_salida_por_fecha[corr.fecha] = corr.comentario_salida

        filas = []
        ultima_sync = None
        for c in filas_rango:
            filas.append({
                "dni": c.dni, "mercaderista": persona.nombre_completo,
                "fecha_str": c.fecha.isoformat(),
                "canal_hoy_mostrar": _canal_para_mostrar(c.canales_marcados, persona.canal),
                "entrada_prog": c.entrada_esperada,
                "entrada_real": c.entrada_real,
                "estado": c.estado,
                "comentario_entrada": _homologar_motivo(override_entrada_por_fecha.get(c.fecha, c.comentario_supervisor)) or "",
                "salida_prog": c.salida_esperada,
                "salida_real": c.salida_real,
                "salida_temprana": bool(c.salida_anticipada_min and c.salida_anticipada_min > SALIDA_ANTICIPADA_MIN),
                "comentario_salida": _homologar_motivo(override_salida_por_fecha.get(c.fecha, c.comentario_supervisor)) or "",
            })
            if c.procesado_en and (ultima_sync is None or c.procesado_en > ultima_sync):
                ultima_sync = c.procesado_en
        ultima_sync_peru = ultima_sync.replace(tzinfo=ZoneInfo("UTC")).astimezone(PERU_TZ) if ultima_sync else None
        return filas, ultima_sync_peru
    finally:
        session.close()


def _motivos_falta():
    session = get_session()
    try:
        return [
            m for (m,) in session.query(CatalogoMotivo.motivo)
            .filter(CatalogoMotivo.categoria != "Descanso")
            .order_by(CatalogoMotivo.motivo).all()
        ]
    finally:
        session.close()


def _motivos_descanso():
    """Motivos del boton "Dia de descanso" (Davor, 2026-08-31) -- catalogo
    separado del de Falta via CatalogoMotivo.categoria == "Descanso": el
    dia SI se le considera a la persona, no se descuenta ni cuenta como
    falta en el Indicador de Asistencia (a diferencia de Falta)."""
    session = get_session()
    try:
        return [
            m for (m,) in session.query(CatalogoMotivo.motivo)
            .filter(CatalogoMotivo.categoria == "Descanso")
            .order_by(CatalogoMotivo.motivo).all()
        ]
    finally:
        session.close()


def _motivos_falta_con_id():
    """Para el panel de administrar motivos (analista/admin) -- necesita el
    id de cada fila para poder borrarla, a diferencia de _motivos_falta()/
    _motivos_descanso() (usadas en los <select>, donde solo importa el
    texto). Separado en 2 listas para poder mostrarlas agrupadas."""
    session = get_session()
    try:
        todos = session.query(CatalogoMotivo).order_by(CatalogoMotivo.motivo).all()
        falta = [(m.id, m.motivo) for m in todos if m.categoria != "Descanso"]
        descanso = [(m.id, m.motivo) for m in todos if m.categoria == "Descanso"]
        return falta, descanso
    finally:
        session.close()


def _pendientes_de_marcar(filas):
    """Personas con Falta/Vacante que todavía no tienen ningún comentario --
    ni de la app móvil de supervisores ni de acá -- igual al panel
    "Pendientes" de la Power App (galPendientes, ver GUIA_POWER_APPS_SUPERVISOR.md)."""
    return [
        f for f in filas
        if _estado_base(f["estado"]) in ("FALTA", "VACANTE") and not f["comentario_entrada"] and not f["marcado_web"]
    ]


def _ya_marcaron(filas):
    """Personas del equipo que tienen al menos una visita REAL del
    aplicativo hoy. Usa `canal_hoy` (canales_marcados), no `entrada_real`
    -- una corrección manual de hora en Tabla 3 pisa `entrada_real` igual
    aunque no haya ninguna visita real ese día (ver docstring de
    _corregidos_a_mano), mientras que `canales_marcados` solo se llena a
    partir de visitas reales y el bloque de corrección de
    motor_clasificacion.py nunca lo toca -- es la señal confiable de "esto
    es una marcación real, no una que le puso el supervisor" (Davor,
    2026-08-24: reportó ver a alguien en "Ya marcaron" que en realidad
    todavía no había marcado, solo tenía la hora que él le puso a mano).
    Excluye a quien tenga una corrección manual activa -- esos viven aparte
    en "Confirmados a mano" TODO el día, ver _corregidos_a_mano (Davor,
    2026-08-25)."""
    return sorted(
        (f for f in filas if f["canal_hoy"] and not (f["entrada_corregida"] or f["entrada_pendiente"])),
        key=lambda f: f["entrada_real"] or "",
    )


def _faltas_vacaciones_vacantes(filas):
    """Faltas/Vacaciones/Vacantes registradas hoy, resueltas o no -- para
    revisar al cierre del día qué quedó cargado en total, no solo lo
    todavía pendiente (a diferencia de _pendientes_de_marcar) (Davor,
    2026-08-24)."""
    return sorted(
        (f for f in filas if _estado_base(f["estado"]) in ("FALTA", "VACANTE", "VACACIONES")),
        key=lambda f: (_estado_base(f["estado"]), f["mercaderista"]),
    )


def _corregidos_a_mano(filas):
    """Gente con una corrección manual de hora activa hoy -- botón rápido
    "Asistió"/"Apoyo zona" (`entrada_pendiente`, MARCADOR_SIN_APP en el
    Estado) O una hora tipeada directamente (`entrada_corregida`, ej. "Ya
    marcaron: Editar"/"Confirmados a mano: Editar", para poder mandar el
    reporte temprano con captura). Antes esta sección solo mostraba a
    quienes el aplicativo TODAVÍA no les había registrado ninguna visita
    real (`canal_hoy` vacío) -- apenas llegaba una visita real desaparecían
    de acá y se mezclaban en "Ya marcaron" mostrando la hora tipeada como si
    fuera la real, sin ninguna señal de que había sido una corrección manual
    (Davor, 2026-08-25: "todos los que ponga hora manual, deben aparecer
    aparte, todo el día, y solo confirmarme cuando ya marcaron con su hora
    de marcación real, solo como detalle"). Ahora se quedan acá TODO el día
    -- _ya_marcaron() los excluye a propósito -- y marcar_vista() les agrega
    `hora_real_detalle`/`canal_real_detalle` (ver _marcaciones_reales_hoy)
    cuando el aplicativo ya registró de verdad, solo como dato informativo,
    sin pisar la hora corregida."""
    return [f for f in filas if f["entrada_corregida"] or f["entrada_pendiente"]]


def _marcaciones_reales_hoy(dnis, fecha):
    """Hora y canal de la primera visita REAL (Postgres `visitas`, ver
    dimension_models.Visita) para cada DNI en `fecha` -- el "detalle"
    informativo que pide _corregidos_a_mano() una vez que el aplicativo ya
    registró de verdad a alguien que se había confirmado a mano. Nunca pisa
    la hora corregida, solo se muestra al lado."""
    if not dnis:
        return {}
    session = get_session()
    try:
        filas_v = (
            session.query(Visita.dni, Visita.hora_inicio, Visita.tipo_negocio)
            .filter(Visita.dni.in_(dnis), Visita.fecha_inicio == fecha, Visita.hora_inicio.isnot(None))
            .all()
        )
    finally:
        session.close()
    reales = {}
    for dni, hora, canal in filas_v:
        if dni not in reales or hora < reales[dni][0]:
            reales[dni] = (hora, canal)
    return reales


def _reemplazos_hoy(fecha, usuario_actual):
    """Personas dadas de alta HOY como reemplazo de una vacante (ver
    reemplazos.py::procesar_reemplazo(), que estampa fecha_registro=hoy y
    reemplaza_a_dni) -- para que el supervisor vea en la misma pantalla que
    se procesó un reemplazo, no solo que una vacante desapareció."""
    session = get_session()
    try:
        Reemplazada = aliased(Persona)
        query = (
            session.query(Persona, Reemplazada.nombre_completo)
            .outerjoin(Reemplazada, Reemplazada.dni == Persona.reemplaza_a_dni)
            .filter(Persona.fecha_registro == fecha, Persona.reemplaza_a_dni.isnot(None))
        )
        cond_scope = condicion_scope(Persona, usuario_actual) if usuario_actual else None
        if cond_scope is not None:
            query = query.filter(cond_scope)
        filas_reemplazo = query.all()
    finally:
        session.close()
    return [
        {
            "nombre_nuevo": p.nombre_completo, "dni_nuevo": p.dni,
            "nombre_reemplazado": nombre_reemplazada or p.reemplaza_a_dni,
            "fecha_ingreso": p.fecha_ingreso,
        }
        for p, nombre_reemplazada in filas_reemplazo
    ]


def _fecha_mas_reciente_con_datos():
    session = get_session()
    try:
        from sqlalchemy import func
        (fecha,) = session.query(func.max(ClasificacionDiaria.fecha)).one()
        return fecha
    finally:
        session.close()


def _filtros_marcar():
    """Filtros de Supervisor/Región/Rol/Ciudad en Marcar asistencia --
    Davor, 2026-08-25: "poner filtro supervisor, region, rol, solo para el
    admin y analista, para supervisor no deberia aparecer filtros" +
    "agrega tambien filtros de ciudad". Acá también alcanza a analista (a
    diferencia de la versión anterior de reportes.py::_filtros_admin(), que
    era solo-admin y se unificó a este mismo criterio) -- un analista ya
    está acotado a su propio cliente por condicion_scope(), así que los
    desplegables igual solo muestran lo suyo."""
    if current_user.rol == "supervisor":
        return {"rol": "", "region": "", "supervisor": "", "ciudad": "", "canal": ""}, [], [], [], [], []
    es_admin = current_user.rol == "admin"
    filtro_args = {
        "rol": request.args.get("rol") or "",
        "region": request.args.get("region") or "",
        "supervisor": request.args.get("supervisor") or "",
        "ciudad": request.args.get("ciudad") or "",
        "canal": (request.args.get("canal") or "") if es_admin else "",
    }
    session = get_session()
    try:
        cond_scope = condicion_scope(Persona, current_user)

        def _valores(columna):
            q = session.query(columna).filter(columna.isnot(None))
            if cond_scope is not None:
                q = q.filter(cond_scope)
            return sorted({v for (v,) in q.distinct().all() if v})

        roles_disponibles = _valores(Persona.rol)
        regiones_disponibles = _valores(Persona.region)
        ciudades_disponibles = _valores(Persona.ciudad)

        SupervisorPersona = aliased(Persona)
        q_sup = (
            session.query(Persona.supervisor_dni, SupervisorPersona.nombre_completo)
            .join(SupervisorPersona, SupervisorPersona.dni == Persona.supervisor_dni)
            .filter(Persona.supervisor_dni.isnot(None))
        )
        if cond_scope is not None:
            q_sup = q_sup.filter(cond_scope)
        supervisores_disponibles = sorted(set(q_sup.distinct().all()), key=lambda t: (t[1] or "").title())
    finally:
        session.close()
    canales_disponibles = CANALES_FILTRABLES if es_admin else []
    return filtro_args, roles_disponibles, regiones_disponibles, supervisores_disponibles, ciudades_disponibles, canales_disponibles


def _vista_reporte(fecha_str, vista):
    """vista: 'reporte' (columnas + edicion completa) o 'cliente' (columnas
    reducidas, sin Hora entrada/salida corregida)."""
    try:
        fecha = dt.date.fromisoformat(fecha_str)
    except ValueError:
        fecha = dt.date.today()
        fecha_str = fecha.isoformat()

    # Canal (Davor, 2026-08-29) -- SOLO admin: "debo tener un filtro para
    # ver Tradicional, Farmacia y AU" -- un analista de canal ya está
    # acotado por condicion_scope(), no lo necesita.
    es_admin = current_user.rol == "admin"
    canal_filtro = (request.args.get("canal") or "") if es_admin else ""
    canales_disponibles = CANALES_FILTRABLES if es_admin else []

    resumen, filas, frescura = _cargar_reporte(fecha, usuario_actual=current_user, canal_filtro=canal_filtro)
    # Si no hay datos para la fecha pedida (ej. hoy, si el pipeline todavia
    # no corrio), sugerir la fecha mas reciente que si tiene -- para poder
    # seguir corrigiendo el pasado sin quedar en un callejon sin salida.
    fecha_reciente = None if filas else _fecha_mas_reciente_con_datos()
    return render_template(
        "asistencia.html", usuario=current_user, activo=vista, vista=vista,
        fecha_reciente=fecha_reciente,
        fecha_str=fecha_str, resumen=resumen, filas=filas, frescura=frescura,
        canal_filtro=canal_filtro, canales_disponibles=canales_disponibles,
    )


@bp.get("/marcar")
@requiere_pagina("asistencia")
def marcar_vista():
    """Pestaña propia (version web de galPendientes de la Power App) --
    antes vivia dentro de "Reporte diario", ahora tiene su propio lugar
    para no mezclar "ver/corregir el detalle" con "marcar rapido a los
    pendientes de hoy"."""
    fecha_str = request.args.get("fecha") or dt.date.today().isoformat()
    try:
        fecha = dt.date.fromisoformat(fecha_str)
    except ValueError:
        fecha = dt.date.today()
        fecha_str = fecha.isoformat()

    filtro_args, roles_disponibles, regiones_disponibles, supervisores_disponibles, ciudades_disponibles, canales_disponibles = _filtros_marcar()
    resumen, filas, frescura = _cargar_reporte(
        fecha, usuario_actual=current_user,
        rol_filtro=filtro_args["rol"], region_filtro=filtro_args["region"], supervisor_filtro=filtro_args["supervisor"],
        ciudad_filtro=filtro_args["ciudad"], canal_filtro=filtro_args["canal"],
    )
    pendientes = _pendientes_de_marcar(filas) if filas else []
    marcaron = _ya_marcaron(filas) if filas else []
    faltas_vacaciones_vacantes = _faltas_vacaciones_vacantes(filas) if filas else []
    corregidos_a_mano = _corregidos_a_mano(filas) if filas else []
    reales_hoy = _marcaciones_reales_hoy([f["dni"] for f in corregidos_a_mano if f["canal_hoy"]], fecha)
    for f in corregidos_a_mano:
        detalle = reales_hoy.get(f["dni"])
        f["hora_real_detalle"] = detalle[0] if detalle else None
        f["canal_real_detalle"] = detalle[1] if detalle else None
    reemplazos = _reemplazos_hoy(fecha, current_user)
    fecha_reciente = None if filas else _fecha_mas_reciente_con_datos()
    hay_pendientes_de_algo = pendientes or marcaron or corregidos_a_mano or faltas_vacaciones_vacantes
    motivos_catalogo_falta, motivos_catalogo_descanso = (
        _motivos_falta_con_id() if current_user.rol != "supervisor" else (None, None)
    )
    return render_template(
        "asistencia_marcar.html", usuario=current_user, activo="marcar",
        fecha_str=fecha_str, resumen=resumen, frescura=frescura, fecha_reciente=fecha_reciente,
        pendientes=pendientes, marcaron=marcaron,
        motivos=_motivos_falta() if hay_pendientes_de_algo else None,
        motivos_descanso=_motivos_descanso() if hay_pendientes_de_algo else None,
        motivos_catalogo=motivos_catalogo_falta, motivos_catalogo_descanso=motivos_catalogo_descanso,
        faltas_vacaciones_vacantes=faltas_vacaciones_vacantes, corregidos_a_mano=corregidos_a_mano,
        reemplazos=reemplazos, marcado=request.args.get("marcado"),
        filtro_args=filtro_args, roles_disponibles=roles_disponibles,
        regiones_disponibles=regiones_disponibles, supervisores_disponibles=supervisores_disponibles,
        ciudades_disponibles=ciudades_disponibles, canales_disponibles=canales_disponibles,
    )


@bp.get("")
@requiere_pagina("asistencia")
def reporte():
    # El supervisor solo debe poder marcar asistencia de su equipo -- ver/
    # editar el detalle completo (horas corregidas, borrar comentarios) es
    # una herramienta de analista.
    if current_user.rol == "supervisor":
        return redirect(url_for("asistencia.marcar_vista"))
    fecha_str = request.args.get("fecha") or dt.date.today().isoformat()
    return _vista_reporte(fecha_str, "reporte")


@bp.get("/cliente")
@requiere_pagina("asistencia")
def cliente():
    if current_user.rol == "supervisor":
        return redirect(url_for("asistencia.marcar_vista"))
    fecha_str = request.args.get("fecha") or dt.date.today().isoformat()
    return _vista_reporte(fecha_str, "cliente")


def _agregar_fila_tabla3(ws, fila_libre, dni, fecha, comentario, hora_ent=None, hora_sal=None, estado_reportado=None):
    """Mismo formato/columnas que aplicar_correcciones.py::_agregar_fila_t3()
    -- no cambiar sin revisar ese archivo primero, el motor de clasificacion
    (local y nube) depende de este layout exacto.

    `estado_reportado` (columna 3, "Estado reportado") es la MISMA columna
    que llena la app Power Apps de los supervisores para "Asistió"/"Apoyo
    zona"/"Vacante" -- ver GUIA_POWER_APPS_SUPERVISOR.md sección 3/4. Hasta
    ahora esta función nunca la llenaba (solo Comentario), asi que esas 3
    acciones no tenian forma de reportarse desde la web. "Falta" sigue
    usando el Comentario con el prefijo "Falta - {motivo}" (columna 3 vacía),
    igual que ya hace el resto del flujo -- motor_clasificacion.py lee
    "Estado reportado" solo como respaldo cuando NO hay un comentario que
    empiece con "FALTA"/tenga "VACANTE"/"VACACIONES".
    """
    ws.cell(row=fila_libre, column=1, value=str(dni))
    ws.cell(row=fila_libre, column=2, value=str(fecha))
    if estado_reportado:
        ws.cell(row=fila_libre, column=3, value=estado_reportado)
    # texto_seguro_excel(): comentario es texto libre tipeado por el usuario
    # (o un motivo agregado por otro usuario) -- sin esto, un valor que
    # empiece con =/+/-/@ se interpretaria como formula al abrir la Tabla 3
    # en Excel (CSV/Excel Injection).
    ws.cell(row=fila_libre, column=4, value=texto_seguro_excel(comentario))
    ws.cell(row=fila_libre, column=5, value="Web mac_cloud")
    ws.cell(row=fila_libre, column=6, value=dt.datetime.now().strftime("%H:%M"))
    if hora_ent:
        ws.cell(row=fila_libre, column=7, value=str(hora_ent))
    if hora_sal:
        ws.cell(row=fila_libre, column=8, value=str(hora_sal))


@bp.post("/guardar")
@requiere_pagina("asistencia")
def guardar():
    # Mismo criterio que reporte()/cliente(): editar horas/comentarios del
    # detalle es cosa de analista, no del supervisor (que solo marca).
    if current_user.rol == "supervisor":
        flash("Esta sección es solo para analistas.", "error")
        return redirect(url_for("asistencia.marcar_vista"))
    fecha_str = request.form["fecha"]
    vista = request.form.get("vista") or "reporte"
    fecha = dt.date.fromisoformat(fecha_str)
    feriados = _cargar_feriados()
    ayer = _dia_habil_anterior(fecha, feriados)

    dnis = request.form.getlist("dni")
    ediciones = []
    for dni in dnis:
        comentario_entrada = request.form.get(f"comentario_entrada_{dni}", "").strip()
        comentario_salida = request.form.get(f"comentario_salida_{dni}", "").strip()
        entrada_corr = request.form.get(f"entrada_corr_{dni}", "").strip() if vista in ("reporte", "marcar") else ""
        salida_corr = request.form.get(f"salida_corr_{dni}", "").strip() if vista == "reporte" else ""
        # "Confirmados a mano, pendiente marcación" (Marcar asistencia) --
        # gente sin equipo/app, el supervisor pone la hora de salida de HOY
        # directo a fin de día en vez de esperar a corregirla mañana desde
        # "Reporte diario" (que siempre corrige la salida de AYER, ver
        # salida_corr arriba). Campo separado a propósito, para no pisar el
        # significado de salida_corr_{dni} en el resto de las vistas.
        salida_corr_hoy = request.form.get(f"salida_corr_hoy_{dni}", "").strip() if vista == "marcar" else ""

        # "Ya marcaron" (Marcar asistencia) arma el comentario desde un
        # Motivo + Detalle separados en vez de un solo cuadro de texto libre
        # (mismo criterio que ya usa marcar() para Pendientes) -- si vienen,
        # pisan lo que haya en comentario_entrada_{dni}.
        motivo_falta = request.form.get(f"motivo_falta_{dni}", "").strip()
        if motivo_falta:
            detalle_falta = request.form.get(f"detalle_falta_{dni}", "").strip()
            comentario_entrada = f"Falta - {motivo_falta}"
            if detalle_falta:
                comentario_entrada += f" ({detalle_falta})"

        # "Borrar" gana sobre lo que haya quedado tipeado en el cuadro de
        # texto -- el usuario tildó el check porque quiere deshacer el
        # comentario, no reemplazarlo por otra cosa a medio escribir.
        if request.form.get(f"borrar_entrada_{dni}"):
            comentario_entrada = MARCADOR_BORRADO
        if request.form.get(f"borrar_salida_{dni}"):
            comentario_salida = MARCADOR_BORRADO

        # Una hora de entrada corregida y un comentario "Falta" son
        # contradictorios -- motor_clasificacion.py prioriza la hora
        # corregida (recalcula ASISTIÓ/TARDANZA) e ignora que el comentario
        # diga "Falta" cuando las dos llegan juntas en la misma fila de
        # Tabla 3 (Davor, 2026-08-25: "mercaderista que marca y se va...
        # deberia ser Falta"). Si el analista cargó las dos a la vez, gana
        # la conversión a Falta -- es la acción explícita, la hora corregida
        # ahí sería un resto de haber tocado el campo por error.
        if comentario_entrada.upper().startswith("FALTA") and entrada_corr:
            entrada_corr = ""

        if comentario_entrada or comentario_salida or entrada_corr or salida_corr or salida_corr_hoy:
            ediciones.append({
                "dni": dni, "comentario_entrada": comentario_entrada, "comentario_salida": comentario_salida,
                "entrada_corr": entrada_corr, "salida_corr": salida_corr, "salida_corr_hoy": salida_corr_hoy,
            })

    endpoint_vista = {"cliente": "asistencia.cliente", "marcar": "asistencia.marcar_vista"}.get(vista, "asistencia.reporte")

    if not ediciones:
        return redirect(url_for(endpoint_vista, fecha=fecha_str))

    ok_lock, motivo_lock = adquirir_lock("tabla3_web", f"web:{current_user.email}", max_minutos=2)
    if not ok_lock:
        return render_template(
            "asistencia_resultado.html", usuario=current_user, activo=vista,
            titulo="No se pudo guardar", ok=False,
            detalle=f"{motivo_lock} -- probá de nuevo en un minuto.",
            volver=url_for(endpoint_vista, fecha=fecha_str),
        )
    try:
        wb = openpyxl.load_workbook(io.BytesIO(descargar(TABLA3_RUTA_GRAPH)))
        ws = wb["Registro diario supervisor"]
        fila_libre = ws.max_row + 1

        # Se arma todo en memoria y se sube a SharePoint (T3, la fuente
        # real que lee el motor) ANTES de comitear a Postgres -- igual que
        # marcar() y el cierre de Falta por reemplazo. Al revés (comitear
        # primero) era el único lugar que lo hacía así: si subir_in_place
        # fallaba después del commit, la fila CorreccionWeb ya existía y
        # sacaba a la persona de "Pendientes de marcar" aunque la
        # corrección nunca hubiera llegado a T3 -- se perdía en silencio.
        correcciones_web = []
        for e in ediciones:
            if e["comentario_entrada"] or e["entrada_corr"]:
                _agregar_fila_tabla3(ws, fila_libre, e["dni"], fecha, e["comentario_entrada"], hora_ent=e["entrada_corr"])
                fila_libre += 1
                correcciones_web.append(CorreccionWeb(
                    dni=e["dni"], fecha=fecha,
                    comentario_entrada=e["comentario_entrada"] or None,
                    hora_entrada_corregida=e["entrada_corr"] or None,
                    registrado_por=current_user.email,
                ))
            if e["comentario_salida"] or e["salida_corr"]:
                _agregar_fila_tabla3(ws, fila_libre, e["dni"], ayer, e["comentario_salida"], hora_sal=e["salida_corr"])
                fila_libre += 1
                correcciones_web.append(CorreccionWeb(
                    dni=e["dni"], fecha=ayer,
                    comentario_salida=e["comentario_salida"] or None,
                    hora_salida_corregida=e["salida_corr"] or None,
                    registrado_por=current_user.email,
                ))
            if e.get("salida_corr_hoy"):
                _agregar_fila_tabla3(ws, fila_libre, e["dni"], fecha, None, hora_sal=e["salida_corr_hoy"])
                fila_libre += 1
                correcciones_web.append(CorreccionWeb(
                    dni=e["dni"], fecha=fecha,
                    hora_salida_corregida=e["salida_corr_hoy"],
                    registrado_por=current_user.email,
                ))

        subir_in_place(TABLA3_RUTA_GRAPH, wb)

        session = get_session()
        try:
            for c in correcciones_web:
                session.add(c)
            session.commit()
        finally:
            session.close()
    finally:
        liberar_lock("tabla3_web")

    plural = "personas" if len(ediciones) != 1 else "persona"
    return render_template(
        "asistencia_resultado.html", usuario=current_user, activo=vista,
        titulo="Correcciones guardadas", ok=True,
        detalle=(f"Se guardaron los cambios de {len(ediciones)} {plural}. Se van a reflejar en el reporte en "
                 "unos minutos -- si querés verlos ya, usá \"Actualizar ahora\"."),
        volver=url_for(endpoint_vista, fecha=fecha_str),
    )


ACCIONES_MARCAR = ("Asistió", "Descanso", "Vacante", "Falta")

# Davor, 2026-08-31: "Apoyo zona" no servía -- cualquier estado reportado
# por el supervisor sin marcación real de la app (incluido "Apoyo zona")
# caía en un balde genérico ("... según supervisor, sin marcación app")
# que motor_compensacion_indicador.py SÍ contaba como falta en el
# Indicador de Asistencia, aunque la persona sí hubiera trabajado. Se
# reemplaza por "Día de descanso" con motivo obligatorio (mismo patrón que
# Falta), que el motor clasifica aparte ("DESCANSO (comentario supervisor)",
# ver motor_clasificacion.py) y que el indicador NO cuenta como negativo --
# "se le considera el día", no se descuenta.
ACCIONES_CON_MOTIVO = ("Falta", "Descanso")


@bp.post("/marcar")
@requiere_pagina("asistencia")
def marcar():
    """Version web de los botones Asistió/Día de descanso/Vacante/Falta de
    la Power App (galPendientes -- ver GUIA_POWER_APPS_SUPERVISOR.md
    secciones 3/4). Escribe a la MISMA Tabla 3 que la app móvil, con el
    mismo criterio que usa motor_clasificacion.py para interpretarlo --
    "Falta"/"Descanso" via Comentario="{Accion} - {motivo}" (mismo patron
    que ya usan las correcciones web y la app), las otras 2 via la columna
    "Estado reportado"."""
    dni = request.form["dni"].strip()
    fecha_str = request.form["fecha"]
    fecha = dt.date.fromisoformat(fecha_str)
    accion = request.form.get("accion", "").strip()
    motivo = request.form.get("motivo", "").strip() or request.form.get("motivo_descanso", "").strip()
    comentario_extra = request.form.get("comentario", "").strip()

    if accion not in ACCIONES_MARCAR:
        return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str))
    if accion in ACCIONES_CON_MOTIVO and not motivo:
        return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str, marcado="falta_sin_motivo"))

    # El detalle adicional va entre paréntesis al final -- _homologar_motivo()
    # / _motivo_limpio() ya recortan cualquier paréntesis final al agrupar
    # "Faltas por motivo" (mismo mecanismo que ya usa MARCADOR_BORRADO), así
    # que el motivo sigue agrupando bien aunque cada persona escriba un
    # detalle distinto acá.
    comentario_con_motivo = f"{accion} - {motivo}"
    if comentario_extra:
        comentario_con_motivo += f" ({comentario_extra})"

    ok_lock, motivo_lock = adquirir_lock("tabla3_web", f"web:{current_user.email}", max_minutos=2)
    if not ok_lock:
        return render_template(
            "asistencia_resultado.html", usuario=current_user, activo="marcar",
            titulo="No se pudo guardar", ok=False,
            detalle=f"{motivo_lock} -- probá de nuevo en un minuto.",
            volver=url_for("asistencia.marcar_vista", fecha=fecha_str),
        )
    try:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(descargar(TABLA3_RUTA_GRAPH)))
            ws = wb["Registro diario supervisor"]
            fila_libre = ws.max_row + 1
            if accion in ACCIONES_CON_MOTIVO:
                _agregar_fila_tabla3(ws, fila_libre, dni, fecha, comentario_con_motivo)
            else:
                _agregar_fila_tabla3(ws, fila_libre, dni, fecha, None, estado_reportado=accion)
            subir_in_place(TABLA3_RUTA_GRAPH, wb)
        except requests.exceptions.RequestException as e:
            # Sin esto, si SharePoint/Graph no responde (timeout, red caida,
            # throttle) la excepcion volaba sin capturar y el navegador
            # quedaba con el boton "cargando" pegado -- ahora corta y avisa
            # (Davor, 2026-08-31: "sale cargando nomas y no se guarda").
            return render_template(
                "asistencia_resultado.html", usuario=current_user, activo="marcar",
                titulo="No se pudo guardar", ok=False,
                detalle=f"Fallo la conexión con SharePoint/Graph ({e}) -- probá de nuevo en un minuto.",
                volver=url_for("asistencia.marcar_vista", fecha=fecha_str),
            )

        # Ademas de Tabla 3 (lo que lee el motor), se guarda en correcciones_web
        # -- asi "Pendientes de marcar" puede sacar a esta persona de la
        # lista de una, sin esperar a que el motor vuelva a correr.
        session = get_session()
        try:
            session.add(CorreccionWeb(
                dni=dni, fecha=fecha,
                comentario_entrada=(comentario_con_motivo if accion in ACCIONES_CON_MOTIVO else accion),
                registrado_por=current_user.email,
            ))
            session.commit()
        finally:
            session.close()
    finally:
        liberar_lock("tabla3_web")

    return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str, marcado="ok"))


@bp.post("/motivos/agregar")
@requiere_pagina("asistencia")
def motivos_agregar():
    fecha_str = request.form.get("fecha") or dt.date.today().isoformat()
    motivo = request.form.get("motivo", "").strip()
    if not motivo:
        return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str))
    # "Descanso" (Davor, 2026-08-31) para que el nuevo motivo caiga en el
    # desplegable de "Día de descanso" en vez del de "Falta" por defecto.
    categoria = "Descanso" if request.form.get("tipo") == "descanso" else "Falta"

    session = get_session()
    try:
        ya_existe = session.query(CatalogoMotivo).filter(CatalogoMotivo.motivo.ilike(motivo)).first()
        if not ya_existe:
            session.add(CatalogoMotivo(motivo=motivo, categoria=categoria))
            session.commit()
    finally:
        session.close()

    return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str, marcado="motivo_agregado"))


@bp.post("/motivos/eliminar")
@_analista_requerido
def motivos_eliminar():
    # Mas sensible que agregar (afecta el desplegable de todos) -- analista
    # o admin, no supervisor (Davor, 2026-08-31: "Vacante" se repetia con
    # el boton de accion rapida y quiso poder sacarlo de la lista).
    fecha_str = request.form.get("fecha") or dt.date.today().isoformat()
    try:
        motivo_id = int(request.form.get("motivo_id", ""))
    except ValueError:
        return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str))

    session = get_session()
    try:
        motivo = session.query(CatalogoMotivo).filter(CatalogoMotivo.id == motivo_id).first()
        if motivo:
            session.delete(motivo)
            session.commit()
    finally:
        session.close()

    return redirect(url_for("asistencia.marcar_vista", fecha=fecha_str, marcado="motivo_eliminado"))


@bp.post("/actualizar")
@requiere_pagina("asistencia")
def actualizar():
    # Dispara el pipeline COMPLETO (no solo reporte_9am.yml) a proposito:
    # reporte_9am.yml unicamente regenera el reporte con lo YA calculado en
    # Postgres -- aplicar_correcciones.py (su unico paso relevante) solo
    # empuja ediciones hechas sobre el snapshot de auditoria Reporte_Asistencia_*.xlsx,
    # y si no encuentra ninguna ahi (que es el caso normal: tanto el guardado
    # de mac_cloud como la app movil de supervisores escriben DIRECTO a
    # Tabla 3) se salta motor_clasificacion.py por completo -- los
    # comentarios/motivos nuevos en Tabla 3 quedan sin aplicar. pipeline_completo.yml
    # si corre motor_clasificacion.py siempre, asi que es lo unico que
    # garantiza reflejar comentarios nuevos (via web o app movil).
    fecha_str = request.form["fecha"]
    vista = request.form.get("vista") or "reporte"
    ok, mensaje = disparar_workflow("pipeline_completo.yml")
    return render_template(
        "asistencia_resultado.html", usuario=current_user, activo=vista,
        titulo="Actualización disparada" if ok else "No se pudo disparar", ok=ok, detalle=mensaje,
        volver=url_for(f"asistencia.{'cliente' if vista == 'cliente' else 'reporte'}", fecha=fecha_str),
        poll_workflow="pipeline_completo.yml" if ok else None,
    )


@bp.get("/actualizar")
@requiere_pagina("asistencia")
def actualizar_get():
    # A diferencia de /guardar y /marcar (que redirigen tras el POST), esta
    # vista renderiza el resultado directo en la URL del POST -- si el
    # navegador vuelve a pedir esa misma URL más tarde (recuperar pestaña
    # tras cerrar Chrome, refrescar, etc.) lo hace por GET, sin el body del
    # formulario, y Flask tira "Método no permitido" (Davor, 2026-08-24: se
    # quedó viendo esa pantalla en blanco). No dispara el workflow de nuevo
    # (eso sería un GET con efecto secundario real) -- solo manda de vuelta
    # al reporte del día.
    return redirect(url_for("asistencia.reporte"))


@bp.get("/estado_workflow")
@login_required
def estado_workflow():
    nombre = request.args.get("wf", "")
    return jsonify(estado_ultima_corrida(nombre) or {})


_WORKFLOWS_ENCADENABLES = ("exportar_dimensiones.yml", "pipeline_completo.yml")


@bp.post("/disparar_siguiente")
@login_required
def disparar_siguiente():
    # Paso 2 de la cadena "actualizar Maestro -> correr motor de clasificación"
    # (ver poll_siguiente en asistencia_resultado.html) -- el JS de polling
    # llama a esto apenas exportar_dimensiones.yml termina bien, para recién
    # ahí disparar pipeline_completo.yml (que lee el Excel que ese primer
    # paso acaba de regenerar). Whitelist a propósito: esto SÍ dispara una
    # acción (a diferencia de estado_workflow, que solo lee), no aceptar
    # cualquier nombre de archivo.
    nombre = request.form.get("wf", "")
    if nombre not in _WORKFLOWS_ENCADENABLES:
        return jsonify({"ok": False}), 400
    ok, _mensaje = disparar_workflow(nombre)
    return jsonify({"ok": ok})


@bp.get("/reemplazo")
@_analista_requerido
def reemplazo_form():
    hoy = dt.date.today().isoformat()
    try:
        pendientes = r9am.reemplazos_pendientes(r9am.dnis_alguna_vez_en_maestro())
        error_pendientes = None
    except Exception as e:
        pendientes = []
        error_pendientes = str(e)

    dni_prellenado = request.args.get("dni_vacante", "").strip()
    nombre_prellenado = None
    if dni_prellenado:
        session = get_session()
        try:
            p = session.query(Persona).filter(Persona.dni == dni_prellenado).first()
            nombre_prellenado = p.nombre_completo if p else None
        finally:
            session.close()

    return render_template(
        "asistencia_reemplazo.html", usuario=current_user, activo="reemplazo",
        hoy=hoy, pendientes=pendientes, error_pendientes=error_pendientes,
        dni_prellenado=dni_prellenado, nombre_prellenado=nombre_prellenado,
    )


@bp.post("/reemplazo")
@_analista_requerido
def reemplazo_submit():
    from reemplazos import procesar_reemplazo

    dni_vacante = request.form["dni_vacante"].strip()
    dni_nuevo = request.form["dni_nuevo"].strip()
    nombre_nuevo = request.form["nombre_nuevo"].strip()
    fecha_ingreso = dt.date.fromisoformat(request.form["fecha_ingreso"].strip())
    motivo_baja = request.form.get("motivo_baja", "").strip() or None

    try:
        log = procesar_reemplazo(dni_vacante, dni_nuevo, nombre_nuevo, fecha_ingreso, motivo_baja=motivo_baja)

        # Cierra el pendiente de HOY de la persona reemplazada -- sin esto,
        # "Agregar reemplazo" resuelve el futuro (quien entra) pero la
        # persona saliente seguia apareciendo en "Pendientes de marcar"
        # porque su Falta de hoy nunca queda justificada por separado.
        try:
            hoy = dt.date.today()
            ok_lock, _ = adquirir_lock("tabla3_web", f"web:{current_user.email}", max_minutos=2)
            if ok_lock:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(descargar(TABLA3_RUTA_GRAPH)))
                    ws = wb["Registro diario supervisor"]
                    _agregar_fila_tabla3(ws, ws.max_row + 1, dni_vacante, hoy, f"Falta - {motivo_baja or 'Reemplazo'}")
                    subir_in_place(TABLA3_RUTA_GRAPH, wb)
                    session = get_session()
                    try:
                        session.add(CorreccionWeb(
                            dni=dni_vacante, fecha=hoy, comentario_entrada=f"Falta - {motivo_baja or 'Reemplazo'}",
                            registrado_por=current_user.email,
                        ))
                        session.commit()
                    finally:
                        session.close()
                finally:
                    liberar_lock("tabla3_web")
        except Exception:
            pass  # el reemplazo en si ya se guardo bien -- esto es solo prolijidad, no bloquea el flujo principal

        # Mismo motivo que headcount_nuevo_submit(): exportar_dimensiones.yml
        # primero para que el Excel Maestro refleje al reemplazo antes de
        # que el motor de clasificación lo lea.
        ok_disparo, _ = disparar_workflow("exportar_dimensiones.yml")
        detalle = "\n".join(log)
        if not ok_disparo:
            detalle += "\n\nEl reemplazo se guardó, pero no se pudo iniciar la actualización automática -- probá \"Actualizar ahora\" desde el reporte diario en unos minutos."
        titulo = "Reemplazo agregado"
    except Exception as e:
        detalle = str(e)
        titulo = "Falló"
        ok_disparo = False

    return render_template(
        "asistencia_resultado.html", usuario=current_user, activo="reemplazo",
        titulo=titulo, ok=(titulo == "Reemplazo agregado"), detalle=detalle,
        volver=url_for("asistencia.reemplazo_form"),
        poll_workflow="exportar_dimensiones.yml" if ok_disparo else None,
        poll_siguiente="pipeline_completo.yml" if ok_disparo else None,
    )


# Sin domingo -- mismo criterio que WD_NORM en horas_semanales.py, ningún
# módulo del pipeline espera un día de semana "Domingo" en PatronRecurrente.
DIAS_PATRON = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]


def _hora_form(valor):
    if not valor:
        return None
    try:
        return dt.time.fromisoformat(valor)
    except ValueError:
        return None


@bp.get("/headcount-nuevo")
@_analista_requerido
def headcount_nuevo_form():
    # Los desplegables se arman con los valores que YA existen en Postgres
    # (acotados al scope del usuario) en vez de una lista fija a mano --
    # mismo criterio que reportes.py::_filtros_admin() -- para no
    # desincronizarse si el Maestro usa un Rol/Canal nuevo que acá no
    # estaba contemplado.
    session = get_session()
    try:
        cond_scope = condicion_scope(Persona, current_user)

        def _valores(columna):
            q = session.query(columna).filter(columna.isnot(None))
            if cond_scope is not None:
                q = q.filter(cond_scope)
            return sorted({v for (v,) in q.distinct().all() if v})

        roles_disponibles = _valores(Persona.rol)
        canales_disponibles = _valores(Persona.canal)
        regiones_disponibles = _valores(Persona.region)
        ciudades_disponibles = _valores(Persona.ciudad)
        zonas_disponibles = _valores(Persona.zona)

        q_sup = session.query(Persona.dni, Persona.nombre_completo).filter(
            Persona.rol == "SUPERVISORES", Persona.estado == "Activo"
        )
        if cond_scope is not None:
            q_sup = q_sup.filter(cond_scope)
        supervisores_disponibles = sorted(q_sup.distinct().all(), key=lambda t: (t[1] or "").title())
    finally:
        session.close()

    return render_template(
        "asistencia_headcount_nuevo.html", usuario=current_user, activo="headcount_nuevo",
        hoy=dt.date.today().isoformat(), dias_patron=DIAS_PATRON,
        roles_disponibles=roles_disponibles, canales_disponibles=canales_disponibles,
        regiones_disponibles=regiones_disponibles, ciudades_disponibles=ciudades_disponibles,
        zonas_disponibles=zonas_disponibles, supervisores_disponibles=supervisores_disponibles,
    )


@bp.post("/headcount-nuevo")
@_analista_requerido
def headcount_nuevo_submit():
    from cargas import crear_persona_individual

    dni = request.form.get("dni", "").strip()
    nombre = request.form.get("nombre", "").strip()
    rol = request.form.get("rol", "").strip()
    canal = request.form.get("canal", "").strip() or None
    region = request.form.get("region", "").strip() or None
    ciudad = request.form.get("ciudad", "").strip() or None
    zona = request.form.get("zona", "").strip() or None
    supervisor_dni = request.form.get("supervisor_dni", "").strip() or None
    correo = request.form.get("correo", "").strip() or None
    fecha_ingreso_str = request.form.get("fecha_ingreso", "").strip()

    if not dni or not nombre or not rol or not fecha_ingreso_str:
        return render_template(
            "asistencia_resultado.html", usuario=current_user, activo="headcount_nuevo",
            titulo="Falló", ok=False, detalle="Faltan campos obligatorios (DNI, Nombre, Rol, Fecha de ingreso).",
            volver=url_for("asistencia.headcount_nuevo_form"),
        )
    try:
        fecha_ingreso = dt.date.fromisoformat(fecha_ingreso_str)
    except ValueError:
        return render_template(
            "asistencia_resultado.html", usuario=current_user, activo="headcount_nuevo",
            titulo="Falló", ok=False, detalle="Fecha de ingreso inválida.",
            volver=url_for("asistencia.headcount_nuevo_form"),
        )

    patron_dias = [
        {
            "dia_semana": dia,
            "hora_entrada": _hora_form(request.form.get(f"entrada_{dia.lower()}", "").strip()),
            "hora_salida": _hora_form(request.form.get(f"salida_{dia.lower()}", "").strip()),
            "canal_dia": request.form.get(f"canal_{dia.lower()}", "").strip() or None,
            "refrigerio": request.form.get(f"refrigerio_{dia.lower()}", "").strip() or None,
        }
        for dia in DIAS_PATRON
    ]

    try:
        es_reingreso, n_patron = crear_persona_individual(
            current_user.email, dni, nombre, rol, canal, region, ciudad, zona,
            supervisor_dni, correo, fecha_ingreso, patron_dias,
        )
        detalle = f"{nombre.title()} (DNI {dni}) fue dado de alta{' como reingreso' if es_reingreso else ''}."
        if n_patron:
            detalle += f" Se guardaron {n_patron} días de horario."
        else:
            detalle += (
                " No se cargó ningún horario -- completalo en \"Cargar Headcount\" (Patrón) o en "
                "\"Historial de cambios\" antes de que empiece a trabajar, sino no se va a poder "
                "clasificar su asistencia."
            )
        # exportar_dimensiones.yml primero, NO pipeline_completo.yml directo
        # -- el motor de clasificación lee el Excel Maestro, que recién
        # queda al día después de ese primer paso (ver poll_siguiente en
        # asistencia_resultado.html; Davor, 2026-08-26: headcount nuevo no
        # aparecía en el reporte por dispararse el paso equivocado).
        ok_disparo, _ = disparar_workflow("exportar_dimensiones.yml")
        if not ok_disparo:
            detalle += "\n\nNo se pudo iniciar la actualización automática -- probá \"Actualizar ahora\" desde el reporte diario en unos minutos."
        titulo = "Headcount agregado"
    except Exception as e:
        detalle = str(e)
        titulo = "Falló"
        ok_disparo = False

    return render_template(
        "asistencia_resultado.html", usuario=current_user, activo="headcount_nuevo",
        titulo=titulo, ok=(titulo == "Headcount agregado"), detalle=detalle,
        volver=url_for("asistencia.headcount_nuevo_form"),
        poll_workflow="exportar_dimensiones.yml" if titulo == "Headcount agregado" and ok_disparo else None,
        poll_siguiente="pipeline_completo.yml" if titulo == "Headcount agregado" and ok_disparo else None,
    )


@bp.get("/dar-de-baja")
@_analista_requerido
def dar_de_baja_form():
    """Contraparte de "Agregar headcount" -- mercaderista es una posición
    por días, así que hace falta poder dar de baja a alguien SIN
    necesariamente tener ya a quien lo reemplace (a diferencia de "Agregar
    reemplazo", que pide los dos DNIs a la vez) (Davor, 2026-08-26). Deja la
    posición en estado Vacante -- "Agregar reemplazo" ya sabe cubrir
    vacantes existentes cuando aparezca quien la tome."""
    session = get_session()
    try:
        cond_scope = condicion_scope(Persona, current_user)
        query = session.query(Persona.dni, Persona.nombre_completo, Persona.rol, Persona.ciudad).filter(
            Persona.estado == "Activo"
        )
        if cond_scope is not None:
            query = query.filter(cond_scope)
        activos = sorted(query.all(), key=lambda t: (t[1] or "").title())
    finally:
        session.close()

    return render_template(
        "asistencia_dar_de_baja.html", usuario=current_user, activo="dar_de_baja",
        hoy=dt.date.today().isoformat(), activos=activos,
    )


@bp.post("/dar-de-baja")
@_analista_requerido
def dar_de_baja_submit():
    dni = request.form.get("dni", "").strip()
    fecha_baja_str = request.form.get("fecha_baja", "").strip()
    motivo = request.form.get("motivo_baja", "").strip() or None

    if not dni or not fecha_baja_str:
        return render_template(
            "asistencia_resultado.html", usuario=current_user, activo="dar_de_baja",
            titulo="Falló", ok=False, detalle="Faltan campos obligatorios (persona, fecha de baja).",
            volver=url_for("asistencia.dar_de_baja_form"),
        )
    try:
        fecha_baja = dt.date.fromisoformat(fecha_baja_str)
    except ValueError:
        return render_template(
            "asistencia_resultado.html", usuario=current_user, activo="dar_de_baja",
            titulo="Falló", ok=False, detalle="Fecha de baja inválida.",
            volver=url_for("asistencia.dar_de_baja_form"),
        )

    session = get_session()
    try:
        # condicion_scope() aplicada ACÁ, no solo confiar en que el
        # desplegable del formulario ya venga acotado -- mismo criterio que
        # reportes.marcaciones()/reportes.ficha() para no confiar en un DNI
        # que llegue por POST directo sin pasar por el desplegable.
        query = session.query(Persona).filter(Persona.dni == dni)
        cond_scope = condicion_scope(Persona, current_user)
        if cond_scope is not None:
            query = query.filter(cond_scope)
        persona = query.first()

        if persona is None:
            titulo, ok = "Falló", False
            detalle = "No se encontró a esa persona o no tenés acceso a darla de baja."
        elif persona.estado != "Activo":
            titulo, ok = "Falló", False
            detalle = f"DNI {dni} no está Activo ahora mismo (está en '{persona.estado}')."
        else:
            nombre = persona.nombre_completo
            persona.estado = "Vacante"
            persona.fecha_baja = fecha_baja
            persona.motivo_baja = motivo
            persona.dado_de_baja_por = current_user.email
            session.commit()
            titulo, ok = "Baja registrada", True
            detalle = (
                f"{(nombre or dni).title()} (DNI {dni}) quedó de baja desde {fecha_baja_str}. "
                "La posición queda Vacante -- usá \"Agregar reemplazo\" cuando tengas a quien la cubra."
            )
    finally:
        session.close()

    ok_disparo = False
    if ok:
        # Misma cadena que headcount_nuevo_submit()/reemplazo_submit(): el
        # motor de clasificación lee el Excel Maestro, no Postgres directo.
        ok_disparo, _ = disparar_workflow("exportar_dimensiones.yml")
        if not ok_disparo:
            detalle += "\n\nNo se pudo iniciar la actualización automática -- probá \"Actualizar ahora\" desde el reporte diario en unos minutos."

    return render_template(
        "asistencia_resultado.html", usuario=current_user, activo="dar_de_baja",
        titulo=titulo, ok=ok, detalle=detalle,
        volver=url_for("asistencia.dar_de_baja_form"),
        poll_workflow="exportar_dimensiones.yml" if ok and ok_disparo else None,
        poll_siguiente="pipeline_completo.yml" if ok and ok_disparo else None,
    )
